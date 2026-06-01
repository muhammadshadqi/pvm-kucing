"""
Price Simulator Engine — Page 3 (Astro Pricing Toolkit)
Compute price scenario impact on GV, GP, Margin, and 3 versions of PI per SKU.

Input:
    - df_master: Query 3 output (per-SKU master data with 11 cols)
    - df_scenario: scenarios DataFrame (product_id + baseline + var_1, var_2, ...)

Output:
    - dict with: df_sim, summary, by_bl, by_l1, pi_distribution, framework_flags, meta
"""
import numpy as np
import pandas as pd
import io


# ─────────────────────────────────────────────────────────────────────────────
# CONSTANTS
# ─────────────────────────────────────────────────────────────────────────────
REQUIRED_MASTER_COLS = [
    'product_id', 'product_name', 'l1_category_name', 'l2_category_name',
    'pricing_bl_25', 'qty', 'selling_price', 'cost_price',
    'avg_comp_price', 'last_comp_price', 'last_price'
]

PI_BINS = [-np.inf, 95, 100, 105, 110, 120, np.inf]
PI_BINS_LBL = ['A.<95', 'B.95-<100', 'C.100-105', 'D.105-110', 'E.110-120', 'F.>120']


# ─────────────────────────────────────────────────────────────────────────────
# COMPUTE — main entry point
# ─────────────────────────────────────────────────────────────────────────────
def compute(df_master, df_scenario):
    """
    Compute price scenario simulation results.

    Args:
        df_master: pandas DataFrame with REQUIRED_MASTER_COLS
        df_scenario: pandas DataFrame with [product_id, baseline, var_1, var_2, ...]

    Returns:
        dict with all simulation outputs
    """
    # Sanity checks
    missing_m = [c for c in REQUIRED_MASTER_COLS if c not in df_master.columns]
    if missing_m:
        raise ValueError(f"File 1 missing columns: {missing_m}")
    if 'product_id' not in df_scenario.columns or 'baseline' not in df_scenario.columns:
        raise ValueError("File 2 must have at least: product_id, baseline")

    # Variant columns (everything except product_id + baseline)
    variant_cols = [c for c in df_scenario.columns
                    if c not in ('product_id', 'baseline')]

    # Force product_id to string for safe join
    df_master = df_master.copy()
    df_scenario = df_scenario.copy()
    df_master['product_id'] = df_master['product_id'].astype(str).str.strip()
    df_scenario['product_id'] = df_scenario['product_id'].astype(str).str.strip()

    # Inner join: only simulate SKU that exist in both
    df = df_master.merge(df_scenario, on='product_id', how='inner', suffixes=('', '_dup'))
    # Drop duplicate columns from merge
    df = df.loc[:, ~df.columns.str.endswith('_dup')]

    # All scenario columns
    scenarios = ['baseline'] + variant_cols

    # ── Per-scenario derived metrics ──
    # Constants per SKU
    df['cogs_total'] = df['qty'] * df['cost_price']
    df['act_gv']     = df['qty'] * df['selling_price']  # actual GV (current selling)
    df['act_gp']     = df['act_gv'] - df['cogs_total']
    df['act_gp_pct'] = np.where(df['act_gv'] != 0,
                                df['act_gp'] / df['act_gv'], np.nan)
    df['act_pi_avg']        = np.where(df['avg_comp_price'] > 0,
                                       df['selling_price'] * 100 / df['avg_comp_price'], np.nan)
    df['act_pi_last']       = np.where(df['last_comp_price'] > 0,
                                       df['selling_price'] * 100 / df['last_comp_price'], np.nan)
    df['pi_last_with_lp']   = np.where(df['last_comp_price'] > 0,
                                       df['last_price'] * 100 / df['last_comp_price'], np.nan)

    # Per scenario: price, GV, GP, GP%, PI Avg, PI Last
    for s in scenarios:
        price_col = s  # already a column in df from merge
        df[f'gv_{s}']         = df['qty'] * df[price_col]
        df[f'gp_{s}']         = df[f'gv_{s}'] - df['cogs_total']
        df[f'gp_pct_{s}']     = np.where(df[f'gv_{s}'] != 0,
                                          df[f'gp_{s}'] / df[f'gv_{s}'], np.nan)
        df[f'pi_avg_{s}']     = np.where(df['avg_comp_price'] > 0,
                                          df[price_col] * 100 / df['avg_comp_price'], np.nan)
        df[f'pi_last_{s}']    = np.where(df['last_comp_price'] > 0,
                                          df[price_col] * 100 / df['last_comp_price'], np.nan)

    # ── Delta vs baseline ──
    for v in variant_cols:
        df[f'd_price_{v}']      = df[v] - df['baseline']
        df[f'd_price_pct_{v}']  = np.where(df['baseline'] != 0,
                                            (df[v] - df['baseline']) / df['baseline'], np.nan)
        df[f'd_gv_{v}']         = df[f'gv_{v}'] - df['gv_baseline']
        df[f'd_gv_pct_{v}']     = np.where(df['gv_baseline'] != 0,
                                            (df[f'gv_{v}'] - df['gv_baseline']) / df['gv_baseline'], np.nan)
        df[f'd_gp_{v}']         = df[f'gp_{v}'] - df['gp_baseline']
        df[f'd_gp_pct_{v}']     = np.where(df['gp_baseline'] != 0,
                                            (df[f'gp_{v}'] - df['gp_baseline']) / df['gp_baseline'], np.nan)
        df[f'd_gp_pp_{v}']      = df[f'gp_pct_{v}'] - df['gp_pct_baseline']
        df[f'd_pi_avg_{v}']     = df[f'pi_avg_{v}'] - df['pi_avg_baseline']
        df[f'd_pi_last_{v}']    = df[f'pi_last_{v}'] - df['pi_last_baseline']

    # ── PI bucket per scenario (using PI Avg Comp as basis) ──
    for s in scenarios:
        df[f'pi_bucket_{s}'] = pd.cut(
            df[f'pi_avg_{s}'], bins=PI_BINS, labels=PI_BINS_LBL, include_lowest=False
        ).astype(str)

    # ── Aggregated summary ──
    summary = build_summary(df, scenarios, variant_cols)

    # ── Per-dimension aggregation ──
    by_bl = aggregate_by_dim(df, 'pricing_bl_25', scenarios, variant_cols)
    by_l1 = aggregate_by_dim(df, 'l1_category_name', scenarios, variant_cols)
    by_l2 = aggregate_by_dim(df, 'l2_category_name', scenarios, variant_cols)

    # ── PI distribution per scenario ──
    pi_distribution = build_pi_distribution(df, scenarios)

    # ── Framework flags per scenario ──
    framework_flags = build_framework_flags(df, scenarios, variant_cols)

    meta = {
        'n_sku_master':    len(df_master),
        'n_sku_scenario':  len(df_scenario),
        'n_sku_simulated': len(df),
        'n_sku_dropped':   len(df_scenario) - len(df),
        'variants':        variant_cols,
        'n_variants':      len(variant_cols),
    }

    return {
        'df_sim':          df,
        'summary':         summary,
        'by_bl':           by_bl,
        'by_l1':           by_l1,
        'by_l2':           by_l2,
        'pi_distribution': pi_distribution,
        'framework_flags': framework_flags,
        'scenarios':       scenarios,
        'variant_cols':    variant_cols,
        'meta':            meta,
    }


# ─────────────────────────────────────────────────────────────────────────────
# OVERALL IMPACT — full Astro universe (File 1 = all sales in range)
# ─────────────────────────────────────────────────────────────────────────────
def compute_overall_impact(df_master, df_scenario):
    """
    Impact of repricing the selected SKU on the FULL universe (all of File 1).

    Logic:
      - Changed SKU (in df_scenario)  -> price = scenario price (baseline / var_N)
      - Unchanged SKU (rest of File1) -> price = original selling_price (constant
        across all scenarios)
      - Overall / per-BL / per-L1 = aggregate of EVERYONE, but only changed SKU move.

    Returns dict with:
      overall : DataFrame per scenario (gv, gp, gp_pct, n_sku_total, n_sku_changed)
      by_bl   : DataFrame per pricing_bl_25 per scenario + deltas
      by_l1   : DataFrame per l1_category_name per scenario + deltas
      variant_cols, scenarios
    """
    df_master = df_master.copy()
    df_scenario = df_scenario.copy()
    df_master['product_id'] = df_master['product_id'].astype(str).str.strip()
    df_scenario['product_id'] = df_scenario['product_id'].astype(str).str.strip()

    variant_cols = [c for c in df_scenario.columns if c not in ('product_id', 'baseline')]
    scenarios = ['baseline'] + variant_cols

    # Map: product_id -> {scenario: price} for changed SKU only
    scn_map = df_scenario.set_index('product_id')

    u = df_master.copy()
    u['cogs_total'] = u['qty'] * u['cost_price']
    u['is_changed'] = u['product_id'].isin(scn_map.index)

    # Effective price per scenario for EVERY SKU
    for s in scenarios:
        # default = original selling_price (unchanged SKU stay flat across scenarios)
        eff = u['selling_price'].copy()
        if s in scn_map.columns or s == 'baseline':
            col = 'baseline' if s == 'baseline' else s
            # changed SKU -> scenario price (fall back to selling_price if NaN)
            mapped = u['product_id'].map(scn_map[col]) if col in scn_map.columns else pd.Series(np.nan, index=u.index)
            eff = np.where(u['is_changed'] & mapped.notna(), mapped, u['selling_price'])
        u[f'gv_{s}'] = u['qty'] * eff
        u[f'gp_{s}'] = u[f'gv_{s}'] - u['cogs_total']

    def _agg(g):
        row = {'n_sku_total': len(g), 'n_sku_changed': int(g['is_changed'].sum()),
               'qty': g['qty'].sum(), 'cogs': g['cogs_total'].sum()}
        for s in scenarios:
            gv = g[f'gv_{s}'].sum(); gp = g[f'gp_{s}'].sum()
            row[f'gv_{s}'] = gv
            row[f'gp_{s}'] = gp
            row[f'gp_pct_{s}'] = gp / gv if gv > 0 else np.nan
        for v in variant_cols:
            row[f'd_gv_{v}']     = row[f'gv_{v}'] - row['gv_baseline']
            row[f'd_gp_{v}']     = row[f'gp_{v}'] - row['gp_baseline']
            row[f'd_gp_pp_{v}']  = row[f'gp_pct_{v}'] - row['gp_pct_baseline']
        return row

    # Overall (single row table per scenario, transposed-friendly)
    overall = pd.DataFrame([{'scenario': s,
                             'gv': u[f'gv_{s}'].sum(),
                             'gp': u[f'gp_{s}'].sum(),
                             'gp_pct': (u[f'gp_{s}'].sum() / u[f'gv_{s}'].sum()
                                        if u[f'gv_{s}'].sum() > 0 else np.nan),
                             'n_sku_total': len(u),
                             'n_sku_changed': int(u['is_changed'].sum())}
                            for s in scenarios])

    def _by_dim(dim):
        if dim not in u.columns:
            return pd.DataFrame()
        rows = []
        for dim_val, g in u.groupby(dim, dropna=False):
            if pd.isna(dim_val):
                continue
            r = {dim: dim_val, **_agg(g)}
            rows.append(r)
        out = pd.DataFrame(rows)
        if not out.empty:
            out = out.sort_values('gp_baseline', ascending=False)
        return out

    return {
        'overall':      overall,
        'by_bl':        _by_dim('pricing_bl_25'),
        'by_l1':        _by_dim('l1_category_name'),
        'scenarios':    scenarios,
        'variant_cols': variant_cols,
    }


# ─────────────────────────────────────────────────────────────────────────────
# AGGREGATIONS
# ─────────────────────────────────────────────────────────────────────────────
def build_summary(df, scenarios, variant_cols):
    """Aggregated KPI per scenario (total across all simulated SKU)."""
    rows = []
    for s in scenarios:
        gv_total = df[f'gv_{s}'].sum()
        gp_total = df[f'gp_{s}'].sum()
        cogs_total = df['cogs_total'].sum()
        gp_pct = gp_total / gv_total if gv_total > 0 else np.nan

        # Weighted avg PI (weighted by qty)
        qty_total = df['qty'].sum()
        if qty_total > 0:
            pi_avg_w  = (df[f'pi_avg_{s}'] * df['qty']).sum() / qty_total
            pi_last_w = (df[f'pi_last_{s}'] * df['qty']).sum() / qty_total
        else:
            pi_avg_w = pi_last_w = np.nan

        rows.append({
            'scenario':    s,
            'gv':          gv_total,
            'cogs':        cogs_total,
            'gp':          gp_total,
            'gp_pct':      gp_pct,
            'pi_avg_w':    pi_avg_w,
            'pi_last_w':   pi_last_w,
            'n_sku':       len(df),
            'qty':         qty_total,
        })

    return pd.DataFrame(rows)


def aggregate_by_dim(df, dim_col, scenarios, variant_cols):
    """Aggregate per dimension (BL/L1/L2)."""
    if dim_col not in df.columns:
        return pd.DataFrame()

    grouped = df.groupby(dim_col, dropna=False)

    out_rows = []
    for dim_val, g in grouped:
        if pd.isna(dim_val):
            continue
        row = {dim_col: dim_val, 'n_sku': len(g), 'qty': g['qty'].sum()}
        cogs = g['cogs_total'].sum()
        row['cogs'] = cogs
        for s in scenarios:
            gv = g[f'gv_{s}'].sum()
            gp = g[f'gp_{s}'].sum()
            row[f'gv_{s}']     = gv
            row[f'gp_{s}']     = gp
            row[f'gp_pct_{s}'] = gp / gv if gv > 0 else np.nan
            # Weighted avg PI
            q = g['qty'].sum()
            if q > 0:
                row[f'pi_avg_{s}']  = (g[f'pi_avg_{s}'] * g['qty']).sum() / q
                row[f'pi_last_{s}'] = (g[f'pi_last_{s}'] * g['qty']).sum() / q
            else:
                row[f'pi_avg_{s}']  = np.nan
                row[f'pi_last_{s}'] = np.nan

        for v in variant_cols:
            row[f'd_gp_{v}'] = row[f'gp_{v}'] - row['gp_baseline']
            row[f'd_gp_pct_{v}'] = ((row[f'gp_{v}'] - row['gp_baseline']) / row['gp_baseline']
                                    if row['gp_baseline'] != 0 else np.nan)
            row[f'd_pi_avg_{v}']  = row[f'pi_avg_{v}']  - row['pi_avg_baseline']
            row[f'd_pi_last_{v}'] = row[f'pi_last_{v}'] - row['pi_last_baseline']

        out_rows.append(row)

    out_df = pd.DataFrame(out_rows)
    if not out_df.empty and 'gp_baseline' in out_df.columns:
        out_df = out_df.sort_values('gp_baseline', ascending=False)
    return out_df


def build_pi_distribution(df, scenarios):
    """SKU count per PI bucket per scenario."""
    rows = []
    for s in scenarios:
        counts = df[f'pi_bucket_{s}'].value_counts().reindex(PI_BINS_LBL, fill_value=0)
        row = {'scenario': s}
        for bucket in PI_BINS_LBL:
            row[bucket] = int(counts.get(bucket, 0))
        row['total'] = counts.sum()
        rows.append(row)
    return pd.DataFrame(rows)


def build_framework_flags(df, scenarios, variant_cols):
    """Apply framework rules per scenario, count flagged SKU."""
    # Rules (same as Page 2):
    # Rule 1: Fresh, PI > 110, Margin <= 15% -> drop price
    # Rule 2: Frozen, PI > 100, Margin <= 15% -> drop price
    # Rule 3: Fresh, PI > 120, Margin >= 70% -> over-priced
    # Rule 4: Dry, PI < 105, Margin <= 0% -> loss leader
    # Rule 5: Dry, PI > 120, Margin > 40% -> over-priced

    rows = []
    for s in scenarios:
        pi = df[f'pi_avg_{s}']
        margin = df[f'gp_pct_{s}']
        bl = df['pricing_bl_25']

        rule1 = (bl == 'Fresh') & (pi > 110) & (margin <= 0.15)
        rule2 = (bl == 'Frozen') & (pi > 100) & (margin <= 0.15)
        rule3 = (bl == 'Fresh') & (pi > 120) & (margin >= 0.70)
        rule4 = (bl == 'Dry') & (pi < 105) & (margin <= 0.00)
        rule5 = (bl == 'Dry') & (pi > 120) & (margin > 0.40)

        rows.append({
            'scenario':            s,
            'rule_1_fresh_dropp':  int(rule1.sum()),
            'rule_2_frozen_dropp': int(rule2.sum()),
            'rule_3_fresh_over':   int(rule3.sum()),
            'rule_4_dry_loss':     int(rule4.sum()),
            'rule_5_dry_over':     int(rule5.sum()),
            'total_flagged':       int((rule1 | rule2 | rule3 | rule4 | rule5).sum()),
        })
    return pd.DataFrame(rows)


# ─────────────────────────────────────────────────────────────────────────────
# EXCEL GENERATOR
# ─────────────────────────────────────────────────────────────────────────────
def generate_excel(result):
    """Generate multi-sheet Excel from compute() result."""
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # ── styles ──
    NAVY = "1B2A4A"
    HEADER_SUB = "E8EDF5"
    BASELINE_BG = "EBF5FB"
    VAR_BG = "EAFAF1"
    DELTA_BG = "FEF9E7"
    WHITE = "FFFFFF"

    NUM0 = '#,##0;(#,##0);"-"'
    NUM2 = '#,##0.00;(#,##0.00);"-"'
    PCT  = '0.0%;(0.0%);"-"'
    DPCT = '+0.0%;-0.0%;"0.0%"'
    DNUM = '+#,##0;-#,##0;"0"'
    DPP  = '+0.0"pp";-0.0"pp";"0pp"'
    PI_FMT = '0.000'

    thin = Side(style="thin", color="D5D8DC")
    bdr = Border(left=thin, right=thin, top=thin, bottom=thin)

    def hcell(ws, r, c, val, bg=NAVY, fg=WHITE, wrap=False):
        cell = ws.cell(r, c, val)
        cell.fill = PatternFill("solid", fgColor=bg)
        cell.font = Font(bold=True, color=fg, size=9)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=wrap)
        cell.border = bdr
        return cell

    def dcell(ws, r, c, val, bg=WHITE, fmt=None, halign="right", bold=False):
        cell = ws.cell(r, c, val)
        cell.fill = PatternFill("solid", fgColor=bg)
        cell.font = Font(bold=bold, size=9)
        cell.alignment = Alignment(horizontal=halign, vertical="center")
        cell.border = bdr
        if fmt:
            cell.number_format = fmt
        return cell

    df = result['df_sim']
    scenarios = result['scenarios']
    variant_cols = result['variant_cols']
    summary = result['summary']
    by_bl = result['by_bl']
    by_l1 = result['by_l1']
    pi_dist = result['pi_distribution']
    flags = result['framework_flags']

    wb = Workbook()

    # ── Sheet 1: Master Data ──
    ws1 = wb.active
    ws1.title = "1. Master Data"
    cols_s1 = ['product_id', 'product_name', 'l1_category_name', 'l2_category_name',
               'pricing_bl_25', 'qty', 'selling_price', 'cost_price',
               'avg_comp_price', 'last_comp_price', 'last_price',
               'act_gv', 'cogs_total', 'act_gp', 'act_gp_pct',
               'act_pi_avg', 'act_pi_last', 'pi_last_with_lp']
    cols_s1 = [c for c in cols_s1 if c in df.columns]

    for ci, c in enumerate(cols_s1, 1):
        hcell(ws1, 1, ci, c.replace('_', ' ').title())
        ws1.column_dimensions[get_column_letter(ci)].width = 16
    for ri, row in enumerate(df[cols_s1].itertuples(index=False), 2):
        for ci, c in enumerate(cols_s1, 1):
            val = getattr(row, c) if hasattr(row, c) else None
            fmt = NUM0 if c in ('qty','selling_price','cost_price','avg_comp_price',
                                'last_comp_price','last_price','act_gv','cogs_total','act_gp') else (
                PCT if c == 'act_gp_pct' else (
                PI_FMT if c in ('act_pi_avg','act_pi_last','pi_last_with_lp') else None))
            halign = "left" if c in ('product_name','l1_category_name','l2_category_name',
                                      'pricing_bl_25','product_id') else "right"
            dcell(ws1, ri, ci, val, fmt=fmt, halign=halign)
    ws1.freeze_panes = "B2"
    ws1.column_dimensions['B'].width = 32

    # ── Sheet 2: SKU Detail (wide pivot) ──
    ws2 = wb.create_sheet("2. SKU Detail")
    write_sku_detail(ws2, df, scenarios, variant_cols, hcell, dcell, NUM0, PCT, DNUM, DPCT, DPP, PI_FMT,
                     BASELINE_BG, VAR_BG, DELTA_BG, get_column_letter)

    # ── Sheet 3: Summary KPI ──
    ws3 = wb.create_sheet("3. Summary KPI")
    write_summary_kpi(ws3, summary, scenarios, variant_cols,
                      hcell, dcell, NUM0, PCT, PI_FMT, DNUM, DPCT, DPP,
                      BASELINE_BG, VAR_BG, DELTA_BG, get_column_letter)

    # ── Sheet 4: By BL ──
    ws4 = wb.create_sheet("4. By BL")
    write_by_dim(ws4, by_bl, 'pricing_bl_25', scenarios, variant_cols,
                 hcell, dcell, NUM0, PCT, PI_FMT, DNUM, DPCT,
                 BASELINE_BG, VAR_BG, DELTA_BG, get_column_letter)

    # ── Sheet 5: By L1 ──
    ws5 = wb.create_sheet("5. By L1")
    write_by_dim(ws5, by_l1, 'l1_category_name', scenarios, variant_cols,
                 hcell, dcell, NUM0, PCT, PI_FMT, DNUM, DPCT,
                 BASELINE_BG, VAR_BG, DELTA_BG, get_column_letter)

    # ── Sheet 6: PI Distribution ──
    ws6 = wb.create_sheet("6. PI Distribution")
    write_pi_dist(ws6, pi_dist, hcell, dcell, NUM0, get_column_letter)

    # ── Sheet 7: Framework Flags ──
    ws7 = wb.create_sheet("7. Framework Flags")
    write_framework_flags(ws7, flags, hcell, dcell, NUM0, get_column_letter)

    # ── Sheet 8: Glossary ──
    ws8 = wb.create_sheet("8. Glossary")
    write_glossary(ws8, hcell, dcell)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def write_sku_detail(ws, df, scenarios, variant_cols,
                     hcell, dcell, NUM0, PCT, DNUM, DPCT, DPP, PI_FMT,
                     BL_BG, V_BG, D_BG, gcl):
    """Wide pivot: per-SKU with all scenarios + deltas."""
    INFO = ['product_id', 'product_name', 'l1_category_name', 'l2_category_name',
            'pricing_bl_25', 'qty', 'cost_price', 'avg_comp_price', 'last_comp_price', 'last_price']

    # Build column structure
    structure = []  # list of (group_label, [col_keys])
    structure.append(("Info SKU", INFO))

    structure.append(("Harga", scenarios))
    structure.append(("GV", [f"gv_{s}" for s in scenarios]))
    structure.append(("COGS", ['cogs_total']))
    structure.append(("GP", [f"gp_{s}" for s in scenarios]))
    structure.append(("GP%", [f"gp_pct_{s}" for s in scenarios]))
    structure.append(("PI Avg Comp", [f"pi_avg_{s}" for s in scenarios]))
    structure.append(("PI Last Day", [f"pi_last_{s}" for s in scenarios]))
    structure.append(("PI Last w/ Last Price", ['pi_last_with_lp']))

    if variant_cols:
        structure.append(("Δ Harga", [f"d_price_{v}" for v in variant_cols]))
        structure.append(("Δ% Harga", [f"d_price_pct_{v}" for v in variant_cols]))
        structure.append(("Δ GV", [f"d_gv_{v}" for v in variant_cols]))
        structure.append(("Δ% GV", [f"d_gv_pct_{v}" for v in variant_cols]))
        structure.append(("Δ GP", [f"d_gp_{v}" for v in variant_cols]))
        structure.append(("Δ% GP", [f"d_gp_pct_{v}" for v in variant_cols]))
        structure.append(("Δ GP%", [f"d_gp_pp_{v}" for v in variant_cols]))
        structure.append(("Δ PI Avg", [f"d_pi_avg_{v}" for v in variant_cols]))
        structure.append(("Δ PI Last", [f"d_pi_last_{v}" for v in variant_cols]))

    # Flatten + write headers
    col_idx = 1
    group_spans = []  # (group_label, start_col, end_col)
    sub_labels = []   # list of (col_idx, label, group_label)
    for group, keys in structure:
        start = col_idx
        for k in keys:
            sub_labels.append((col_idx, k, group))
            col_idx += 1
        end = col_idx - 1
        group_spans.append((group, start, end))

    # Row 1: group header
    for group, start, end in group_spans:
        hcell(ws, 1, start, group)
        if end > start:
            ws.merge_cells(start_row=1, start_column=start, end_row=1, end_column=end)
    ws.row_dimensions[1].height = 22

    # Row 2: sub-header (cleaned label)
    def clean_lbl(k, group):
        if group == "Info SKU":
            return k.replace('_', ' ').title()
        if group in ("GV", "GP", "GP%", "PI Avg Comp", "PI Last Day", "Harga"):
            # extract scenario suffix
            return k.split('_')[-1] if '_' in k else k
        if group.startswith('Δ'):
            # e.g. "d_gp_var_1" -> "var_1"
            return k.split('_', 2)[-1] if k.count('_') >= 2 else k
        return k.replace('_', ' ')

    for ci, k, grp in sub_labels:
        lbl = clean_lbl(k, grp)
        hcell(ws, 2, ci, lbl, bg=HEADER_SUB if False else "E8EDF5", fg="1B2A4A")
    ws.row_dimensions[2].height = 20

    # Data rows
    DATA_START = 3
    for ri, row in enumerate(df.itertuples(index=False), DATA_START):
        for ci, k, grp in sub_labels:
            val = getattr(row, k, None) if hasattr(row, k) else None
            # Determine format & color
            if grp == "Info SKU":
                if k in ('qty', 'cost_price', 'avg_comp_price', 'last_comp_price', 'last_price'):
                    fmt = NUM0
                    halign = "right"
                else:
                    fmt = None
                    halign = "left"
                bg = WHITE_HEX
            elif grp == "Harga":
                fmt = NUM0
                halign = "right"
                bg = BL_BG if k == 'baseline' else V_BG
            elif grp in ("GV", "GP"):
                fmt = NUM0
                halign = "right"
                bg = BL_BG if k.endswith('baseline') else V_BG
            elif grp == "COGS":
                fmt = NUM0
                halign = "right"
                bg = WHITE_HEX
            elif grp == "GP%":
                fmt = PCT
                halign = "right"
                bg = BL_BG if k.endswith('baseline') else V_BG
            elif grp in ("PI Avg Comp", "PI Last Day"):
                fmt = PI_FMT
                halign = "right"
                bg = BL_BG if k.endswith('baseline') else V_BG
            elif grp == "PI Last w/ Last Price":
                fmt = PI_FMT
                halign = "right"
                bg = "F4F6F9"  # subtotal grey
            elif grp == "Δ Harga":
                fmt = DNUM
                halign = "right"
                bg = D_BG
            elif grp == "Δ% Harga":
                fmt = DPCT
                halign = "right"
                bg = D_BG
            elif grp == "Δ GV":
                fmt = DNUM
                halign = "right"
                bg = D_BG
            elif grp == "Δ% GV":
                fmt = DPCT
                halign = "right"
                bg = D_BG
            elif grp == "Δ GP":
                fmt = DNUM
                halign = "right"
                bg = D_BG
            elif grp == "Δ% GP":
                fmt = DPCT
                halign = "right"
                bg = D_BG
            elif grp == "Δ GP%":
                fmt = DPP
                halign = "right"
                bg = D_BG
            elif grp in ("Δ PI Avg", "Δ PI Last"):
                fmt = '+0.000;-0.000;"0.000"'
                halign = "right"
                bg = D_BG
            else:
                fmt = None
                halign = "right"
                bg = WHITE_HEX

            # NaN check
            if isinstance(val, float) and (val != val):  # NaN
                val = None
            dcell(ws, ri, ci, val, bg=bg, fmt=fmt, halign=halign)

    # Column widths
    for ci, k, grp in sub_labels:
        if grp == "Info SKU":
            ws.column_dimensions[gcl(ci)].width = 22 if k == 'product_name' else 14
        else:
            ws.column_dimensions[gcl(ci)].width = 13

    ws.freeze_panes = "B3"


def write_summary_kpi(ws, summary, scenarios, variant_cols,
                      hcell, dcell, NUM0, PCT, PI_FMT, DNUM, DPCT, DPP,
                      BL_BG, V_BG, D_BG, gcl):
    """Summary KPI: one row per scenario + delta rows."""
    cols = ['scenario', 'gv', 'cogs', 'gp', 'gp_pct', 'pi_avg_w', 'pi_last_w', 'n_sku', 'qty']
    labels = ['Scenario', 'GV', 'COGS', 'GP', 'GP%', 'PI Avg (W)', 'PI Last (W)', '# SKU', 'Qty Total']

    for ci, lbl in enumerate(labels, 1):
        hcell(ws, 1, ci, lbl)
        ws.column_dimensions[gcl(ci)].width = 18

    for ri, row in enumerate(summary.itertuples(index=False), 2):
        is_baseline = row.scenario == 'baseline'
        bg = BL_BG if is_baseline else V_BG
        dcell(ws, ri, 1, row.scenario, bg=bg, halign="left", bold=is_baseline)
        dcell(ws, ri, 2, row.gv,        bg=bg, fmt=NUM0)
        dcell(ws, ri, 3, row.cogs,      bg=bg, fmt=NUM0)
        dcell(ws, ri, 4, row.gp,        bg=bg, fmt=NUM0)
        dcell(ws, ri, 5, row.gp_pct,    bg=bg, fmt=PCT)
        dcell(ws, ri, 6, row.pi_avg_w,  bg=bg, fmt=PI_FMT)
        dcell(ws, ri, 7, row.pi_last_w, bg=bg, fmt=PI_FMT)
        dcell(ws, ri, 8, row.n_sku,     bg=bg, fmt=NUM0)
        dcell(ws, ri, 9, row.qty,       bg=bg, fmt=NUM0)

    # Delta rows (per variant vs baseline)
    if variant_cols:
        baseline_row = summary[summary['scenario'] == 'baseline'].iloc[0]
        delta_start = len(summary) + 3
        for ci, lbl in enumerate(['Δ Variant', 'Δ GV', '—', 'Δ GP', 'Δ GP% (pp)', 'Δ PI Avg', 'Δ PI Last', '—', '—'], 1):
            hcell(ws, delta_start - 1, ci, lbl, bg="5D4E8E")
        for ri_offset, v in enumerate(variant_cols, delta_start):
            v_row = summary[summary['scenario'] == v].iloc[0]
            dcell(ws, ri_offset, 1, f"Δ {v}", bg=D_BG, halign="left", bold=True)
            dcell(ws, ri_offset, 2, v_row.gv - baseline_row.gv,             bg=D_BG, fmt=DNUM)
            dcell(ws, ri_offset, 3, "—",                                     bg=D_BG, halign="center")
            dcell(ws, ri_offset, 4, v_row.gp - baseline_row.gp,             bg=D_BG, fmt=DNUM)
            dcell(ws, ri_offset, 5, (v_row.gp_pct - baseline_row.gp_pct),   bg=D_BG, fmt=DPP)
            dcell(ws, ri_offset, 6, v_row.pi_avg_w - baseline_row.pi_avg_w, bg=D_BG, fmt='+0.000;-0.000;"0.000"')
            dcell(ws, ri_offset, 7, v_row.pi_last_w - baseline_row.pi_last_w, bg=D_BG, fmt='+0.000;-0.000;"0.000"')
            dcell(ws, ri_offset, 8, "—", bg=D_BG, halign="center")
            dcell(ws, ri_offset, 9, "—", bg=D_BG, halign="center")


def write_by_dim(ws, by_dim, dim_col, scenarios, variant_cols,
                 hcell, dcell, NUM0, PCT, PI_FMT, DNUM, DPCT,
                 BL_BG, V_BG, D_BG, gcl):
    """Aggregated per dimension."""
    if by_dim is None or by_dim.empty:
        hcell(ws, 1, 1, "(no data)")
        return

    # Choose columns to display (keep it readable)
    display_cols = [dim_col, 'n_sku']
    for s in scenarios:
        display_cols += [f'gv_{s}', f'gp_{s}', f'gp_pct_{s}', f'pi_avg_{s}']
    for v in variant_cols:
        display_cols += [f'd_gp_{v}', f'd_pi_avg_{v}']

    display_cols = [c for c in display_cols if c in by_dim.columns]

    # Headers
    for ci, c in enumerate(display_cols, 1):
        bg = "1B2A4A" if not c.startswith('d_') else "5D4E8E"
        hcell(ws, 1, ci, c.replace('_', ' ').title(), bg=bg)
        ws.column_dimensions[gcl(ci)].width = 14

    # Rows
    for ri, row in enumerate(by_dim.itertuples(index=False), 2):
        for ci, c in enumerate(display_cols, 1):
            val = getattr(row, c, None) if hasattr(row, c) else None
            if c == dim_col:
                dcell(ws, ri, ci, val, halign="left")
            elif c.startswith('d_pi'):
                dcell(ws, ri, ci, val, bg=D_BG, fmt='+0.000;-0.000;"0.000"')
            elif c.startswith('d_'):
                dcell(ws, ri, ci, val, bg=D_BG, fmt=DNUM)
            elif 'gp_pct' in c:
                bg = BL_BG if c.endswith('baseline') else V_BG
                dcell(ws, ri, ci, val, bg=bg, fmt=PCT)
            elif 'pi_avg' in c or 'pi_last' in c:
                bg = BL_BG if c.endswith('baseline') else V_BG
                dcell(ws, ri, ci, val, bg=bg, fmt=PI_FMT)
            elif c == 'n_sku':
                dcell(ws, ri, ci, val, fmt=NUM0)
            else:
                bg = BL_BG if c.endswith('baseline') else V_BG
                dcell(ws, ri, ci, val, bg=bg, fmt=NUM0)


def write_pi_dist(ws, pi_dist, hcell, dcell, NUM0, gcl):
    """PI bucket distribution per scenario."""
    cols = list(pi_dist.columns)
    for ci, c in enumerate(cols, 1):
        hcell(ws, 1, ci, c)
        ws.column_dimensions[gcl(ci)].width = 13
    for ri, row in enumerate(pi_dist.itertuples(index=False), 2):
        for ci, c in enumerate(cols, 1):
            val = getattr(row, c, None) if hasattr(row, c) else None
            if c == 'scenario':
                dcell(ws, ri, ci, val, halign="left", bold=True)
            else:
                dcell(ws, ri, ci, val, fmt=NUM0)


def write_framework_flags(ws, flags, hcell, dcell, NUM0, gcl):
    """Framework flags per scenario."""
    cols = list(flags.columns)
    for ci, c in enumerate(cols, 1):
        hcell(ws, 1, ci, c.replace('_', ' ').title())
        ws.column_dimensions[gcl(ci)].width = 18
    for ri, row in enumerate(flags.itertuples(index=False), 2):
        for ci, c in enumerate(cols, 1):
            val = getattr(row, c, None) if hasattr(row, c) else None
            if c == 'scenario':
                dcell(ws, ri, ci, val, halign="left", bold=True)
            else:
                dcell(ws, ri, ci, val, fmt=NUM0)


def write_glossary(ws, hcell, dcell):
    """Glossary sheet."""
    hcell(ws, 1, 1, "Term", bg="1B2A4A")
    hcell(ws, 1, 2, "Formula", bg="1B2A4A")
    hcell(ws, 1, 3, "Description", bg="1B2A4A")
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 80

    rows = [
        ('GV (Goods Value)',     '= qty × selling_price',          'Revenue per SKU per scenario'),
        ('GP (Gross Profit)',    '= qty × (selling_price - cost_price)', 'Profit absolut'),
        ('COGS Total',           '= qty × cost_price',             'Cost of goods sold'),
        ('GP%',                  '= GP / GV',                       'Margin %'),
        ('PI Avg Comp',          '= selling_price × 100 / avg_comp_price',  'Posisi rata-rata vs blended avg comp price periode tsb'),
        ('PI Last Day',          '= selling_price × 100 / last_comp_price', 'Posisi vs blended comp price di hari terakhir periode'),
        ('PI Last w/ Last Price','= last_price × 100 / last_comp_price',    'Posisi current snapshot (constant, scenario-independent)'),
        ('Δ vs Baseline',        '= var_value - baseline_value',   'Selisih variant terhadap baseline scenario'),
        ('PI Bucket',            'A.<95 / B.95-100 / C.100-105 / D.105-110 / E.110-120 / F.>120', 'Klasifikasi posisi PI'),
        ('Framework Rules',      'Rule 1-5 lihat Glossary Page 5', 'Aturan SKU butuh repricing action'),
    ]
    for ri, (term, formula, desc) in enumerate(rows, 2):
        dcell(ws, ri, 1, term, halign="left", bold=True)
        dcell(ws, ri, 2, formula, halign="left")
        dcell(ws, ri, 3, desc, halign="left")


# Define WHITE_HEX outside for write_sku_detail
WHITE_HEX = "FFFFFF"
HEADER_SUB_HEX = "E8EDF5"
