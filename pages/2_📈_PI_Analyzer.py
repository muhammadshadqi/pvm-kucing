"""
PI Analyzer — Streamlit Web App (Page 2)
Pricing Index movement decomposition vs competitor.

Author: Shadqi (Pricing Strategy Analyst, Astro)
"""
import io
import sys
import os
import streamlit as st
import pandas as pd
import numpy as np
import plotly.graph_objects as go

# Add parent dir to path so we can import pi_analyzer_v1
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from pi_analyzer_v1 import compute as pi_compute, generate_excel as pi_generate_excel

# ─────────────────────────────────────────────────────────────────────────────
# PAGE CONFIG
# ─────────────────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="PI Analyzer — Astro Pricing",
    page_icon="📈",
    layout="wide",
    initial_sidebar_state="auto",
)


# ─────────────────────────────────────────────────────────────────────────────
# CACHED HELPERS (analytical computation cached by file hash)
# ─────────────────────────────────────────────────────────────────────────────
@st.cache_data(show_spinner=False, max_entries=3)
def cached_pi_compute(file_bytes, file_name):
    """Cache PI compute result by file content hash. Excel NOT included."""
    import io
    if file_name.lower().endswith('.csv'):
        df = pd.read_csv(io.BytesIO(file_bytes))
    else:
        df = pd.read_excel(io.BytesIO(file_bytes))
    return pi_compute(df)


@st.cache_data(show_spinner=False, max_entries=2)
def cached_pi_excel_bytes(file_bytes, file_name):
    """Cache Excel workbook bytes (separate cache from compute)."""
    import io
    result = cached_pi_compute(file_bytes, file_name)
    wb = pi_generate_excel(result)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

# Custom CSS — mirror Page 1
st.markdown("""
<style>
    .main > div { padding-top: 1rem; }
    .kpi-card {
        border: 1px solid #E5E7EB;
        border-radius: 10px;
        padding: 16px 20px;
        background: #FFFFFF;
        height: 100%;
    }
    .kpi-label { font-size: 12px; color: #6B7280; text-transform: uppercase; letter-spacing: 0.5px; font-weight: 600; }
    .kpi-value { font-size: 28px; font-weight: 700; color: #111827; margin: 4px 0; }
    .kpi-delta-pos { font-size: 14px; color: #059669; font-weight: 600; }
    .kpi-delta-neg { font-size: 14px; color: #DC2626; font-weight: 600; }
    .kpi-delta-neu { font-size: 14px; color: #6B7280; font-weight: 600; }
    .kpi-sub { font-size: 11px; color: #9CA3AF; margin-top: 4px; }
    .banner-warn {
        background: #FEF3C7;
        border-left: 4px solid #F59E0B;
        padding: 12px 16px;
        border-radius: 6px;
        margin-bottom: 16px;
    }
    .banner-info {
        background: #DBEAFE;
        border-left: 4px solid #2563EB;
        padding: 12px 16px;
        border-radius: 6px;
        margin-bottom: 16px;
    }
    .section-header {
        font-size: 22px;
        font-weight: 700;
        color: #111827;
        margin-top: 32px;
        margin-bottom: 12px;
        padding-bottom: 8px;
        border-bottom: 2px solid #E5E7EB;
    }
</style>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────────────────────
# CONSTANTS
# ─────────────────────────────────────────────────────────────────────────────
REQUIRED_COLS = [
    'product_id', 'product_name', 'l1_category_name', 'pricing_bl_25',
    'pareto_classification', 'source_status',
    'price', 'next_price', 'cogs', 'next_cogs',
    'comp_price', 'next_comp_price',
    'normal_comp_price', 'next_normal_comp_price',
    'pi', 'next_pi',
]
PERIOD_PAIRS = [('week_key', 'next_week'), ('month_key', 'next_month')]

SEGMENTS = ['Dry', 'Fresh', 'Frozen']

PI_BINS_LBL  = ["A.<95", "B.95-<100", "C.100-105", "D.105-110", "E.110-120", "F.>120"]
CI_BINS_LBL  = ["A.<70", "B.70-85", "C.85-95", "D.95-105", "E.>105"]
MG_BINS_LBL  = ["A.<-20%", "B.-20to-10%", "C.-10to0%", "D.0to10%",
                "E.10to20%", "F.20to30%", "G.30to50%", "H.>50%"]

# ─────────────────────────────────────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────────────────────────────────────
def fmt_pi(v, decimals=2):
    """Format PI value (just a number, no unit since PI is unitless ratio)."""
    if pd.isna(v): return "—"
    return f"{v:,.{decimals}f}"

def fmt_pi_delta(v, decimals=3):
    """Format PI delta with sign."""
    if pd.isna(v): return "—"
    return f"{v:+.{decimals}f}"

def fmt_pct(v, decimals=2):
    if pd.isna(v): return "—"
    return f"{v*100:.{decimals}f}%"

def fmt_count(v):
    if pd.isna(v): return "—"
    return f"{int(v):,}"

def kpi_card(label, value, delta=None, delta_label=None, sub=None, delta_inverse=False):
    """Build HTML for KPI card."""
    delta_html = ""
    if delta is not None and delta_label is not None:
        if delta == 0:
            cls = "kpi-delta-neu"
            arrow = "—"
        else:
            # delta_inverse=True means lower=better (e.g. PI lower vs competitor)
            is_positive = delta > 0
            if delta_inverse:
                cls = "kpi-delta-neg" if is_positive else "kpi-delta-pos"
            else:
                cls = "kpi-delta-pos" if is_positive else "kpi-delta-neg"
            arrow = "▲" if is_positive else "▼"
        delta_html = f'<div class="{cls}">{arrow} {delta_label}</div>'

    sub_html = f'<div class="kpi-sub">{sub}</div>' if sub else ""

    return f'''<div class="kpi-card">
<div class="kpi-label">{label}</div>
<div class="kpi-value">{value}</div>
{delta_html}
{sub_html}
</div>'''


def gradient_color(val, vmax):
    """Return CSS background+text color for gradient red-white-green."""
    if pd.isna(val): return ''
    if vmax == 0: return ''
    norm = max(-1, min(1, val / vmax))
    if norm > 0:
        alpha = norm
        r = int(255 - (255-22) * alpha)
        g = int(255 - (255-163) * alpha)
        b = int(255 - (255-74) * alpha)
        return f'background-color: rgb({r},{g},{b}); color: #111827;'
    elif norm < 0:
        alpha = -norm
        r = int(255 - (255-220) * alpha)
        g = int(255 - (255-38) * alpha)
        b = int(255 - (255-38) * alpha)
        text_color = '#FFFFFF' if alpha > 0.6 else '#111827'
        return f'background-color: rgb({r},{g},{b}); color: {text_color};'
    return ''


def make_template_pi_excel():
    """Generate a template Excel with required columns."""
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Raw Input Template"

    headers = ['week_key', 'next_week'] + REQUIRED_COLS
    for c, h in enumerate(headers, 1):
        ws.cell(1, c, h)

    # Add note
    ws.cell(3, 1, "Note: Required period cols: 'week_key' + 'next_week' OR 'month_key' + 'next_month'")
    ws.cell(4, 1, "All numeric cols (price, cogs, comp_price, normal_comp_price, pi) bisa NaN untuk New/Departing SKU.")

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


# ─────────────────────────────────────────────────────────────────────────────
# TORNADO CHART (Zone 3)
# ─────────────────────────────────────────────────────────────────────────────
def make_pi_tornado(decomp_dict, scope_label='Overall'):
    """
    Tornado horizontal bar chart of 5 PI effect components.
    Components: eff_dep (Churned), eff_price, eff_normal_comp, eff_discount_comp, eff_new.
    """
    components = [
        ('1. Churned SKU Effect', decomp_dict['eff_dep']),
        ('2. Price Change Effect',       decomp_dict['eff_price']),
        ('3.1 Normal Comp Price Effect', decomp_dict['eff_normal_comp']),
        ('3.2 Discount (Blended) Comp Price Effect', decomp_dict['eff_discount_comp']),
        ('4. New SKU Effect',     decomp_dict['eff_new']),
    ]
    # Sort by magnitude desc
    sorted_comp = sorted(components, key=lambda x: abs(x[1]), reverse=True)
    labels = [c[0] for c in sorted_comp][::-1]
    values = [c[1] for c in sorted_comp][::-1]
    colors = ['#059669' if v >= 0 else '#DC2626' for v in values]
    text_vals = [f"{v:+.3f}" for v in values]

    total = decomp_dict.get('total', sum(c[1] for c in components))

    fig = go.Figure(go.Bar(
        x=values,
        y=labels,
        orientation='h',
        marker_color=colors,
        text=text_vals,
        textposition='outside',
        cliponaxis=False,
    ))
    fig.update_layout(
        title=f"PI Movement Decomposition — {scope_label}  |  Total Δ: {total:+.3f} (A={decomp_dict['A']:.2f} → E={decomp_dict['E']:.2f})",
        showlegend=False,
        height=420,
        margin=dict(l=20, r=80, t=60, b=40),
        xaxis_title='PI points',
        plot_bgcolor='white',
        bargap=0.35,
    )
    fig.update_xaxes(showgrid=True, gridcolor='#F3F4F6', zeroline=True,
                     zerolinecolor='#9CA3AF', zerolinewidth=2)
    fig.update_yaxes(showgrid=False)
    return fig


# ─────────────────────────────────────────────────────────────────────────────
# BUILDER: Tabel PI Decomposition per Scope (Zone 3 Tabel)
# ─────────────────────────────────────────────────────────────────────────────
def build_pi_decomp_table(overall, segments, contribs):
    """
    Tabel 1: Per-segment PI decomposition with overall column.
    Methodology mirrors Sheet 2 in Excel output.
    Rows: A baseline, 5 effect components, E result, Total Δ.
    Cols: Overall, Dry, Fresh, Frozen.
    """
    rows = []
    rows.append({
        'Step': 'Baseline — Avg PI Prev (A)',
        'Overall': overall['A'],
        'Dry':     segments['Dry']['A'],
        'Fresh':   segments['Fresh']['A'],
        'Frozen':  segments['Frozen']['A'],
    })
    for label, key in [
        ('1. Churned SKU Effect',          'eff_dep'),
        ('2. Price Change Effect',         'eff_price'),
        ('3. Comp Price Effect',  'eff_comp'),
        ('    3.1 Normal Comp Price Effect',       'eff_normal_comp'),
        ('    3.2 Discount (Blended) Comp Price Effect', 'eff_discount_comp'),
        ('4. New SKU Effect',              'eff_new'),
    ]:
        rows.append({
            'Step': label,
            'Overall': overall[key],
            'Dry':     segments['Dry'][key],
            'Fresh':   segments['Fresh'][key],
            'Frozen':  segments['Frozen'][key],
        })
    rows.append({
        'Step': 'Total Δ',
        'Overall': overall['total'],
        'Dry':     segments['Dry']['total'],
        'Fresh':   segments['Fresh']['total'],
        'Frozen':  segments['Frozen']['total'],
    })
    rows.append({
        'Step': 'Result — Avg PI Cur (E)',
        'Overall': overall['E'],
        'Dry':     segments['Dry']['E'],
        'Fresh':   segments['Fresh']['E'],
        'Frozen':  segments['Frozen']['E'],
    })
    return pd.DataFrame(rows)


def build_pi_contrib_table(overall, contribs):
    """
    Tabel 2: Per-segment contribution to OVERALL effect (exact math identity).
    Each cell = exact_contributions formula. Sum across segments = overall (exact).
    Structure:
      - Row 1: Overall PI Prev (A) [baseline reference, top]
      - Rows 2-7: 6 effects (Churned, Price, Comp, Normal, Discount, New)
      - Row 8: Total Δ (sum of effects)
      - Row 9: Overall PI Cur (E) [result reference, bottom]
    """
    rows = []

    # Row 1: Baseline reference (top)
    rows.append({
        'Effect': 'Overall PI Prev (A)',
        'Dry contrib':    None,
        'Fresh contrib':  None,
        'Frozen contrib': None,
        'Overall (sum)':  None,
        'Overall actual': overall['A'],
    })

    # Rows 2-7: 6 effects
    for label, key in [
        ('1. Churned SKU Effect',                  'eff_dep'),
        ('2. Price Change Effect',                 'eff_price'),
        ('3. Comp Price Effect',                   'eff_comp'),
        ('  3.1 Normal Comp Price Effect',         'eff_normal_comp'),
        ('  3.2 Discount (Blended) Comp Price Effect', 'eff_discount_comp'),
        ('4. New SKU Effect',                      'eff_new'),
    ]:
        rows.append({
            'Effect': label,
            'Dry contrib':    contribs['Dry'][key],
            'Fresh contrib':  contribs['Fresh'][key],
            'Frozen contrib': contribs['Frozen'][key],
            'Overall (sum)':  contribs['Dry'][key] + contribs['Fresh'][key] + contribs['Frozen'][key],
            'Overall actual': overall[key],
        })

    # Row 8: Total Δ (sum of leaf effects: dep + price + comp + new). eff_comp already = normal+discount
    total_dry    = contribs['Dry']['eff_dep'] + contribs['Dry']['eff_price'] + contribs['Dry']['eff_comp'] + contribs['Dry']['eff_new']
    total_fresh  = contribs['Fresh']['eff_dep'] + contribs['Fresh']['eff_price'] + contribs['Fresh']['eff_comp'] + contribs['Fresh']['eff_new']
    total_frozen = contribs['Frozen']['eff_dep'] + contribs['Frozen']['eff_price'] + contribs['Frozen']['eff_comp'] + contribs['Frozen']['eff_new']
    rows.append({
        'Effect': 'Total Δ',
        'Dry contrib':    total_dry,
        'Fresh contrib':  total_fresh,
        'Frozen contrib': total_frozen,
        'Overall (sum)':  total_dry + total_fresh + total_frozen,
        'Overall actual': overall['total'],
    })

    # Row 9: Result reference (bottom)
    rows.append({
        'Effect': 'Overall PI Cur (E)',
        'Dry contrib':    None,
        'Fresh contrib':  None,
        'Frozen contrib': None,
        'Overall (sum)':  None,
        'Overall actual': overall['E'],
    })
    return pd.DataFrame(rows)


# ─────────────────────────────────────────────────────────────────────────────
# BUILDER: L1 Decomposition (Zone 4)
# ─────────────────────────────────────────────────────────────────────────────
def build_l1_pi_table(df, scope='Overall'):
    """
    For each L1 category, run decompose() and return table.
    scope: 'Overall' = all SKU, else filter by BL (Dry/Fresh/Frozen).
    Adds Margin Before/After + Driver Effect identification.
    """
    from pi_analyzer_v1 import decompose

    if scope != 'Overall':
        df_scope = df[df['pricing_bl_25'] == scope].copy()
    else:
        df_scope = df.copy()

    if 'l1_category_name' not in df_scope.columns or len(df_scope) == 0:
        return pd.DataFrame()

    # Effect-key → friendly name map (for driver identification)
    EFFECT_NAMES = {
        'eff_dep':          '1. Churned',
        'eff_price':        '2. Price',
        'eff_normal_comp':  '3.1 Normal Comp',
        'eff_discount_comp':'3.2 Discount Comp',
        'eff_new':          '4. New SKU',
    }

    rows = []
    for l1, g in df_scope.groupby('l1_category_name', dropna=False):
        if pd.isna(l1):
            continue
        r = decompose(g)
        ex_sub = g[g['sku_type'] == 'Existing']

        # Margin Existing P1 vs P2
        mg_p1 = ex_sub['margin_pct_prev'].mean() if 'margin_pct_prev' in ex_sub.columns and len(ex_sub) > 0 else np.nan
        mg_p2 = ex_sub['margin_pct_cur'].mean()  if 'margin_pct_cur'  in ex_sub.columns and len(ex_sub) > 0 else np.nan
        mg_delta = (mg_p2 - mg_p1) if (pd.notna(mg_p1) and pd.notna(mg_p2)) else np.nan

        # Driver effect — find which leaf component (non-aggregate) has the biggest magnitude
        # We use NORMAL & DISCOUNT comp separately (not aggregated eff_comp)
        leaf_effects = {
            'eff_dep':          r['eff_dep'],
            'eff_price':        r['eff_price'],
            'eff_normal_comp':  r['eff_normal_comp'],
            'eff_discount_comp':r['eff_discount_comp'],
            'eff_new':          r['eff_new'],
        }
        # Pick the effect with the same sign as Total Δ and largest magnitude
        total = r['total']
        if pd.notna(total) and total != 0:
            same_sign = {k: v for k, v in leaf_effects.items() if (v >= 0) == (total >= 0)}
            if same_sign:
                driver_key = max(same_sign, key=lambda k: abs(same_sign[k]))
                driver_name = EFFECT_NAMES[driver_key]
                driver_val  = leaf_effects[driver_key]
            else:
                driver_key = max(leaf_effects, key=lambda k: abs(leaf_effects[k]))
                driver_name = EFFECT_NAMES[driver_key]
                driver_val  = leaf_effects[driver_key]
        else:
            driver_name = '—'
            driver_val = 0

        rows.append({
            'L1 Category': l1,
            'BL (sample)': g['pricing_bl_25'].iloc[0] if 'pricing_bl_25' in g.columns else '—',
            'n Existing':  r['n_ex'],
            'n Departing': r['n_dep'],
            'n New':       r['n_new'],
            'Avg PI Prev (A)': r['A'],
            'Avg PI Cur (E)':  r['E'],
            'Total Δ':         r['total'],
            'Driver':          driver_name,
            'Driver Value':    driver_val,
            'Margin % P1': mg_p1,
            'Margin % P2': mg_p2,
            'Margin Δ pp': mg_delta,
            '1. Churned Eff':  r['eff_dep'],
            '2. Price Eff':    r['eff_price'],
            '3. Comp Price Eff':     r['eff_comp'],
            '3.1 Normal Comp Eff': r['eff_normal_comp'],
            '3.2 Discount Comp Eff': r['eff_discount_comp'],
            '4. New SKU Eff':      r['eff_new'],
        })

    out = pd.DataFrame(rows)
    if not out.empty:
        out = out.sort_values('Total Δ', ascending=False).reset_index(drop=True)
    return out


# ─────────────────────────────────────────────────────────────────────────────
# BUILDER: Pareto Decomposition (Zone 11)
# ─────────────────────────────────────────────────────────────────────────────
def build_pareto_pi_table(df, scope='Overall'):
    """Same structure as L1 but dimension = pareto_classification."""
    from pi_analyzer_v1 import decompose

    if scope != 'Overall':
        df_scope = df[df['pricing_bl_25'] == scope].copy()
    else:
        df_scope = df.copy()

    if 'pareto_classification' not in df_scope.columns or len(df_scope) == 0:
        return pd.DataFrame()

    rows = []
    for pareto, g in df_scope.groupby('pareto_classification', dropna=False):
        if pd.isna(pareto):
            continue
        r = decompose(g)
        rows.append({
            'Pareto Class':    pareto,
            'n Existing':      r['n_ex'],
            'n Departing':     r['n_dep'],
            'n New':           r['n_new'],
            'Avg PI Prev (A)': r['A'],
            'Avg PI Cur (E)':  r['E'],
            'Total Δ':         r['total'],
            '1. Churned Eff':  r['eff_dep'],
            '2. Price Eff':    r['eff_price'],
            '3. Comp Price Eff':     r['eff_comp'],
            '3.1 Normal Comp Eff': r['eff_normal_comp'],
            '3.2 Discount Comp Eff': r['eff_discount_comp'],
            '4. New SKU Eff':      r['eff_new'],
        })

    out = pd.DataFrame(rows)
    if not out.empty:
        out = out.sort_values('Pareto Class').reset_index(drop=True)
    return out


# ─────────────────────────────────────────────────────────────────────────────
# BUILDER: Price × Comp Matrix (Zone 5)
# ─────────────────────────────────────────────────────────────────────────────
def build_price_comp_matrix(df, scope='Overall'):
    """
    3×3 matrix: rows = price_tag (Down/Stay/Up), cols = comp_tag (Down/Stay/Up).
    Cells = count of Existing SKU.
    """
    if scope != 'Overall':
        df_scope = df[df['pricing_bl_25'] == scope].copy()
    else:
        df_scope = df.copy()

    ex = df_scope[df_scope['sku_type'] == 'Existing']
    if len(ex) == 0:
        return None, None

    DIRS = ['Down', 'Stay', 'Up']
    matrix = np.zeros((3, 3), dtype=int)
    for ri, pr in enumerate(DIRS):
        for ci, cr in enumerate(DIRS):
            matrix[ri, ci] = int(((ex['price_tag'] == pr) & (ex['comp_tag'] == cr)).sum())

    df_matrix = pd.DataFrame(
        matrix,
        index=[f'PRICE {d}' for d in DIRS],
        columns=[f'COMP {d}' for d in DIRS],
    )
    return df_matrix, len(ex)


# ─────────────────────────────────────────────────────────────────────────────
# BUILDER: PI Distribution (Zone 6)
# ─────────────────────────────────────────────────────────────────────────────
def build_pi_distribution(df, scope='Overall'):
    """PI bucket distribution Prev vs Cur. Existing SKU only."""
    if scope != 'Overall':
        df_scope = df[df['pricing_bl_25'] == scope].copy()
    else:
        df_scope = df.copy()

    ex = df_scope[df_scope['sku_type'] == 'Existing']
    n = len(ex)
    if n == 0:
        return pd.DataFrame()

    rows = []
    for lbl in PI_BINS_LBL:
        nc = int((ex['pi_group_prev'].astype(str) == lbl).sum())
        nn = int((ex['pi_group_cur'].astype(str) == lbl).sum())
        rows.append({
            'PI Bucket': lbl,
            'n Prev':    nc,
            'n Cur':     nn,
            'Δ Count':   nn - nc,
            '% Prev':    nc/n if n > 0 else 0,
            '% Cur':     nn/n if n > 0 else 0,
            'Δ %':       (nn - nc)/n if n > 0 else 0,
        })
    return pd.DataFrame(rows)


# ─────────────────────────────────────────────────────────────────────────────
# BUILDER: Top Movers SKU (Zone 7)
# ─────────────────────────────────────────────────────────────────────────────
def build_top_movers_pi(df, n=30):
    """Top N gainers + losers by absolute |diff_pi|."""
    ex = df[df['sku_type'] == 'Existing'].dropna(subset=['diff_pi']).copy()
    if len(ex) == 0:
        return pd.DataFrame(), pd.DataFrame()

    gainers = ex.nlargest(n, 'diff_pi')
    losers = ex.nsmallest(n, 'diff_pi')
    return gainers, losers


def fmt_mover_row(r, idx):
    """Format single row for mover table display.
    Each dimension (Price, Comp, COGS) has 3 separate columns: P1, P2, Δ%.
    """
    pi_p1 = r.get('pi', np.nan)
    pi_p2 = r.get('next_pi', np.nan)
    mgn_p1 = r.get('margin_pct_prev', np.nan)
    mgn_p2 = r.get('margin_pct_cur', np.nan)

    # Price before/after
    price_p1 = r.get('price', np.nan)
    price_p2 = r.get('next_price', np.nan)
    price_diff_pct = ((price_p2 - price_p1) / price_p1) if (pd.notna(price_p1) and pd.notna(price_p2) and price_p1 != 0) else np.nan

    # Competitor BLENDED price before/after
    comp_p1 = r.get('comp_price', np.nan)
    comp_p2 = r.get('next_comp_price', np.nan)
    comp_diff_pct = ((comp_p2 - comp_p1) / comp_p1) if (pd.notna(comp_p1) and pd.notna(comp_p2) and comp_p1 != 0) else np.nan

    # COGS before/after
    cogs_p1 = r.get('cogs', np.nan)
    cogs_p2 = r.get('next_cogs', np.nan)
    cogs_diff_pct = ((cogs_p2 - cogs_p1) / cogs_p1) if (pd.notna(cogs_p1) and pd.notna(cogs_p2) and cogs_p1 != 0) else np.nan

    def fmt_n(v):
        return f"{v:,.0f}" if pd.notna(v) else "—"

    return {
        '#': idx,
        'Product': str(r.get('product_name', '—'))[:35],
        'BL': r.get('pricing_bl_25', '—'),
        'L1': r.get('l1_category_name', '—'),
        'PI P1 → P2': f"{pi_p1:.1f} → {pi_p2:.1f}" if pd.notna(pi_p1) and pd.notna(pi_p2) else "—",
        'Δ PI': r.get('diff_pi', 0),
        # Price (3 cols)
        'Price P1':    fmt_n(price_p1),
        'Price P2':    fmt_n(price_p2),
        'Price Δ%':    price_diff_pct,
        # Comp Blended (3 cols)
        'Comp P1':     fmt_n(comp_p1),
        'Comp P2':     fmt_n(comp_p2),
        'Comp Δ%':     comp_diff_pct,
        # COGS (3 cols)
        'COGS P1':     fmt_n(cogs_p1),
        'COGS P2':     fmt_n(cogs_p2),
        'COGS Δ%':     cogs_diff_pct,
        # Margin
        'Mgn P1 → P2': f"{mgn_p1*100:.1f}% → {mgn_p2*100:.1f}%" if pd.notna(mgn_p1) and pd.notna(mgn_p2) else "—",
        # Effects
        'Price Eff':       r.get('eff_price', 0),
        'Comp Price Eff':  r.get('eff_comp', 0),
        'Normal Comp Eff': r.get('eff_normal_comp', 0),
        'Discount Comp Eff': r.get('eff_discount_comp', 0),
        'Framework':       r.get('framework_check', '') or '',
    }


# ─────────────────────────────────────────────────────────────────────────────
# BUILDER: Framework Check (Zone 9)
# ─────────────────────────────────────────────────────────────────────────────
def build_framework_check_table(df):
    """Filter SKU where framework_check == 'TRUE' and classify by rule."""
    fc = df[df['framework_check'] == 'TRUE'].copy()
    if len(fc) == 0:
        return pd.DataFrame()

    # Classify rule based on conditions
    def classify_rule(row):
        bl = row.get('pricing_bl_25', '')
        pi = row.get('next_pi', np.nan)
        mg = row.get('margin_pct_cur', np.nan)
        if pd.isna(pi) or pd.isna(mg):
            return 'Unknown'
        if bl == 'Fresh':
            if pi > 110 and mg <= 0.15: return 'Rule 1: Fresh PI>110 + Margin≤15% (room to drop price)'
            if pi > 120 and mg >= 0.70: return 'Rule 3: Fresh PI>120 + Margin≥70% (over-priced premium)'
        elif bl == 'Frozen':
            if pi > 100 and mg <= 0.15: return 'Rule 2: Frozen PI>100 + Margin≤15% (room to drop)'
        elif bl == 'Dry':
            if pi < 105 and mg <= 0.00: return 'Rule 4: Dry PI<105 + Margin≤0% (loss leader, raise)'
            if pi > 120 and mg > 0.40: return 'Rule 5: Dry PI>120 + Margin>40% (over-priced, drop)'
        return 'Other'

    fc['Rule'] = fc.apply(classify_rule, axis=1)

    cols = ['Rule', 'product_id', 'product_name', 'pricing_bl_25', 'l1_category_name',
            'pi', 'next_pi', 'diff_pi', 'margin_pct_prev', 'margin_pct_cur',
            'price', 'next_price', 'cogs', 'next_cogs', 'comp_price', 'next_comp_price']
    cols = [c for c in cols if c in fc.columns]
    out = fc[cols].sort_values(['Rule', 'next_pi'], ascending=[True, False]).reset_index(drop=True)
    return out


# ─────────────────────────────────────────────────────────────────────────────
# BUILDER: Structural Loss (Zone 8a) + COGS Need Improve (Zone 8b)
# ─────────────────────────────────────────────────────────────────────────────
def build_structural_loss(df):
    """L1 categories with high CI Group D (95-105) or E (>105)."""
    ex = df[df['ci_cur'].notna() & df['next_pi'].notna() & df['margin_pct_cur'].notna()].copy()
    if len(ex) == 0:
        return pd.DataFrame()

    cat_total = ex.groupby('l1_category_name').size().to_dict()
    rows = []
    for grp_label, grp_filter in [
        ('COGS Index D (95-105)', 'D.95-105'),
        ('COGS Index E (>105)',   'E.>105'),
    ]:
        sub = ex[ex['ci_group_cur'] == grp_filter]
        if len(sub) == 0:
            continue
        by_cat = sub.groupby('l1_category_name').agg(
            n_sku=('product_id', 'count'),
            avg_ci=('ci_cur', 'mean'),
            avg_mg=('margin_pct_cur', 'mean'),
            avg_pi=('next_pi', 'mean')
        ).sort_values('n_sku', ascending=False).reset_index()
        for _, row in by_cat.iterrows():
            pct = row['n_sku'] / cat_total.get(row['l1_category_name'], 1)
            rows.append({
                'CI Group': grp_label,
                'L1 Category': row['l1_category_name'],
                'n SKU': int(row['n_sku']),
                '% of Category': pct,
                'Avg COGS Index': row['avg_ci'],
                'Avg Margin %': row['avg_mg'],
                'Avg PI (Cur)': row['avg_pi'],
            })

    return pd.DataFrame(rows)


def build_cogs_need_improve(df):
    """SKU list with CI Group D or E. Sorted by CI desc."""
    ex = df[df['ci_group_cur'].isin(['D.95-105', 'E.>105'])].copy()
    if len(ex) == 0:
        return pd.DataFrame()
    ex = ex.sort_values('ci_cur', ascending=False)
    cols = ['product_id', 'product_name', 'l1_category_name', 'pricing_bl_25',
            'pareto_classification', 'sku_type',
            'ci_group_cur', 'ci_cur', 'next_pi', 'margin_pct_cur',
            'next_price', 'next_cogs', 'next_comp_price']
    cols = [c for c in cols if c in ex.columns]
    return ex[cols].reset_index(drop=True)


# ─────────────────────────────────────────────────────────────────────────────
# SESSION STATE
# ─────────────────────────────────────────────────────────────────────────────
if 'pi_analysis' not in st.session_state:
    st.session_state.pi_analysis = None
if 'pi_file_bytes' not in st.session_state:
    st.session_state.pi_file_bytes = None
if 'pi_uploaded_file_name' not in st.session_state:
    st.session_state.pi_uploaded_file_name = None
if 'pi_uploaded_file_size' not in st.session_state:
    st.session_state.pi_uploaded_file_size = None


# ─────────────────────────────────────────────────────────────────────────────
# HEADER
# ─────────────────────────────────────────────────────────────────────────────
col_title, col_clear = st.columns([5, 1])
with col_title:
    st.title("📈 PI Analyzer")
    st.caption("Pricing Index Movement Decomposition vs Competitor — Astro Pricing Strategy")
with col_clear:
    st.write("")
    if st.session_state.pi_analysis is not None:
        if st.button("🗑️ Clear data", type="secondary", use_container_width=True, key='pi_clear'):
            st.session_state.pi_analysis = None
            st.session_state.pi_file_bytes = None
            st.session_state.pi_uploaded_file_name = None
            st.session_state.pi_uploaded_file_size = None
            st.session_state.pi_excel_ready = False
            if 'pi_excel_bytes' in st.session_state:
                del st.session_state.pi_excel_bytes
            st.rerun()


# ═══════════════════════════════════════════════════════════════════════════
# ZONE 1 — UPLOAD
# ═══════════════════════════════════════════════════════════════════════════
st.markdown('<div class="section-header">📁 Upload Data</div>', unsafe_allow_html=True)

up_col1, up_col2 = st.columns([3, 1])
with up_col1:
    uploaded = st.file_uploader(
        "Upload file Excel atau CSV (raw input per-SKU per-periode)",
        type=['xlsx', 'xls', 'csv'],
        help=f"Format mengikuti pi_analyzer_v1.py. Required columns: {', '.join(REQUIRED_COLS[:8])}...",
        key='pi_uploader',
    )
with up_col2:
    st.write("")
    st.write("")
    template_bytes = make_template_pi_excel()
    st.download_button(
        "📥 Download Template",
        data=template_bytes,
        file_name="pi_template.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
        key='pi_template_download',
    )

# ─────────────────────────────────────────────────────────────────────────────
# PROCESS UPLOAD
# ─────────────────────────────────────────────────────────────────────────────
file_changed = (
    uploaded is not None and
    (st.session_state.pi_uploaded_file_name != uploaded.name or
     st.session_state.pi_uploaded_file_size != uploaded.size or
     st.session_state.pi_analysis is None)
)

if file_changed:
    # Read file bytes ONCE for caching (read once, use everywhere)
    file_bytes = uploaded.getvalue()
    try:
        if uploaded.name.lower().endswith('.csv'):
            df_input = pd.read_csv(io.BytesIO(file_bytes))
        else:
            df_input = pd.read_excel(io.BytesIO(file_bytes))
    except Exception as e:
        st.error(f"❌ Gagal load file: {e}")
        st.stop()

    # Validation panel
    st.markdown("### ✅ Validation")
    cols_in = list(df_input.columns)

    val_col1, val_col2 = st.columns(2)
    with val_col1:
        st.markdown(f"**File:** `{uploaded.name}`")
        st.markdown(f"**Rows:** {len(df_input):,}")
        st.markdown(f"**Cols:** {len(cols_in)}")

        # Period detection
        period_pair = None
        for p1c, p2c in PERIOD_PAIRS:
            if p1c in cols_in and p2c in cols_in:
                period_pair = (p1c, p2c)
                break
        if period_pair:
            try:
                p1_val = str(df_input[period_pair[0]].dropna().iloc[0])[:10]
                p2_val = str(df_input[period_pair[1]].dropna().iloc[0])[:10]
                st.success(f"✅ Period: `{period_pair[0]}` ({p1_val}) vs `{period_pair[1]}` ({p2_val})")
            except Exception:
                st.warning(f"⚠️ Period cols ada tapi nilai tidak terbaca")
        else:
            st.error(f"❌ Period cols tidak ditemukan. Butuh: {PERIOD_PAIRS}")
            st.stop()

    with val_col2:
        # Required column check
        missing = [c for c in REQUIRED_COLS if c not in cols_in]
        if missing:
            st.error(f"❌ Missing kolom: `{missing}`")
            st.stop()
        else:
            st.success(f"✅ Semua {len(REQUIRED_COLS)} required columns OK")

    # Process button — FAST path now (cached compute, no Excel build yet)
    if st.button("🚀 Process Data", type="primary", key='pi_process'):
        with st.spinner("Running PI analysis (fast path)..."):
            try:
                result = cached_pi_compute(file_bytes, uploaded.name)
            except Exception as e:
                st.error(f"❌ Analyze error: {e}")
                st.exception(e)
                st.stop()

            st.session_state.pi_analysis = result
            st.session_state.pi_file_bytes = file_bytes  # for lazy Excel
            st.session_state.pi_uploaded_file_name = uploaded.name
            st.session_state.pi_uploaded_file_size = uploaded.size
            st.success(f"✅ Selesai! {len(result['df_enriched']):,} rows processed.")
            st.rerun()


# ═══════════════════════════════════════════════════════════════════════════
# MAIN DASHBOARD (only if analysis available)
# ═══════════════════════════════════════════════════════════════════════════
if st.session_state.pi_analysis is not None:
    result = st.session_state.pi_analysis
    df = result['df_enriched']
    overall = result['overall']
    segments = result['segments']
    contribs = result['contribs']
    period_p1 = result['period_p1']
    period_p2 = result['period_p2']

    # ─────────────────────────────────────────────────────────────────────────
    # HASIL ANALISIS HEADER
    # ─────────────────────────────────────────────────────────────────────────
    st.markdown(f'<div class="section-header">📈 Hasil Analisis — {period_p1} vs {period_p2}</div>',
                unsafe_allow_html=True)
    st.markdown(f"📌 File: `{st.session_state.pi_uploaded_file_name}` · {len(df):,} SKU diproses")

    # ─────────────────────────────────────────────────────────────────────────
    # ZONE 2 — EXECUTIVE SUMMARY (KPI Cards)
    # ─────────────────────────────────────────────────────────────────────────
    st.markdown('<div class="section-header">2️⃣ Executive Summary</div>', unsafe_allow_html=True)

    # Compute KPIs
    pi_a = overall['A']
    pi_e = overall['E']
    pi_delta = overall['total']
    n_ex = overall['n_ex']
    n_dep = overall['n_dep']
    n_new = overall['n_new']

    # ── Margin: P1 vs P2 (avg over Existing SKU) ──
    ex_only = df[df['sku_type'] == 'Existing']
    avg_margin_p1 = ex_only['margin_pct_prev'].mean() if 'margin_pct_prev' in ex_only.columns else np.nan
    avg_margin_p2 = ex_only['margin_pct_cur'].mean()  if 'margin_pct_cur'  in ex_only.columns else np.nan
    margin_delta_pp = (avg_margin_p2 - avg_margin_p1) if (pd.notna(avg_margin_p1) and pd.notna(avg_margin_p2)) else np.nan

    # ── Action SKU counts ──
    n_framework = int((df['framework_check'] == 'TRUE').sum())
    n_cogs_improve = int(df['ci_group_cur'].isin(['D.95-105', 'E.>105']).sum())

    # ── PI Distribution: Premium / Match / Undercut, P1 vs P2 ──
    #   Premium = D, E, F (PI > 105 - includes D for >105)
    #   Match = B, C (95 - 105)
    #   Undercut = A (< 95)
    PREMIUM_BUCKETS = ['D.105-110', 'E.110-120', 'F.>120']
    MATCH_BUCKETS   = ['B.95-<100', 'C.100-105']
    UNDERCUT_BUCKETS = ['A.<95']

    n_ex_total = len(ex_only) if len(ex_only) > 0 else 1

    n_prem_p2  = int(ex_only['pi_group_cur'].isin(PREMIUM_BUCKETS).sum())
    n_match_p2 = int(ex_only['pi_group_cur'].isin(MATCH_BUCKETS).sum())
    n_under_p2 = int(ex_only['pi_group_cur'].isin(UNDERCUT_BUCKETS).sum())

    n_prem_p1  = int(ex_only['pi_group_prev'].isin(PREMIUM_BUCKETS).sum())
    n_match_p1 = int(ex_only['pi_group_prev'].isin(MATCH_BUCKETS).sum())
    n_under_p1 = int(ex_only['pi_group_prev'].isin(UNDERCUT_BUCKETS).sum())

    # Convert to % of Existing
    pct_prem_p1 = n_prem_p1 / n_ex_total * 100
    pct_prem_p2 = n_prem_p2 / n_ex_total * 100
    pct_match_p1 = n_match_p1 / n_ex_total * 100
    pct_match_p2 = n_match_p2 / n_ex_total * 100
    pct_under_p1 = n_under_p1 / n_ex_total * 100
    pct_under_p2 = n_under_p2 / n_ex_total * 100

    # ── Row 1: PI Current, Total Δ PI, Avg Margin (with delta) ──
    c1, c2, c3 = st.columns(3)
    with c1:
        st.markdown(kpi_card(
            "Avg PI Current (E)",
            fmt_pi(pi_e),
            delta=pi_delta,
            delta_label=fmt_pi_delta(pi_delta),
            sub=f"P1 (A): {fmt_pi(pi_a)}"
        ), unsafe_allow_html=True)
    with c2:
        total_abs = abs(overall['eff_dep']) + abs(overall['eff_price']) + abs(overall['eff_comp']) + abs(overall['eff_new'])
        st.markdown(kpi_card(
            "Total Δ PI",
            fmt_pi_delta(pi_delta),
            sub=f"Sum |effects| = {total_abs:.3f}"
        ), unsafe_allow_html=True)
    with c3:
        # Margin delta in percentage points (pp)
        if pd.notna(margin_delta_pp):
            delta_label_str = f"{margin_delta_pp*100:+.2f} pp"
            sub_str = f"P1: {avg_margin_p1*100:.2f}%"
        else:
            delta_label_str = None
            sub_str = "Avg over SKU Existing"
        st.markdown(kpi_card(
            "Avg Margin % (Existing)",
            fmt_pct(avg_margin_p2) if pd.notna(avg_margin_p2) else "—",
            delta=margin_delta_pp if pd.notna(margin_delta_pp) else None,
            delta_label=delta_label_str,
            sub=sub_str
        ), unsafe_allow_html=True)

    # ── Row 2: SKU counts (Existing / New / Departing) ──
    c4, c5, c6 = st.columns(3)
    with c4:
        st.markdown(kpi_card(
            "# SKU Existing",
            fmt_count(n_ex),
            sub=f"Punya PI di P1 dan P2"
        ), unsafe_allow_html=True)
    with c5:
        st.markdown(kpi_card(
            "# SKU New",
            fmt_count(n_new),
            sub=f"Muncul di P2 only"
        ), unsafe_allow_html=True)
    with c6:
        st.markdown(kpi_card(
            "# SKU Departing",
            fmt_count(n_dep),
            sub=f"Ada di P1, hilang di P2"
        ), unsafe_allow_html=True)

    # ── Row 3: Action SKU + spacer ──
    c7, c8, c9 = st.columns(3)
    with c7:
        st.markdown(kpi_card(
            "🚨 Framework Triggered",
            fmt_count(n_framework),
            sub="SKU butuh repricing action"
        ), unsafe_allow_html=True)
    with c8:
        st.markdown(kpi_card(
            "⚠️ COGS Need Improve",
            fmt_count(n_cogs_improve),
            sub="CI Group D (95-105) atau E (>105)"
        ), unsafe_allow_html=True)
    with c9:
        st.markdown(kpi_card(
            "📦 Total SKU Existing (base)",
            fmt_count(n_ex_total),
            sub="Base of PI Distribution di bawah"
        ), unsafe_allow_html=True)

    # ── Row 4: PI Distribution (3 cards) — Premium / Match / Undercut, P1 → P2 ──
    st.markdown('<div style="margin-top: 8px; margin-bottom: 4px; font-size: 13px; color: #6B7280; font-weight: 600;">📊 PI POSITIONING (% of Existing SKU) — P1 → P2</div>', unsafe_allow_html=True)
    c10, c11, c12 = st.columns(3)

    with c10:
        d_prem = pct_prem_p2 - pct_prem_p1
        # Premium: lebih mahal = bad direction kalau naik (delta_inverse=True)
        st.markdown(kpi_card(
            "🔴 SKU Premium Priced",
            f"{pct_prem_p2:.1f}%",
            delta=d_prem,
            delta_label=f"{d_prem:+.1f} pp",
            sub=f"{n_prem_p2:,} SKU di PI > 105 (was {pct_prem_p1:.1f}% / {n_prem_p1:,}) · butuh repricing review",
            delta_inverse=True
        ), unsafe_allow_html=True)
    with c11:
        d_match = pct_match_p2 - pct_match_p1
        # Match: neutral position
        st.markdown(kpi_card(
            "🟡 SKU Match Comp",
            f"{pct_match_p2:.1f}%",
            delta=d_match,
            delta_label=f"{d_match:+.1f} pp",
            sub=f"{n_match_p2:,} SKU di PI 95-105 (was {pct_match_p1:.1f}% / {n_match_p1:,})"
        ), unsafe_allow_html=True)
    with c12:
        d_under = pct_under_p2 - pct_under_p1
        # Undercut: lebih murah dari comp. Naik = lebih kompetitif (bisa juga over-discount)
        st.markdown(kpi_card(
            "🟢 SKU Undercut Comp",
            f"{pct_under_p2:.1f}%",
            delta=d_under,
            delta_label=f"{d_under:+.1f} pp",
            sub=f"{n_under_p2:,} SKU di PI < 95 (was {pct_under_p1:.1f}% / {n_under_p1:,}) · loss leader / advantage"
        ), unsafe_allow_html=True)

    # ─────────────────────────────────────────────────────────────────────────
    # ZONE 3 — PI DECOMPOSITION (Tornado + Tabel 1 + Tabel 2)
    # ─────────────────────────────────────────────────────────────────────────
    st.markdown('<div class="section-header">3️⃣ PI Decomposition (Shapley)</div>', unsafe_allow_html=True)

    # Narrative
    direction = "naik" if pi_delta > 0 else "turun"
    # Top driver
    drivers = [
        ('1. Churned SKU', overall['eff_dep']),
        ('2. Price Change Effect', overall['eff_price']),
        ('3. Comp Effect', overall['eff_comp']),
        ('4. New SKU', overall['eff_new']),
    ]
    top_driver = max(drivers, key=lambda x: abs(x[1]))
    st.markdown(f"""
    PI overall **{direction}** dari **{pi_a:.2f}** ke **{pi_e:.2f}** (Δ **{pi_delta:+.3f}**).
    Driver utama: **{top_driver[0]}** ({top_driver[1]:+.3f}).
    Comp effect breakdown — Normal: **{overall['eff_normal_comp']:+.3f}**, Blended (Discount): **{overall['eff_discount_comp']:+.3f}**.
    """)

    # Scope filter for tornado
    scope_t = st.radio(
        "Scope:",
        ['Overall', 'Dry', 'Fresh', 'Frozen'],
        horizontal=True,
        key='pi_zone3_scope'
    )
    if scope_t == 'Overall':
        decomp_for_chart = overall
    else:
        decomp_for_chart = segments[scope_t]

    fig_tornado = make_pi_tornado(decomp_for_chart, scope_label=scope_t)
    st.plotly_chart(fig_tornado, use_container_width=True)

    # Tabel 1: Per-segment decomposition
    st.markdown("##### 📊 Tabel 1: PI Decomposition per Segment")
    st.caption("Step-by-step PI movement per segment. Sum dari 5 effect = Total Δ (math identity exact via Shapley).")

    t1 = build_pi_decomp_table(overall, segments, contribs)
    seg_cols = ['Overall', 'Dry', 'Fresh', 'Frozen']
    # Format display
    def fmt_t1_row(row):
        label = row['Step']
        is_baseline = 'Baseline' in label or 'Result' in label
        out = {'Step': label}
        for c in seg_cols:
            v = row[c]
            if is_baseline:
                out[c] = f"{v:.4f}" if pd.notna(v) else "—"
            else:
                out[c] = f"{v:+.4f}" if pd.notna(v) else "—"
        return out
    t1_display = pd.DataFrame([fmt_t1_row(r) for _, r in t1.iterrows()])

    # Color the effect rows
    def style_t1(orig, disp):
        # Compute vmax from non-baseline rows
        effect_mask = ~orig['Step'].str.contains('Baseline|Result', regex=True)
        vals = orig.loc[effect_mask, seg_cols].values.flatten()
        vals = [v for v in vals if pd.notna(v)]
        vmax = max(abs(v) for v in vals) if vals else 1
        if vmax == 0: vmax = 1

        def apply_row(row):
            label = row['Step']
            is_baseline = ('Baseline' in label) or ('Result' in label)
            is_total = 'Total Δ' in label
            styles = []
            for col in disp.columns:
                if col == 'Step':
                    if is_baseline:
                        styles.append('background-color: #F3F4F6; font-weight: 700;')
                    elif is_total:
                        styles.append('background-color: #DBEAFE; font-weight: 700;')
                    elif label.startswith('    '):
                        styles.append('padding-left: 24px;')
                    else:
                        styles.append('')
                else:
                    if is_baseline:
                        styles.append('background-color: #F3F4F6; font-weight: 700;')
                    elif is_total:
                        v = orig.loc[row.name, col]
                        styles.append(gradient_color(v, vmax) + ' font-weight: 700;')
                    else:
                        v = orig.loc[row.name, col]
                        styles.append(gradient_color(v, vmax))
            return styles
        return disp.style.apply(apply_row, axis=1)

    st.dataframe(style_t1(t1, t1_display), use_container_width=True, hide_index=True)

    # Tabel 2: Per-segment contribution to OVERALL
    st.markdown("##### 📊 Tabel 2: Segment Contribution to Overall (Exact Math Identity)")
    st.caption("Σ Dry + Fresh + Frozen = Overall (exact, zero residual). Methodology: exact_contributions formula.")

    t2 = build_pi_contrib_table(overall, contribs)
    t2_cols = ['Dry contrib', 'Fresh contrib', 'Frozen contrib', 'Overall (sum)', 'Overall actual']

    def fmt_t2_row(row):
        out = {'Effect': row['Effect']}
        is_ref = row['Effect'] in ('Overall PI Prev (A)', 'Overall PI Cur (E)')
        for c in t2_cols:
            if pd.isna(row[c]):
                out[c] = "—"
            elif is_ref:
                # Reference rows: PI value (not delta, not signed)
                out[c] = f"{row[c]:.4f}"
            else:
                # Effect & Total rows: signed delta format
                out[c] = f"{row[c]:+.4f}"
        return out
    t2_display = pd.DataFrame([fmt_t2_row(r) for _, r in t2.iterrows()])

    def style_t2(orig, disp):
        # vmax from EFFECT rows only (exclude reference + Total Δ to keep contrast)
        effect_mask = ~orig['Effect'].isin(['Overall PI Prev (A)', 'Overall PI Cur (E)', 'Total Δ'])
        vals = orig.loc[effect_mask, t2_cols].values.flatten()
        vals = [v for v in vals if pd.notna(v)]
        vmax = max(abs(v) for v in vals) if vals else 1
        if vmax == 0: vmax = 1

        def apply_row(row):
            label = orig.loc[row.name, 'Effect']
            is_ref = label in ('Overall PI Prev (A)', 'Overall PI Cur (E)')
            is_total = label == 'Total Δ'
            styles = []
            for col in disp.columns:
                if col == 'Effect':
                    if is_ref:
                        styles.append('background-color: #F3F4F6; font-weight: 700;')
                    elif is_total:
                        styles.append('background-color: #DBEAFE; font-weight: 700;')
                    else:
                        styles.append('')
                else:
                    if is_ref:
                        # Highlight only Overall actual cell, dim others
                        if col == 'Overall actual':
                            styles.append('background-color: #DBEAFE; font-weight: 700;')
                        else:
                            styles.append('background-color: #F3F4F6; color: #9CA3AF;')
                    elif is_total:
                        v = orig.loc[row.name, col]
                        styles.append(gradient_color(v, vmax) + ' font-weight: 700;')
                    else:
                        v = orig.loc[row.name, col]
                        styles.append(gradient_color(v, vmax))
            return styles
        return disp.style.apply(apply_row, axis=1)

    st.dataframe(style_t2(t2, t2_display), use_container_width=True, hide_index=True)

    # ─────────────────────────────────────────────────────────────────────────
    # ZONE 4 — L1 CATEGORY BREAKDOWN
    # ─────────────────────────────────────────────────────────────────────────
    st.markdown('<div class="section-header">4️⃣ L1 Category Movement</div>', unsafe_allow_html=True)
    st.caption("PI movement per L1 category dengan 5-component decomposition.")

    l1_scope = st.radio(
        "Scope L1:",
        ['Overall', 'Dry', 'Fresh', 'Frozen'],
        horizontal=True,
        key='pi_l1_scope'
    )
    l1_table = build_l1_pi_table(df, scope=l1_scope)
    if l1_table.empty:
        st.info("Tidak ada data L1 untuk scope ini.")
    else:
        # Format columns
        l1_display = l1_table.copy()
        for c in ['Avg PI Prev (A)', 'Avg PI Cur (E)']:
            l1_display[c] = l1_display[c].apply(lambda v: f"{v:.2f}" if pd.notna(v) else "—")
        for c in ['Total Δ', 'Driver Value', '1. Churned Eff', '2. Price Eff', '3. Comp Price Eff',
                  '3.1 Normal Comp Eff', '3.2 Discount Comp Eff', '4. New SKU Eff']:
            l1_display[c] = l1_display[c].apply(lambda v: f"{v:+.3f}" if pd.notna(v) else "—")
        for c in ['n Existing', 'n Departing', 'n New']:
            l1_display[c] = l1_display[c].apply(lambda v: f"{int(v):,}" if pd.notna(v) else "—")
        for c in ['Margin % P1', 'Margin % P2']:
            l1_display[c] = l1_display[c].apply(lambda v: f"{v*100:.1f}%" if pd.notna(v) else "—")
        l1_display['Margin Δ pp'] = l1_display['Margin Δ pp'].apply(
            lambda v: f"{v*100:+.2f} pp" if pd.notna(v) else "—"
        )

        # Reorder columns to put Driver + Margin right after Total Δ
        col_order = ['L1 Category', 'BL (sample)',
                     'n Existing', 'n Departing', 'n New',
                     'Avg PI Prev (A)', 'Avg PI Cur (E)', 'Total Δ',
                     'Driver', 'Driver Value',
                     'Margin % P1', 'Margin % P2', 'Margin Δ pp',
                     '1. Churned Eff', '2. Price Eff', '3. Comp Price Eff',
                     '3.1 Normal Comp Eff', '3.2 Discount Comp Eff', '4. New SKU Eff']
        col_order = [c for c in col_order if c in l1_display.columns]
        l1_display = l1_display[col_order]

        # Gradient on effect columns
        effect_cols_l1 = ['Total Δ', 'Driver Value', '1. Churned Eff', '2. Price Eff', '3. Comp Price Eff',
                         '3.1 Normal Comp Eff', '3.2 Discount Comp Eff', '4. New SKU Eff']
        vals = l1_table[[c for c in effect_cols_l1 if c in l1_table.columns]].values.flatten()
        vals = [v for v in vals if pd.notna(v)]
        vmax = max(abs(v) for v in vals) if vals else 1
        if vmax == 0: vmax = 1

        # Margin gradient
        mg_vals = l1_table['Margin Δ pp'].dropna().values
        mg_vmax = max(abs(v) for v in mg_vals) if len(mg_vals) > 0 else 0.05
        if mg_vmax == 0: mg_vmax = 0.05

        def apply_l1_style(row):
            styles = []
            for col in l1_display.columns:
                if col in effect_cols_l1:
                    v = l1_table.loc[row.name, col]
                    styles.append(gradient_color(v, vmax))
                elif col == 'Margin Δ pp':
                    v = l1_table.loc[row.name, col]
                    styles.append(gradient_color(v, mg_vmax))
                elif col == 'Driver':
                    # color driver name pill
                    styles.append('background-color: #F3F4F6; font-weight: 600;')
                else:
                    styles.append('')
            return styles

        st.markdown(f"**Scope: {l1_scope}** · {len(l1_display)} L1 categories (sorted by Total Δ desc)")
        st.caption("💡 **Driver** = effect dengan magnitude terbesar yang punya sign sama dengan Total Δ. "
                  "Untuk L1 dengan Total Δ positif: driver = effect yang push PI naik. Negatif: driver = effect yang push PI turun.")
        st.dataframe(
            l1_display.style.apply(apply_l1_style, axis=1),
            use_container_width=True, hide_index=True, height=600
        )

    # ─────────────────────────────────────────────────────────────────────────
    # ZONE 5 — PRICE × COMP MATRIX
    # ─────────────────────────────────────────────────────────────────────────
    st.markdown('<div class="section-header">5️⃣ Price × Comp Direction Matrix</div>', unsafe_allow_html=True)
    st.caption("Behavior pattern: SKU yang harga naik/turun vs gerakan kompetitor. Existing SKU only.")

    matrix_scope = st.radio(
        "Scope Matrix:",
        ['Overall', 'Dry', 'Fresh', 'Frozen'],
        horizontal=True,
        key='pi_matrix_scope'
    )
    matrix_df, n_total = build_price_comp_matrix(df, scope=matrix_scope)
    if matrix_df is None:
        st.info("Tidak ada data Existing SKU untuk scope ini.")
    else:
        # Display matrix + percentage with view toggle (Amount vs %)
        col_label, col_view = st.columns([3, 1])
        with col_label:
            st.markdown(f"**Scope: {matrix_scope}** · Total {n_total:,} Existing SKU")
        with col_view:
            matrix_view = st.radio(
                "View:",
                ['Amount (SKU)', '% of Total'],
                horizontal=False,
                key='pi_matrix_view',
                label_visibility='collapsed'
            )

        # Up/Down/Stay legend (threshold sesuai engine pi_analyzer_v1.py line 182-184)
        with st.expander("ℹ️ Apa arti Up / Stay / Down? (Threshold)", expanded=False):
            st.markdown("""
            **Threshold tagging** (sama untuk Price, COGS, Comp):

            | Tag | Kondisi |
            |---|---|
            | **Up** | Δ absolute ≥ **+5,000 IDR** **OR** Δ percent ≥ **+5%** |
            | **Down** | Δ absolute ≤ **-5,000 IDR** **OR** Δ percent ≤ **-5%** |
            | **Stay** | selain Up dan Down (Δ < 5,000 IDR dan abs Δ % < 5%) |

            **PRICE direction** = pergerakan harga Astro dari P1 ke P2
            **COMP direction** = pergerakan harga blended competitor dari P1 ke P2

            **Cell interpretation:**
            - **Diagonal (Down-Down / Stay-Stay / Up-Up)**: Astro **follow** gerakan competitor — pricing aligned
            - **Off-diagonal**: Astro **NOT follow** competitor — misaligned
            - **PRICE Stay × COMP Up**: competitor naik harga, Astro stay → PI Astro turun otomatis (lebih kompetitif), bisa jadi opportunity raise price
            - **PRICE Stay × COMP Down**: competitor turun harga, Astro stay → PI Astro naik otomatis (uncompetitive), butuh respond
            """)

        # Count matrix with Total row/col
        m_count = matrix_df.copy()
        m_count['Total'] = m_count.sum(axis=1)
        m_count.loc['TOTAL'] = m_count.sum(axis=0)

        # Pct matrix WITH Total row/col (lo minta ini)
        m_pct = matrix_df / n_total * 100
        m_pct['Total'] = m_pct.sum(axis=1)
        m_pct.loc['TOTAL'] = m_pct.sum(axis=0)

        if matrix_view == 'Amount (SKU)':
            # Render count matrix
            max_c = matrix_df.values.max() if matrix_df.values.size > 0 else 1
            def color_amt(val):
                if pd.isna(val) or val == 0: return ''
                # Skip Total cells (they have own styling)
                alpha = min(1.0, val / max_c) if max_c > 0 else 0
                r = int(255 - (255-22) * alpha)
                g = int(255 - (255-163) * alpha)
                b = int(255 - (255-74) * alpha)
                return f'background-color: rgb({r},{g},{b}); color: #111827;'

            # Style: only color non-Total cells
            def style_count(df_):
                styles = pd.DataFrame('', index=df_.index, columns=df_.columns)
                for i in df_.index:
                    for c in df_.columns:
                        if i != 'TOTAL' and c != 'Total':
                            styles.loc[i, c] = color_amt(df_.loc[i, c])
                        else:
                            styles.loc[i, c] = 'background-color: #F3F4F6; font-weight: 700;'
                return styles

            st.markdown("##### 📊 Count (Amount of SKU)")
            st.dataframe(
                m_count.style.format("{:,}").apply(style_count, axis=None),
                use_container_width=True
            )
        else:
            # Render % matrix
            max_p = (matrix_df / n_total * 100).values.max() if matrix_df.values.size > 0 else 1
            def color_pct(val):
                if pd.isna(val) or val == 0: return ''
                alpha = min(1.0, val / max_p) if max_p > 0 else 0
                r = int(255 - (255-22) * alpha)
                g = int(255 - (255-163) * alpha)
                b = int(255 - (255-74) * alpha)
                return f'background-color: rgb({r},{g},{b}); color: #111827;'

            def style_pct(df_):
                styles = pd.DataFrame('', index=df_.index, columns=df_.columns)
                for i in df_.index:
                    for c in df_.columns:
                        if i != 'TOTAL' and c != 'Total':
                            styles.loc[i, c] = color_pct(df_.loc[i, c])
                        else:
                            styles.loc[i, c] = 'background-color: #F3F4F6; font-weight: 700;'
                return styles

            st.markdown("##### 📊 % of Total Existing")
            st.dataframe(
                m_pct.style.format("{:.1f}%").apply(style_pct, axis=None),
                use_container_width=True
            )

        # Diagonal vs off-diagonal insight
        diag_sum = sum(matrix_df.iloc[i, i] for i in range(3))
        st.caption(f"💡 Diagonal (Price follows Comp direction): **{diag_sum:,}** SKU "
                  f"({diag_sum/n_total*100:.1f}%). "
                  f"Off-diagonal = misaligned dengan kompetitor.")

    # ─────────────────────────────────────────────────────────────────────────
    # ZONE 6 — PI DISTRIBUTION
    # ─────────────────────────────────────────────────────────────────────────
    st.markdown('<div class="section-header">6️⃣ PI Distribution & Movement</div>', unsafe_allow_html=True)
    st.caption("Berapa SKU yang bergerak antar PI bucket. Existing SKU only.")

    dist_scope = st.radio(
        "Scope Distribution:",
        ['Overall', 'Dry', 'Fresh', 'Frozen'],
        horizontal=True,
        key='pi_dist_scope'
    )
    dist_df = build_pi_distribution(df, scope=dist_scope)
    if dist_df.empty:
        st.info("Tidak ada data Existing untuk scope ini.")
    else:
        # Bar chart side-by-side Prev vs Cur
        fig_dist = go.Figure()
        fig_dist.add_trace(go.Bar(
            name='Prev',
            x=dist_df['PI Bucket'],
            y=dist_df['n Prev'],
            marker_color='#9CA3AF',
            text=dist_df['n Prev'].apply(lambda v: f"{int(v):,}"),
            textposition='outside',
        ))
        fig_dist.add_trace(go.Bar(
            name='Cur',
            x=dist_df['PI Bucket'],
            y=dist_df['n Cur'],
            marker_color='#2563EB',
            text=dist_df['n Cur'].apply(lambda v: f"{int(v):,}"),
            textposition='outside',
        ))
        fig_dist.update_layout(
            title=f"PI Bucket Distribution — {dist_scope}",
            barmode='group',
            height=400,
            plot_bgcolor='white',
            margin=dict(t=60, b=40, l=40, r=20),
        )
        fig_dist.update_yaxes(showgrid=True, gridcolor='#F3F4F6')
        st.plotly_chart(fig_dist, use_container_width=True)

        # Distribution table with delta
        dist_disp = dist_df.copy()
        dist_disp['n Prev']  = dist_disp['n Prev'].apply(lambda v: f"{int(v):,}")
        dist_disp['n Cur']   = dist_disp['n Cur'].apply(lambda v: f"{int(v):,}")
        dist_disp['Δ Count'] = dist_disp['Δ Count'].apply(lambda v: f"{v:+,}")
        dist_disp['% Prev']  = dist_disp['% Prev'].apply(lambda v: f"{v*100:.1f}%")
        dist_disp['% Cur']   = dist_disp['% Cur'].apply(lambda v: f"{v*100:.1f}%")
        dist_disp['Δ %']     = dist_disp['Δ %'].apply(lambda v: f"{v*100:+.1f}%")
        st.dataframe(dist_disp, use_container_width=True, hide_index=True)

    # ─────────────────────────────────────────────────────────────────────────
    # ZONE 7 — TOP MOVERS SKU
    # ─────────────────────────────────────────────────────────────────────────
    st.markdown('<div class="section-header">7️⃣ Top Movers SKU</div>', unsafe_allow_html=True)
    st.caption("Top 30 SKU per dimensi mover: PI overall, Price Effect, atau Comp Price Effect. "
               "Existing SKU only.")

    mover_tabs = st.tabs([
        "📊 By Δ PI (overall)",
        "💰 By Price Effect",
        "🏪 By Comp Price Effect",
    ])

    fmt_mover = {
        'Δ PI': '{:+.2f}',
        'Price Δ%':  '{:+.2%}',
        'Comp Δ%':   '{:+.2%}',
        'COGS Δ%':   '{:+.2%}',
        'Price Eff': '{:+.3f}',
        'Comp Price Eff': '{:+.3f}',
        'Normal Comp Eff': '{:+.3f}',
        'Discount Comp Eff': '{:+.3f}',
    }

    # Tab 1: by Δ PI
    with mover_tabs[0]:
        gainers, losers = build_top_movers_pi(df, n=30)
        mv_col1, mv_col2 = st.columns(2)
        with mv_col1:
            st.markdown("**🔼 Top 30 PI Gainers (PI naik = harga makin mahal vs comp)**")
            if gainers.empty:
                st.info("Tidak ada data.")
            else:
                gdf = pd.DataFrame([fmt_mover_row(r, i+1) for i, (_, r) in enumerate(gainers.iterrows())])
                st.dataframe(gdf.style.format(fmt_mover),
                             use_container_width=True, hide_index=True, height=600)
        with mv_col2:
            st.markdown("**🔽 Top 30 PI Losers (PI turun = lebih kompetitif vs comp)**")
            if losers.empty:
                st.info("Tidak ada data.")
            else:
                ldf = pd.DataFrame([fmt_mover_row(r, i+1) for i, (_, r) in enumerate(losers.iterrows())])
                st.dataframe(ldf.style.format(fmt_mover),
                             use_container_width=True, hide_index=True, height=600)

    # Tab 2: by Price Effect
    with mover_tabs[1]:
        st.caption("Mover dengan **Price Effect** (eff_price) terbesar — kontribusi perubahan harga Astro ke PI movement.")
        ex_p = df[df['sku_type'] == 'Existing'].dropna(subset=['eff_price']).copy()
        if len(ex_p) == 0:
            st.info("Tidak ada data dengan Price Effect.")
        else:
            price_gain = ex_p.nlargest(30, 'eff_price')
            price_loss = ex_p.nsmallest(30, 'eff_price')
            p_col1, p_col2 = st.columns(2)
            with p_col1:
                st.markdown("**🔼 Top 30 — Price Effect Push UP (Astro naik harga)**")
                df_top = pd.DataFrame([fmt_mover_row(r, i+1) for i, (_, r) in enumerate(price_gain.iterrows())])
                st.dataframe(df_top.style.format(fmt_mover),
                             use_container_width=True, hide_index=True, height=600)
            with p_col2:
                st.markdown("**🔽 Top 30 — Price Effect Push DOWN (Astro turun harga)**")
                df_top = pd.DataFrame([fmt_mover_row(r, i+1) for i, (_, r) in enumerate(price_loss.iterrows())])
                st.dataframe(df_top.style.format(fmt_mover),
                             use_container_width=True, hide_index=True, height=600)

    # Tab 3: by Comp Price Effect
    with mover_tabs[2]:
        st.caption("Mover dengan **Comp Price Effect** (eff_comp = Normal + Discount) terbesar — kontribusi perubahan harga competitor ke PI.")
        ex_c = df[df['sku_type'] == 'Existing'].dropna(subset=['eff_comp']).copy()
        if len(ex_c) == 0:
            st.info("Tidak ada data dengan Comp Effect.")
        else:
            comp_gain = ex_c.nlargest(30, 'eff_comp')
            comp_loss = ex_c.nsmallest(30, 'eff_comp')
            c_col1, c_col2 = st.columns(2)
            with c_col1:
                st.markdown("**🔼 Top 30 — Comp Effect Push UP (comp TURUN harga → PI Astro naik)**")
                df_top = pd.DataFrame([fmt_mover_row(r, i+1) for i, (_, r) in enumerate(comp_gain.iterrows())])
                st.dataframe(df_top.style.format(fmt_mover),
                             use_container_width=True, hide_index=True, height=600)
            with c_col2:
                st.markdown("**🔽 Top 30 — Comp Effect Push DOWN (comp NAIK harga → PI Astro turun)**")
                df_top = pd.DataFrame([fmt_mover_row(r, i+1) for i, (_, r) in enumerate(comp_loss.iterrows())])
                st.dataframe(df_top.style.format(fmt_mover),
                             use_container_width=True, hide_index=True, height=600)

    # ─────────────────────────────────────────────────────────────────────────
    # ZONE 8 — PI DISTRIBUTION (alt visual already in Zone 6)
    # We use this slot for: STRUCTURAL LOSS + COGS NEED IMPROVE
    # ─────────────────────────────────────────────────────────────────────────
    st.markdown('<div class="section-header">8️⃣ Quadrant Analysis (PI × Margin × CI)</div>', unsafe_allow_html=True)
    st.caption("Cross-tables untuk identify SKU clusters yang misaligned. Filter di kanan untuk customize view.")

    # Display label remap (fix spacing issue: '-20to-10%' → '-20% to -10%')
    MG_LBL_DISPLAY = {
        "A.<-20%":      "A. <-20%",
        "B.-20to-10%":  "B. -20% to -10%",
        "C.-10to0%":    "C. -10% to 0%",
        "D.0to10%":     "D. 0% to 10%",
        "E.10to20%":    "E. 10% to 20%",
        "F.20to30%":    "F. 20% to 30%",
        "G.30to50%":    "G. 30% to 50%",
        "H.>50%":       "H. >50%",
    }
    CI_LBL_DISPLAY = {
        "A.<70":   "A. <70",
        "B.70-85": "B. 70-85",
        "C.85-95": "C. 85-95",
        "D.95-105":"D. 95-105",
        "E.>105":  "E. >105",
    }
    PI_LBL_DISPLAY = {
        "A.<95":     "A. <95",
        "B.95-<100": "B. 95-100",
        "C.100-105": "C. 100-105",
        "D.105-110": "D. 105-110",
        "E.110-120": "E. 110-120",
        "F.>120":    "F. >120",
    }

    # ── FILTERS (3 controls) ──
    flt_col1, flt_col2, flt_col3 = st.columns(3)
    with flt_col1:
        quad_scope = st.radio(
            "Scope:",
            ['Overall', 'Dry', 'Fresh', 'Frozen'],
            horizontal=True,
            key='pi_quad_scope'
        )
    with flt_col2:
        quad_period = st.radio(
            "PI period:",
            ['Current (P2)', 'Prev (P1)'],
            horizontal=True,
            key='pi_quad_period'
        )
    with flt_col3:
        quad_view = st.radio(
            "View:",
            ['Amount (SKU)', '% of Total'],
            horizontal=True,
            key='pi_quad_view'
        )

    # Filter dataframe
    if quad_scope != 'Overall':
        ex_only_q = df[(df['sku_type'] == 'Existing') & (df['pricing_bl_25'] == quad_scope)]
    else:
        ex_only_q = df[df['sku_type'] == 'Existing']

    # Period suffix: '_prev' or '_cur'
    p_suffix = '_prev' if quad_period == 'Prev (P1)' else '_cur'
    pi_col = f'pi_group{p_suffix}'
    mg_col = f'margin_group{p_suffix}'
    ci_col = f'ci_group{p_suffix}'

    n_q_total = len(ex_only_q) if len(ex_only_q) > 0 else 1

    st.caption(f"**Scope: {quad_scope}** · **PI period: {quad_period}** · "
               f"**View: {quad_view}** · {n_q_total:,} Existing SKU")
    st.caption("📌 *Note: 'PI period' filter berlaku untuk tab 1-3. Tab 'PI Prev vs PI Cur Transition' selalu pakai Prev × Cur.*")

    quad_tabs = st.tabs([
        "PI vs Margin",
        "COGS Index vs PI",
        "COGS Index vs Margin",
        "🔁 PI Prev vs PI Cur (Transition)"
    ])

    def render_cross_matrix(cross, row_remap, col_remap, view_mode, n_total):
        """Render cross-tab with label remap + Total row/col + gradient."""
        # Remap labels (display only)
        cross_disp = cross.copy()
        cross_disp.index = [row_remap.get(i, i) for i in cross_disp.index]
        cross_disp.columns = [col_remap.get(c, c) for c in cross_disp.columns]

        if view_mode == 'Amount (SKU)':
            # Add Total row + col
            m = cross_disp.copy()
            m['Total'] = m.sum(axis=1)
            m.loc['TOTAL'] = m.sum(axis=0)

            # Gradient based on non-Total cells only
            inner = cross_disp.values
            max_v = inner.max() if inner.size > 0 else 1

            def style_amt(df_):
                styles = pd.DataFrame('', index=df_.index, columns=df_.columns)
                for i in df_.index:
                    for c in df_.columns:
                        if i == 'TOTAL' or c == 'Total':
                            styles.loc[i, c] = 'background-color: #F3F4F6; font-weight: 700;'
                        else:
                            val = df_.loc[i, c]
                            if pd.notna(val) and val > 0:
                                alpha = min(1.0, val / max_v) if max_v > 0 else 0
                                r = int(255 - (255-22) * alpha)
                                g = int(255 - (255-163) * alpha)
                                b = int(255 - (255-74) * alpha)
                                styles.loc[i, c] = f'background-color: rgb({r},{g},{b}); color: #111827;'
                return styles

            st.dataframe(m.style.format("{:,}").apply(style_amt, axis=None),
                         use_container_width=True)
        else:
            # % mode — % of n_total
            pct = cross_disp / n_total * 100
            pct_with_total = pct.copy()
            pct_with_total['Total'] = pct_with_total.sum(axis=1)
            pct_with_total.loc['TOTAL'] = pct_with_total.sum(axis=0)

            inner_pct = pct.values
            max_v = inner_pct.max() if inner_pct.size > 0 else 1

            def style_pct(df_):
                styles = pd.DataFrame('', index=df_.index, columns=df_.columns)
                for i in df_.index:
                    for c in df_.columns:
                        if i == 'TOTAL' or c == 'Total':
                            styles.loc[i, c] = 'background-color: #F3F4F6; font-weight: 700;'
                        else:
                            val = df_.loc[i, c]
                            if pd.notna(val) and val > 0:
                                alpha = min(1.0, val / max_v) if max_v > 0 else 0
                                r = int(255 - (255-22) * alpha)
                                g = int(255 - (255-163) * alpha)
                                b = int(255 - (255-74) * alpha)
                                styles.loc[i, c] = f'background-color: rgb({r},{g},{b}); color: #111827;'
                return styles

            st.dataframe(pct_with_total.style.format("{:.1f}%").apply(style_pct, axis=None),
                         use_container_width=True)

    with quad_tabs[0]:
        if pi_col in ex_only_q.columns and mg_col in ex_only_q.columns:
            cross1 = pd.crosstab(
                ex_only_q[pi_col],
                ex_only_q[mg_col]
            ).reindex(index=PI_BINS_LBL, columns=MG_BINS_LBL, fill_value=0)
            st.markdown(f"**PI Bucket (rows) × Margin Bucket (cols)**")
            render_cross_matrix(cross1, PI_LBL_DISPLAY, MG_LBL_DISPLAY, quad_view, n_q_total)
            st.caption("💡 High PI + Low Margin = uncompetitive + low profit (double whammy). "
                      "Low PI + High Margin = good underpriced position (room to test price up).")

    with quad_tabs[1]:
        if ci_col in ex_only_q.columns and pi_col in ex_only_q.columns:
            cross2 = pd.crosstab(
                ex_only_q[ci_col],
                ex_only_q[pi_col]
            ).reindex(index=CI_BINS_LBL, columns=PI_BINS_LBL, fill_value=0)
            st.markdown(f"**COGS Index (rows) × PI Bucket (cols)**")
            render_cross_matrix(cross2, CI_LBL_DISPLAY, PI_LBL_DISPLAY, quad_view, n_q_total)
            st.caption("💡 High CI (D/E) + High PI = structurally over-priced (cost lebih mahal dari comp + jual lebih mahal). "
                      "Vendor negotiation needed.")

    with quad_tabs[2]:
        if ci_col in ex_only_q.columns and mg_col in ex_only_q.columns:
            cross3 = pd.crosstab(
                ex_only_q[ci_col],
                ex_only_q[mg_col]
            ).reindex(index=CI_BINS_LBL, columns=MG_BINS_LBL, fill_value=0)
            st.markdown(f"**COGS Index (rows) × Margin Bucket (cols)**")
            render_cross_matrix(cross3, CI_LBL_DISPLAY, MG_LBL_DISPLAY, quad_view, n_q_total)
            st.caption("💡 High CI (D/E) + Low Margin = structural loss (cost mahal + margin tipis).")

    # ── Tab 4: PI Prev → PI Cur Transition Matrix ──
    with quad_tabs[3]:
        st.markdown(f"**PI Bucket Prev (rows, P1) × PI Bucket Cur (cols, P2)** — SKU movement antar bucket")
        # Note: transition matrix ignores quad_period filter (always uses Prev × Cur)
        # but respects quad_scope and quad_view
        if 'pi_group_prev' in ex_only_q.columns and 'pi_group_cur' in ex_only_q.columns:
            cross_trans = pd.crosstab(
                ex_only_q['pi_group_prev'],
                ex_only_q['pi_group_cur']
            ).reindex(index=PI_BINS_LBL, columns=PI_BINS_LBL, fill_value=0)
            render_cross_matrix(cross_trans, PI_LBL_DISPLAY, PI_LBL_DISPLAY, quad_view, n_q_total)

            # Movement summary
            n_total_trans = int(cross_trans.values.sum())
            n_diag = int(sum(cross_trans.iloc[i, i] for i in range(len(PI_BINS_LBL))))
            n_off = n_total_trans - n_diag

            # Direction (upper triangle = moved to higher bucket = more expensive vs comp; lower triangle = inverse)
            n_up = int(sum(cross_trans.iloc[i, j] for i in range(len(PI_BINS_LBL))
                                                  for j in range(len(PI_BINS_LBL)) if j > i))
            n_down = int(sum(cross_trans.iloc[i, j] for i in range(len(PI_BINS_LBL))
                                                    for j in range(len(PI_BINS_LBL)) if j < i))

            st.caption(
                f"💡 **Diagonal** (stay di bucket sama): **{n_diag:,}** SKU "
                f"({n_diag/n_total_trans*100:.1f}%). "
                f"**Off-diagonal** (shift): **{n_off:,}** SKU. "
                f"Upper triangle (PI naik = lebih mahal vs comp): **{n_up:,}**. "
                f"Lower triangle (PI turun = lebih kompetitif): **{n_down:,}**."
            )
            st.caption(
                "**Interpretation:** Pola transition kasih insight macro tentang shift positioning portfolio Astro. "
                "Mayoritas di diagonal = portfolio stabil. Banyak movement ke bawah-kiri = portfolio jadi lebih kompetitif. "
                "Banyak movement ke atas-kanan = portfolio jadi lebih premium / mahal."
            )

    # ─────────────────────────────────────────────────────────────────────────
    # ZONE 9 — FRAMEWORK CHECK
    # ─────────────────────────────────────────────────────────────────────────
    st.markdown('<div class="section-header">9️⃣ Framework Check — Actionable SKU List</div>',
                unsafe_allow_html=True)
    st.caption("SKU yang trigger business rules untuk repricing. Definisi 5 rules: lihat caption per row.")

    fc_df = build_framework_check_table(df)
    if fc_df.empty:
        st.info("✅ Tidak ada SKU yang trigger framework rules.")
    else:
        # Summary by rule
        rule_summary = fc_df['Rule'].value_counts().reset_index()
        rule_summary.columns = ['Rule', 'Count']
        st.markdown(f"**Total {len(fc_df)} SKU trigger framework. Breakdown:**")
        st.dataframe(rule_summary, hide_index=True, use_container_width=True)

        # Detail SKU di-hide dulu — expand on click
        with st.expander(f"📋 Detail SKU ({len(fc_df)} rows) — Klik untuk tampilkan", expanded=False):
            fc_disp = fc_df.copy()
            if 'pi' in fc_disp.columns:
                fc_disp['pi'] = fc_disp['pi'].apply(lambda v: f"{v:.1f}" if pd.notna(v) else "—")
                fc_disp['next_pi'] = fc_disp['next_pi'].apply(lambda v: f"{v:.1f}" if pd.notna(v) else "—")
                fc_disp['diff_pi'] = fc_disp['diff_pi'].apply(lambda v: f"{v:+.2f}" if pd.notna(v) else "—")
            for c in ['margin_pct_prev', 'margin_pct_cur']:
                if c in fc_disp.columns:
                    fc_disp[c] = fc_disp[c].apply(lambda v: f"{v*100:.1f}%" if pd.notna(v) else "—")
            for c in ['price', 'next_price', 'cogs', 'next_cogs', 'comp_price', 'next_comp_price']:
                if c in fc_disp.columns:
                    fc_disp[c] = fc_disp[c].apply(lambda v: f"{v:,.0f}" if pd.notna(v) else "—")
            st.dataframe(fc_disp, use_container_width=True, hide_index=True, height=500)

    # ─────────────────────────────────────────────────────────────────────────
    # ZONE 10 — STRUCTURAL LOSS + COGS NEED IMPROVE
    # ─────────────────────────────────────────────────────────────────────────
    st.markdown('<div class="section-header">🔟 Structural Loss & COGS Need Improve</div>',
                unsafe_allow_html=True)
    st.caption("L1 categories dengan banyak SKU di CI Group D/E (cost mahal vs comp) — procurement action needed.")

    # ── 10A: Structural Loss by L1 Category ──
    st.markdown("##### 📋 Structural Loss by L1 Category")
    st.caption("L1 dengan SKU di CI Group D (95-105) atau E (>105) — cost lebih mahal dari competitor. "
              "Sort by n SKU desc.")
    sl_df = build_structural_loss(df)
    if sl_df.empty:
        st.info("Tidak ada L1 di CI Group D/E.")
    else:
        sl_disp = sl_df.copy()
        sl_disp['% of Category'] = sl_disp['% of Category'].apply(lambda v: f"{v*100:.1f}%")
        sl_disp['Avg COGS Index'] = sl_disp['Avg COGS Index'].apply(lambda v: f"{v:.1f}")
        sl_disp['Avg Margin %'] = sl_disp['Avg Margin %'].apply(lambda v: f"{v*100:.1f}%")
        sl_disp['Avg PI (Cur)'] = sl_disp['Avg PI (Cur)'].apply(lambda v: f"{v:.1f}")
        sl_disp['n SKU'] = sl_disp['n SKU'].apply(lambda v: f"{v:,}")
        st.dataframe(sl_disp, use_container_width=True, hide_index=True, height=400)

    # ── 10B: COGS Need Improve — SKU summary ──
    st.markdown("##### 📋 COGS Need Improve — SKU Summary by BL × L1")
    ci_df = build_cogs_need_improve(df)
    if ci_df.empty:
        st.info("Tidak ada SKU di CI D/E.")
    else:
        # Enriched summary: BL × L1, with margin avg, PI avg, CI avg
        ci_summary = ci_df.groupby(['pricing_bl_25', 'l1_category_name']).agg(
            n_sku=('product_id', 'count'),
            avg_ci=('ci_cur', 'mean'),
            avg_pi=('next_pi', 'mean'),
            avg_mg=('margin_pct_cur', 'mean'),
        ).reset_index()
        ci_summary.columns = ['BL', 'L1 Category', 'n SKU', 'Avg COGS Index', 'Avg PI (Cur)', 'Avg Margin %']
        ci_summary = ci_summary.sort_values('n SKU', ascending=False)

        # Display summary
        ci_disp = ci_summary.copy()
        ci_disp['n SKU'] = ci_disp['n SKU'].apply(lambda v: f"{v:,}")
        ci_disp['Avg COGS Index'] = ci_disp['Avg COGS Index'].apply(lambda v: f"{v:.1f}")
        ci_disp['Avg PI (Cur)'] = ci_disp['Avg PI (Cur)'].apply(lambda v: f"{v:.1f}")
        ci_disp['Avg Margin %'] = ci_disp['Avg Margin %'].apply(lambda v: f"{v*100:.1f}%")
        st.caption(f"Total **{len(ci_df):,} SKU** need COGS improvement (CI Group D + E).")
        st.dataframe(ci_disp, use_container_width=True, hide_index=True, height=400)

        # Optional detail SKU list (hidden behind expander)
        with st.expander(f"📋 Detail SKU list ({len(ci_df):,} rows) — klik untuk tampilkan", expanded=False):
            ci_detail = ci_df.copy()
            for c in ['ci_cur', 'next_pi']:
                if c in ci_detail.columns:
                    ci_detail[c] = ci_detail[c].apply(lambda v: f"{v:.1f}" if pd.notna(v) else "—")
            if 'margin_pct_cur' in ci_detail.columns:
                ci_detail['margin_pct_cur'] = ci_detail['margin_pct_cur'].apply(
                    lambda v: f"{v*100:.1f}%" if pd.notna(v) else "—"
                )
            for c in ['next_price', 'next_cogs', 'next_comp_price']:
                if c in ci_detail.columns:
                    ci_detail[c] = ci_detail[c].apply(lambda v: f"{v:,.0f}" if pd.notna(v) else "—")
            st.dataframe(ci_detail, use_container_width=True, hide_index=True, height=500)

    # ─────────────────────────────────────────────────────────────────────────
    # ZONE 11 — PARETO CLASS MOVEMENT
    # ─────────────────────────────────────────────────────────────────────────
    st.markdown('<div class="section-header">1️⃣1️⃣ Pareto Class Movement</div>', unsafe_allow_html=True)
    st.caption("PI movement per Pareto class (A/B/C) — strategic view by SKU importance.")

    pareto_scope = st.radio(
        "Scope Pareto:",
        ['Overall', 'Dry', 'Fresh', 'Frozen'],
        horizontal=True,
        key='pi_pareto_scope'
    )
    pareto_table = build_pareto_pi_table(df, scope=pareto_scope)
    if pareto_table.empty:
        st.info("Tidak ada data Pareto.")
    else:
        pareto_display = pareto_table.copy()
        for c in ['Avg PI Prev (A)', 'Avg PI Cur (E)']:
            pareto_display[c] = pareto_display[c].apply(lambda v: f"{v:.2f}" if pd.notna(v) else "—")
        for c in ['Total Δ', '1. Churned Eff', '2. Price Eff', '3. Comp Price Eff',
                  '3.1 Normal Comp Eff', '3.2 Discount Comp Eff', '4. New SKU Eff']:
            pareto_display[c] = pareto_display[c].apply(lambda v: f"{v:+.3f}" if pd.notna(v) else "—")
        for c in ['n Existing', 'n Departing', 'n New']:
            pareto_display[c] = pareto_display[c].apply(lambda v: f"{int(v):,}" if pd.notna(v) else "—")

        effect_cols_p = ['Total Δ', '1. Churned Eff', '2. Price Eff', '3. Comp Price Eff',
                        '3.1 Normal Comp Eff', '3.2 Discount Comp Eff', '4. New SKU Eff']
        vals = pareto_table[effect_cols_p].values.flatten()
        vals = [v for v in vals if pd.notna(v)]
        vmax = max(abs(v) for v in vals) if vals else 1
        if vmax == 0: vmax = 1

        def apply_pareto_style(row):
            styles = []
            for col in pareto_display.columns:
                if col in effect_cols_p:
                    v = pareto_table.loc[row.name, col]
                    styles.append(gradient_color(v, vmax))
                else:
                    styles.append('')
            return styles

        st.dataframe(
            pareto_display.style.apply(apply_pareto_style, axis=1),
            use_container_width=True, hide_index=True
        )

    # ─────────────────────────────────────────────────────────────────────────
    # DOWNLOAD EXCEL (LAZY — only build when user clicks)
    # ─────────────────────────────────────────────────────────────────────────
    st.markdown('<div class="section-header">📥 Download Full Report</div>', unsafe_allow_html=True)
    st.caption("Download complete 13-sheet Excel report (sama persis output script pi_analyzer_v1.py). "
               "Build Excel akan butuh ~50-60 detik karena ada styling per-cell. Klik tombol di bawah saat lo butuh.")

    out_name = f"PI_Analysis_{result['period_type']}_{period_p1}_vs_{period_p2}.xlsx"

    # Lazy generation: only build Excel when user clicks "Prepare Excel"
    if 'pi_excel_ready' not in st.session_state:
        st.session_state.pi_excel_ready = False

    if not st.session_state.pi_excel_ready:
        if st.button("🔨 Build Excel Report (~60s)", type="primary", key='pi_excel_prepare'):
            with st.spinner("Building 13-sheet Excel workbook... ini bisa makan ~60 detik."):
                try:
                    excel_bytes = cached_pi_excel_bytes(
                        st.session_state.pi_file_bytes,
                        st.session_state.pi_uploaded_file_name
                    )
                    st.session_state.pi_excel_bytes = excel_bytes
                    st.session_state.pi_excel_ready = True
                    st.rerun()
                except Exception as e:
                    st.error(f"❌ Excel build error: {e}")
                    st.exception(e)
    else:
        st.success("✅ Excel report ready. Click below to download.")
        st.download_button(
            "📥 Download Excel (13 sheets)",
            data=st.session_state.pi_excel_bytes,
            file_name=out_name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary",
            use_container_width=False,
            key='pi_excel_download'
        )
        if st.button("🔄 Rebuild Excel", type="secondary", key='pi_excel_rebuild'):
            st.session_state.pi_excel_ready = False
            if 'pi_excel_bytes' in st.session_state:
                del st.session_state.pi_excel_bytes
            st.rerun()

else:
    st.info("⬆️ Upload file PI raw data untuk mulai analisis. Format mengikuti `pi_analyzer_v1.py`.")
