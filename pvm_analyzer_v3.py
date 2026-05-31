import os, sys, glob
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Colors & Formats ──────────────────────────────────────────────────────────
C_DARK="1F3864"; C_MID="2F5496"; C_LIGHT="D9E1F2"
C_TOTAL="BDD7EE"; C_GREY="F2F2F2"; C_GREEN2="E2EFDA"
C_OR="FCE4D6"; C_RED="FFE0E0"; C_WHITE="FFFFFF"; C_SUB="F7F9FC"
C_AMBER="FFF2CC"; C_PURPLE="EDE7F6"
F_NUM='#,##0'; F_PCT='0.00%'; F_PT='0.00'
BLS = ['Dry','Fresh','Frozen','PL']

def cl(n): return get_column_letter(n)

def progress(msg, step=None, total=None):
    if step and total:
        pct = int(step/total*100)
        bar = '█'*int(pct/5) + '░'*(20-int(pct/5))
        print(f"  [{bar}] {pct:3d}% | {msg}", flush=True)
    else:
        print(f"  ► {msg}", flush=True)

# ── Style helpers (NO diagonal — all headers in single row) ──────────────────
def st(ws, row, col, value=None, bold=False, italic=False,
       bg=None, fc="000000", align="left", fmt=None,
       bb=False, bt=False, size=10, wrap=False):
    c = ws.cell(row=row, column=col)
    if value is not None: c.value = value
    c.font = Font(name="Arial", bold=bold, italic=italic, color=fc, size=size)
    if bg: c.fill = PatternFill("solid", start_color=bg)
    ha = "center" if align=="center" else ("right" if align=="right" else "left")
    c.alignment = Alignment(horizontal=ha, vertical="center", wrap_text=wrap)
    if fmt: c.number_format = fmt
    sides = {}
    if bb: sides['bottom'] = Side(style='thin')
    if bt: sides['top']    = Side(style='medium')
    if sides: c.border = Border(**sides)
    return c

def hdr(ws, row, col, val, bg=C_MID, fc="FFFFFF"):
    """Single-row header — wrap_text=False prevents diagonal appearance"""
    c = ws.cell(row=row, column=col, value=val)
    c.font = Font(name="Arial", bold=True, color=fc, size=10)
    c.fill = PatternFill("solid", start_color=bg)
    c.alignment = Alignment(horizontal="center", vertical="center",
                             wrap_text=True)   # wrap not diagonal
    c.border = Border(bottom=Side(style='thin'))
    return c

def title_row(ws, row, cols, text, bg=C_DARK, fc="FFFFFF", h=20):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    st(ws, row, 1, text, bold=True, bg=bg, fc=fc, align="center", size=11)
    ws.row_dimensions[row].height = h

def sec(ws, row, cols, text, bg=C_MID, fc="FFFFFF"):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    st(ws, row, 1, text, bold=True, bg=bg, fc=fc, align="left")
    ws.row_dimensions[row].height = 16

def note_row(ws, row, cols, text):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    st(ws, row, 1, text, italic=True, fc="595959", bg="F8F8F8", align="left")
    ws.row_dimensions[row].height = 14

def set_col_widths(ws, widths):
    """widths: list of (col_letter, width)"""
    for col, w in widths:
        ws.column_dimensions[col].width = w

# ── File selector ─────────────────────────────────────────────────────────────
def select_file():
    sd = os.path.dirname(os.path.abspath(__file__))
    files = [f for f in
             glob.glob(os.path.join(sd,'*.csv')) +
             glob.glob(os.path.join(sd,'*.xlsx')) +
             glob.glob(os.path.join(sd,'*.xls'))
             if not os.path.basename(f).startswith('~') and
             '_enriched' not in os.path.basename(f)]
    if not files: print("No files found."); sys.exit(1)
    print("\n" + "="*60)
    print("  PVM ANALYZER v3.0")
    print("="*60)
    for i,f in enumerate(files,1):
        print(f"  [{i}] {os.path.basename(f)}")
    print()
    while True:
        try:
            c = int(input("  Select file number: "))
            if 1 <= c <= len(files): return files[c-1]
            print(f"  Enter 1-{len(files)}")
        except ValueError:
            print("  Enter a valid number")

def load_data(fp):
    ext = os.path.splitext(fp)[1].lower()
    return pd.read_csv(fp) if ext=='.csv' else pd.read_excel(fp)

def detect_period(df):
    if 'week_key' in df.columns and 'next_week' in df.columns:
        return 'week_key','next_week'
    if 'week_key' in df.columns and 'next_key' in df.columns:
        return 'week_key','next_key'
    if 'month_key' in df.columns and 'next_month' in df.columns:
        return 'month_key','next_month'
    return None,None

def get_dates(df,p1c,p2c):
    p1=str(df[p1c].dropna().iloc[0])[:10] if p1c and p1c in df.columns else 'P1'
    p2=str(df[p2c].dropna().iloc[0])[:10] if p2c and p2c in df.columns else 'P2'
    return p1,p2

def ensure_cols(df):
    for col in ['comp_price','comp_price1','pi','pi1','avg_stock','avg_stock1',
                'pareto_classification','margin_pct','margin1_pct']:
        if col not in df.columns: df[col]=np.nan
    return df

# ── Coverage status ───────────────────────────────────────────────────────────
def cov_status(x):
    if pd.isna(x): return None
    if x < 70:   return "(A) < 70"
    if x < 85:   return "(B) 70-85"
    if x < 95:   return "(C) 85-95"
    if x < 105:  return "(D) 95-105"
    return "(E) > 105"

def gp_grp(x):
    if pd.isna(x): return None
    if x < -0.20: return "(A) < -20%"
    if x < -0.10: return "(B) -20% to -10%"
    if x < 0:     return "(C) -10% to 0%"
    if x < 0.10:  return "(D) 0% to 10%"
    if x < 0.20:  return "(E) 10% to 20%"
    if x < 0.30:  return "(F) 20% to 30%"
    if x < 0.50:  return "(G) 30% to 50%"
    return "(H) > 50%"

def pi_grp(x):
    if pd.isna(x): return None
    if x < 80:  return "(A) < 80"
    if x < 95:  return "(B) 80-95"
    if x < 105: return "(C) 95-105"
    if x < 115: return "(D) 105-115"
    if x < 120: return "(E) 115-120"
    return "(F) > 120"

FLAG_MAP = {
    ('Up','Up','Flat'):'Priority',('Up','Flat','Flat'):'Review',
    ('Up','Drop','Flat'):'Review',('Flat','Up','Flat'):'Review',
    ('Flat','Flat','Flat'):'Hold',('Flat','Drop','Flat'):'Review',
    ('Drop','Up','Flat'):'Adjust',('Drop','Flat','Flat'):'Adjust',
    ('Drop','Drop','Flat'):'Priority',
}

def status_tag(diff,pct,ta=5000,tp=0.05):
    if pd.isna(diff) or pd.isna(pct): return None
    if diff>ta or pct>tp: return "Up"
    if diff<-ta or pct<-tp: return "Drop"
    return "Flat"

def fw_check(row):
    bl=row.get('pricing_bl',''); pi1=row.get('pi_p2',np.nan); m1=row.get('margin_p2',np.nan)
    if pd.isna(pi1) or pd.isna(m1): return None
    if bl=='Fresh' and ((pi1>110 and m1<=0.15) or (pi1>120 and m1>0.70)): return True
    if bl=='Frozen' and pi1>100 and m1<=0.15: return True
    if bl=='Dry' and ((pi1<105 and m1<=0) or (pi1>120 and m1>=0.40)): return True
    return None

# ── Enrich raw data ───────────────────────────────────────────────────────────
def enrich(df, p1c, p2c):
    df = df.copy()

    # Rename period cols
    if p1c and p1c in df.columns: df.rename(columns={p1c:'period_1'}, inplace=True)
    if p2c and p2c in df.columns: df.rename(columns={p2c:'period_2'}, inplace=True)

    # Rename all columns to clean names
    rename_map = {
        'selling_price':'price_p1','selling_price1':'price_p2',
        'cost_price':'cogs_p1','cost_price1':'cogs_p2',
        'qty':'qty_p1','qty1':'qty_p2',
        'comp_price':'comp_price_p1','comp_price1':'comp_price_p2',
        'pi':'pi_p1','pi1':'pi_p2',
        'avg_stock':'avg_stock_p1','avg_stock1':'avg_stock_p2',
        'margin_pct':'margin_pct_raw_p1','margin1_pct':'margin_pct_raw_p2',
        'pricing_bl_25':'pricing_bl',
        'l1_category_name':'l1_category',
        'business_lines_2025':'business_line',
        'pareto_classification':'pareto_class',
    }
    df.rename(columns={k:v for k,v in rename_map.items() if k in df.columns}, inplace=True)

    progress("SKU status tagging...", 1, 10)
    def sku_stat(r):
        qn  = pd.isna(r.get('qty_p1'))  or r.get('qty_p1')==0
        q1n = pd.isna(r.get('qty_p2'))  or r.get('qty_p2')==0
        if not qn and not q1n: return 'Existing'
        if qn  and not q1n:    return 'New'
        if not qn and q1n:     return 'Deprecated'
        return 'Unknown'
    df['sku_status'] = df.apply(sku_stat, axis=1)

    progress("GV, GP, unit margins...", 2, 10)
    df['gv_p1']  = df['qty_p1']  * df['price_p1']
    df['gv_p2']  = df['qty_p2']  * df['price_p2']
    df['gv_diff']= df['gv_p2']   - df['gv_p1']
    df['gv_diff_pct'] = df['gv_diff'] / df['gv_p1']

    df['gp_p1']  = df['qty_p1']  * (df['price_p1'] - df['cogs_p1'])
    df['gp_p2']  = df['qty_p2']  * (df['price_p2'] - df['cogs_p2'])
    df['gp_diff']= df['gp_p2']   - df['gp_p1']
    df['gp_diff_pct'] = df['gp_diff'] / df['gp_p1']

    df['unit_margin_p1']  = df['price_p1'] - df['cogs_p1']
    df['unit_margin_p2']  = df['price_p2'] - df['cogs_p2']
    df['unit_margin_diff']= df['unit_margin_p2'] - df['unit_margin_p1']
    df['unit_margin_diff_pct'] = df['unit_margin_diff'] / df['unit_margin_p1']

    df['margin_p1']  = df['gp_p1']  / df['gv_p1']
    df['margin_p2']  = df['gp_p2']  / df['gv_p2']
    df['margin_diff']= df['margin_p2'] - df['margin_p1']

    progress("Diffs...", 3, 10)
    df['qty_diff']     = df['qty_p2']   - df['qty_p1']
    df['qty_diff_pct'] = df['qty_diff'] / df['qty_p1']

    df['price_diff']     = df['price_p2']   - df['price_p1']
    df['price_diff_pct'] = df['price_diff'] / df['price_p1']

    df['cogs_diff']     = df['cogs_p2']   - df['cogs_p1']
    df['cogs_diff_pct'] = df['cogs_diff'] / df['cogs_p1']

    df['comp_price_diff']     = df['comp_price_p2']   - df['comp_price_p1']
    df['comp_price_diff_pct'] = df['comp_price_diff'] / df['comp_price_p1']

    df['pi_diff']     = df['pi_p2']   - df['pi_p1']
    df['pi_diff_pct'] = df['pi_diff'] / df['pi_p1']

    df['avg_stock_diff']     = df['avg_stock_p2']   - df['avg_stock_p1']
    df['avg_stock_diff_pct'] = df['avg_stock_diff'] / df['avg_stock_p1']

    progress("PVM effects (q_P1 basis)...", 4, 10)
    ex = df['sku_status']=='Existing'
    df.loc[ex,'cogs_effect_rp']    = -(df.loc[ex,'qty_p1'] * df.loc[ex,'cogs_diff'])
    df.loc[ex,'price_effect_rp']   =   df.loc[ex,'qty_p1'] * df.loc[ex,'price_diff']
    # vol_mix = residual: gp_p2 - gp_hyp2 where gp_hyp2 = qty_p1*(price_p2-cogs_p2)
    df.loc[ex,'gp_hyp2'] = df.loc[ex,'qty_p1'] * (df.loc[ex,'price_p2'] - df.loc[ex,'cogs_p2'])
    df.loc[ex,'vol_mix_effect_rp'] = df.loc[ex,'gp_p2'] - df.loc[ex,'gp_hyp2']

    progress("Tagging...", 5, 10)
    df['oos_flag']     = df['avg_stock_diff_pct'].apply(
        lambda x: 'OOS' if (not pd.isna(x) and x<=-0.10) else ('Normal' if not pd.isna(x) else None))
    df['cogs_status']  = df.apply(lambda r: status_tag(r.get('cogs_diff'),  r.get('cogs_diff_pct')),  axis=1)
    df['price_status'] = df.apply(lambda r: status_tag(r.get('price_diff'), r.get('price_diff_pct')), axis=1)
    df['comp_status']  = df.apply(lambda r: status_tag(r.get('comp_price_diff'), r.get('comp_price_diff_pct')), axis=1)

    progress("Framework check & flag price...", 6, 10)
    df['framework_check'] = df.apply(fw_check, axis=1)
    ex_mask = df['sku_status']=='Existing'
    df.loc[ex_mask,'flag_price'] = df[ex_mask].apply(
        lambda r: FLAG_MAP.get((r.get('cogs_status'),r.get('comp_status'),r.get('price_status')),None), axis=1)

    progress("Groupings...", 7, 10)
    df['gp_group_p1']  = df['margin_p1'].apply(gp_grp)
    df['gp_group_p2']  = df['margin_p2'].apply(gp_grp)
    df['pi_group_p1']  = df['pi_p1'].apply(pi_grp)
    df['pi_group_p2']  = df['pi_p2'].apply(pi_grp)

    progress("COGS vs Comp coverage...", 8, 10)
    df['cogs_vs_comp_p1'] = df['cogs_p1'] * 100 / df['comp_price_p1']
    df['cogs_vs_comp_p2'] = df['cogs_p2'] * 100 / df['comp_price_p2']
    df['coverage_status_p1'] = df['cogs_vs_comp_p1'].apply(cov_status)
    df['coverage_status_p2'] = df['cogs_vs_comp_p2'].apply(cov_status)

    return df

# ── Compute PVM ───────────────────────────────────────────────────────────────
def compute_pvm(df):
    progress("Computing PVM bridge...", 9, 10)
    ex  = df[df['sku_status']=='Existing'].copy()
    new = df[df['sku_status']=='New'].copy()
    dep = df[df['sku_status']=='Deprecated'].copy()

    def bl_calc(bl):
        e = ex[ex['pricing_bl']==bl]   if bl!='TOTAL' else ex
        n = new[new['pricing_bl']==bl] if bl!='TOTAL' else new
        d = dep[dep['pricing_bl']==bl] if bl!='TOTAL' else dep

        gv_ex1=(e['qty_p1']*e['price_p1']).sum()
        gv_ex2=(e['qty_p2']*e['price_p2']).sum()
        gp_ex1=(e['qty_p1']*(e['price_p1']-e['cogs_p1'])).sum()
        gp_ex2=(e['qty_p2']*(e['price_p2']-e['cogs_p2'])).sum()
        gv_dep=(d['qty_p1']*d['price_p1']).sum() if len(d) else 0
        gp_dep=(d['qty_p1']*(d['price_p1']-d['cogs_p1'])).sum() if len(d) else 0
        gv_new=(n['qty_p2']*n['price_p2']).sum() if len(n) else 0
        gp_new=(n['qty_p2']*(n['price_p2']-n['cogs_p2'])).sum() if len(n) else 0

        gp_start=gp_ex1+gp_dep; gp_end=gp_ex2+gp_new
        gv_start=gv_ex1+gv_dep; gv_end=gv_ex2+gv_new
        m_base=gp_start/gv_start if gv_start>0 else 0
        m_ex1=gp_ex1/gv_ex1 if gv_ex1>0 else 0
        m_ex2=gp_ex2/gv_ex2 if gv_ex2>0 else 0
        m_end=gp_end/gv_end if gv_end>0 else 0

        # Hypotheticals
        gv_h1=(e['qty_p1']*e['price_p1']).sum()          # = gv_ex1
        cogs_h1=(e['qty_p1']*e['cogs_p2']).sum()
        gp_h1=gv_h1-cogs_h1; m_h1=gp_h1/gv_h1 if gv_h1>0 else 0

        gv_h2=(e['qty_p1']*e['price_p2']).sum()
        gp_h2=gv_h2-cogs_h1; m_h2=gp_h2/gv_h2 if gv_h2>0 else 0

        cogs_rp=gp_h1-gp_ex1; pp_cogs=m_h1-m_ex1
        price_rp=gp_h2-gp_h1; pp_price=m_h2-m_h1
        volmix_rp=gp_ex2-gp_h2; pp_volmix=m_ex2-m_h2
        pp_B=m_ex1-m_base; pp_G=m_end-m_ex2

        return dict(
            bl=bl,
            gv_ex1=gv_ex1,gv_ex2=gv_ex2,gp_ex1=gp_ex1,gp_ex2=gp_ex2,
            gv_dep=gv_dep,gp_dep=gp_dep,gv_new=gv_new,gp_new=gp_new,
            gp_start=gp_start,gp_end=gp_end,gv_start=gv_start,gv_end=gv_end,
            m_base=m_base,m_ex1=m_ex1,m_ex2=m_ex2,m_end=m_end,
            m_h1=m_h1,m_h2=m_h2,
            cogs_rp=cogs_rp,price_rp=price_rp,volmix_rp=volmix_rp,
            pp_B=pp_B,pp_cogs=pp_cogs,pp_price=pp_price,pp_volmix=pp_volmix,pp_G=pp_G,
            pp_total=m_end-m_base,
            new_cnt=len(n),dep_cnt=len(d),
            n_ex=len(e),
            gv_w1=0,gv_w2=0,
        )

    pvm={bl:bl_calc(bl) for bl in BLS+['TOTAL']}
    tot1=pvm['TOTAL']['gv_start']; tot2=pvm['TOTAL']['gv_end']
    for bl in BLS:
        pvm[bl]['gv_w1']=pvm[bl]['gv_start']/tot1 if tot1>0 else 0
        pvm[bl]['gv_w2']=pvm[bl]['gv_end']/tot2   if tot2>0 else 0
    pvm['TOTAL']['gv_w1']=1.0; pvm['TOTAL']['gv_w2']=1.0
    return pvm

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 0 — Formula Reference
# ══════════════════════════════════════════════════════════════════════════════
def write_s0(wb, p1, p2):
    # Insert at position 0 (before all other sheets)
    ws = wb.create_sheet("0. Formula Reference", 0)
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 55
    ws.column_dimensions['C'].width = 32
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 35

    row = 1
    title_row(ws, row, 5, f"FORMULA & TAGGING REFERENCE  |  {p1} vs {p2}")
    row += 1
    ws.merge_cells(f'A{row}:E{row}')
    st(ws, row, 1, "Semua formula dan flag logic yang dipakai di workbook ini. Kolom diff/diff_pct tidak ditampilkan.",
       italic=True, fc="595959", bg=C_LIGHT)
    row += 2

    def wsec(title, entries, bg_t=C_MID):
        nonlocal row
        sec(ws, row, 5, title, bg=bg_t); row += 1
        for i,h_ in enumerate(["Kolom / Tag","Formula / Logic","Input","Contoh Nilai","Notes"],1):
            hdr(ws,row,i,h_)
        ws.row_dimensions[row].height = 24; row += 1
        for idx,(col,formula,inputs,example,notes) in enumerate(entries):
            bg_ = C_GREY if idx%2==0 else C_WHITE
            st(ws,row,1,col,bold=True,bg=bg_)
            st(ws,row,2,formula,bg=bg_,align="left",wrap=True)
            st(ws,row,3,inputs,bg=bg_,align="left",italic=True,fc="595959")
            st(ws,row,4,example,bg=bg_,align="left")
            st(ws,row,5,notes,bg=bg_,align="left",italic=True,fc="595959")
            ws.row_dimensions[row].height = max(15, formula.count('\n')*14 + 15)
            row += 1
        row += 1

    wsec("1. SKU STATUS", [
        ("sku_status = Existing",
         "NOT NULL(qty_p1) AND NOT NULL(qty_p2)",
         "qty_p1, qty_p2","Existing","Ada di kedua periode"),
        ("sku_status = New",
         "NULL(qty_p1) AND NOT NULL(qty_p2)",
         "qty_p1, qty_p2","New","Hanya ada di P2"),
        ("sku_status = Deprecated",
         "NOT NULL(qty_p1) AND NULL(qty_p2)",
         "qty_p1, qty_p2","Deprecated","Hanya ada di P1"),
    ])

    wsec("2. DERIVED METRICS (per SKU)", [
        ("gv_p1 / gv_p2","qty_p1 x price_p1  |  qty_p2 x price_p2",
         "qty, price","1,250,000","Goods Value"),
        ("gp_p1 / gp_p2","qty_p1 x (price_p1 - cogs_p1)  |  qty_p2 x (price_p2 - cogs_p2)",
         "qty, price, cogs","312,500","Gross Profit"),
        ("margin_p1 / p2","gp_p1 / gv_p1  |  gp_p2 / gv_p2",
         "gp, gv","25.00%","Inventory Margin %"),
        ("unit_margin_p1 / p2","price_p1 - cogs_p1  |  price_p2 - cogs_p2",
         "price, cogs","25,000","Unit margin per item"),
    ])

    wsec("3. PVM EFFECTS — q_P1 BASIS (existing SKU only)", [
        ("cogs_effect_rp",
         "-(qty_p1 x (cogs_p2 - cogs_p1))",
         "qty_p1, cogs_p1, cogs_p2","(5,000,000)","Negatif = COGS naik"),
        ("price_effect_rp",
         "qty_p1 x (price_p2 - price_p1)",
         "qty_p1, price_p1, price_p2","+8,000,000","Positif = harga naik"),
        ("vol_mix_effect_rp",
         "gp_p2 - gp_hyp2\ngp_hyp2 = qty_p1 x (price_p2 - cogs_p2)",
         "qty_p1, qty_p2, price_p2, cogs_p2","+12,000,000","Residual. Absorbs Dq x Dmargin"),
    ])

    wsec("4. MARGIN BRIDGE — pp PER STEP", [
        ("1. Churned SKU Effect",
         "pp = M_ex_P1 - M_base\nM_ex_P1 = GP_ex_P1 / GV_ex_P1\nM_base = GP_start / GV_start",
         "GP & GV existing P1, dep P1","(0.0057pp)","+ = churned SKU margin jelek (cleaning)"),
        ("2. Existing SKU Effect",
         "pp = 2.1 + 2.2 + 2.3 (aggregate)",
         "Sum of COGS + Price + Vol/Mix effects","+0.1376pp","Total impact dari existing SKU"),
        ("  2.1 COGS Effect",
         "pp = M_hyp1 - M_ex_P1\nM_hyp1 = (GV_ex_P1 - COGS_hyp1) / GV_ex_P1\nCOGS_hyp1 = sum(q_P1 x cogs_P2)",
         "qty_p1, cogs_p1, cogs_p2","(0.0248pp)","GV tidak berubah. Paling bersih."),
        ("  2.2 Price Effect",
         "pp = M_hyp2 - M_hyp1\nGV_hyp2 = sum(q_P1 x price_P2)\nGP_hyp2 = GV_hyp2 - COGS_hyp1",
         "qty_p1, price_p2, cogs_p2","+0.0367pp","Qty masih P1. Paling bersih kedua."),
        ("  2.3 Vol/Mix Effect",
         "pp = M_ex_P2 - M_hyp2",
         "Actual P2 GP & GV existing","+0.1257pp","Residual. Absorbs interaction term."),
        ("3. New SKU Effect",
         "pp = M_end - M_ex_P2\nM_end = GP_all_P2 / GV_all_P2",
         "GP & GV all P2","+0.0157pp","+ = new SKU margin > existing"),
    ])

    wsec("5. COGS vs COMP COVERAGE", [
        ("cogs_vs_comp_p1 / p2",
         "cogs_p1 x 100 / comp_price_p1",
         "cogs, comp_price","82.5","Dalam poin, bukan %"),
        ("coverage_status",
         "(A) < 70  ruang besar vs kompetitor\n(B) 70-85 sehat\n(C) 85-95 watch\n(D) 95-105 at risk\n(E) > 105 critical",
         "cogs_vs_comp","(B) 70-85","Makin tinggi = makin berisiko"),
    ])

    wsec("6. GP% GROUP & PI GROUP TAGGING", [
        ("gp_group_p1 / p2",
         "(A)<-20%  (B)-20to-10%  (C)-10to0%\n(D)0-10%  (E)10-20%  (F)20-30%\n(G)30-50%  (H)>50%",
         "margin_p1 atau margin_p2","(E) 10% to 20%","Berdasarkan margin % aktual"),
        ("pi_group_p1 / p2",
         "(A)<80  (B)80-95  (C)95-105\n(D)105-115  (E)115-120  (F)>120",
         "pi_p1 atau pi_p2","(C) 95-105","PI = price x 100 / comp_price"),
    ])

    wsec("7. STATUS TAGS", [
        ("cogs_status / price_status / comp_status",
         "Up   = diff > 5,000 AND diff_pct > 5%\nDrop = diff < -5,000 AND diff_pct < -5%\nFlat = selain Up atau Drop",
         "diff, diff_pct masing-masing","Up / Drop / Flat","Threshold: 5,000 Rp dan 5%"),
        ("oos_flag",
         "OOS    = avg_stock_diff_pct <= -10%\nNormal = avg_stock_diff_pct > -10%",
         "avg_stock_p1, avg_stock_p2","Normal","Out of Stock indicator"),
    ])

    wsec("8. FRAMEWORK CHECK (existing SKU)", [
        ("framework_check (Fresh)",
         "TRUE jika: PI > 110 AND margin_p2 <= 15%\nATAU: PI > 120 AND margin_p2 > 70%",
         "pi_p2, margin_p2, pricing_bl","TRUE","Overpriced + margin tipis"),
        ("framework_check (Frozen)",
         "TRUE jika: PI > 100 AND margin_p2 <= 15%",
         "pi_p2, margin_p2","FALSE",""),
        ("framework_check (Dry)",
         "TRUE jika: PI < 105 AND margin_p2 <= 0%\nATAU: PI > 120 AND margin_p2 >= 40%",
         "pi_p2, margin_p2","FALSE",""),
    ])

    wsec("9. FLAG PRICE (existing SKU — matrix COGS x Comp x Price status)", [
        ("Priority",
         "COGS Up  + Comp Up  + Price Flat\nCOGS Drop + Comp Drop + Price Flat",
         "cogs_status, comp_status, price_status","Priority","Immediate action"),
        ("Review",
         "COGS Up  + Comp Flat + Price Flat\nCOGS Up  + Comp Drop + Price Flat\nCOGS Flat + Comp Up  + Price Flat\nCOGS Flat + Comp Drop + Price Flat",
         "cogs_status, comp_status, price_status","Review","Monitor closely"),
        ("Adjust",
         "COGS Drop + Comp Up  + Price Flat\nCOGS Drop + Comp Flat + Price Flat",
         "cogs_status, comp_status, price_status","Adjust","Opportunity to adjust price"),
        ("Hold",
         "COGS Flat + Comp Flat + Price Flat",
         "cogs_status, comp_status, price_status","Hold","Stable"),
    ])

    wsec("10. GROWTH QUALITY FLAG (L1 Category)", [
        ("Healthy",
         "margin_diff >= -0.5% AND vol_growth >= 0%",
         "diff_margin, vol_growth%","Healthy",""),
        ("Healthy — Mix Driven",
         "Healthy AND largest abs pp = Vol/Mix",
         "pp_cogs, pp_price, pp_volmix","Healthy — Mix Driven",""),
        ("Dilutive — Cost",  "margin_diff < -0.5% AND pp_cogs < -0.5%",  "diff_margin, pp_cogs","Dilutive — Cost",""),
        ("Dilutive — Price", "margin_diff < -0.5% AND pp_price < -0.5%", "diff_margin, pp_price","Dilutive — Price",""),
        ("Dilutive — Cost+Price","margin_diff < -0.5% AND keduanya < -0.5%","diff_margin, pp_cogs, pp_price","Dilutive — Cost+Price",""),
        ("Shrinking",        "vol_growth < 0%",                           "vol_growth%","Shrinking","GP turun + vol turun"),
    ])

    wsec("11. CONTEXT FLAG (Section E L1 Category — Price pp negatif)", [
        ("Intentional Promo",
         "price_diff < 0 AND vol_growth > +10%",
         "price_diff, vol_growth%","Intentional Promo","Strategic price cut"),
        ("Comp Pressure",
         "price_diff < 0 AND >50% SKU di kategori PI turun",
         "price_diff, pi_p1, pi_p2","Comp Pressure","Mengikuti kompetitor"),
        ("Price Erosion",
         "price_diff < 0 AND tidak masuk dua kondisi di atas",
         "price_diff","Price Erosion","Margin loss tanpa volume benefit"),
    ])

    wsec("12. CONCENTRATION FLAG (COGS Pressure)", [
        ("Concentrated",
         "Top 20% SKU by GV P1 drive > 60% of total |COGS impact|",
         "qty_p1, cogs_diff, gv_p1","Concentrated","Masalah terkonsentrasi di SKU besar"),
        ("Widespread",
         "Selain Concentrated",
         "qty_p1, cogs_diff, gv_p1","Widespread","Masalah merata di seluruh SKU"),
    ])


# ══════════════════════════════════════════════════════════════════════════════
# SHEET — Executive Overview
# ══════════════════════════════════════════════════════════════════════════════
def write_exec(wb, df, pvm, p1, p2):
    progress("Sheet: Executive Overview...", 0, 8)
    ws = wb.create_sheet("Executive Overview", 1)

    ws.column_dimensions['A'].width = 26
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 16
    ws.column_dimensions['F'].width = 14

    def B(x): return x  # full Rp, no conversion

    ex  = df[df['sku_status']=='Existing'].copy()
    new = df[df['sku_status']=='New'].copy()
    dep = df[df['sku_status']=='Deprecated'].copy()

    for d_ in [ex, new, dep]:
        if 'gv_p1' not in d_.columns:
            d_['gv_p1'] = d_['qty_p1'] * d_['price_p1'] if 'qty_p1' in d_.columns else 0
        if 'gv_p2' not in d_.columns:
            d_['gv_p2'] = d_['qty_p2'] * d_['price_p2'] if 'qty_p2' in d_.columns else 0
        if 'gp_p1' not in d_.columns:
            d_['gp_p1'] = d_['qty_p1'] * (d_['price_p1'] - d_['cogs_p1']) if 'qty_p1' in d_.columns else 0
        if 'gp_p2' not in d_.columns:
            d_['gp_p2'] = d_['qty_p2'] * (d_['price_p2'] - d_['cogs_p2']) if 'qty_p2' in d_.columns else 0

    qty_p1  = ex['qty_p1'].sum()  + dep['qty_p1'].sum()
    qty_p2  = ex['qty_p2'].sum()  + new['qty_p2'].sum()
    gv_p1   = ex['gv_p1'].sum()   + dep['gv_p1'].sum()
    gv_p2   = ex['gv_p2'].sum()   + new['gv_p2'].sum()
    cogs_p1 = (ex['qty_p1']*ex['cogs_p1']).sum() + (dep['qty_p1']*dep['cogs_p1']).sum()
    cogs_p2 = (ex['qty_p2']*ex['cogs_p2']).sum() + (new['qty_p2']*new['cogs_p2']).sum()
    gp_p1   = ex['gp_p1'].sum()   + dep['gp_p1'].sum()
    gp_p2   = ex['gp_p2'].sum()   + new['gp_p2'].sum()
    m_p1    = gp_p1/gv_p1 if gv_p1>0 else 0
    m_p2    = gp_p2/gv_p2 if gv_p2>0 else 0
    n_ex=len(ex); n_new=len(new); n_dep=len(dep)
    d_tot = pvm['TOTAL']

    row = 1
    title_row(ws, row, 6, f"EXECUTIVE OVERVIEW  |  {p1}  →  {p2}", h=24)
    row += 2

    # ── Section 1: Key Metrics ────────────────────────────────────────────────
    sec(ws, row, 6, "KEY METRICS — Overall (All SKU)")
    row += 1
    for i,h_ in enumerate(["Metric", p1, p2, "Diff", "Diff %", ""], 1):
        hdr(ws, row, i, h_)
    ws.row_dimensions[row].height = 24
    row += 1

    def mrow(label, v1, v2, fmt1, fmt2, is_pp=False, bg_=C_WHITE, bold=False):
        nonlocal row
        diff = v2 - v1
        diff_pct = diff/v1 if (v1!=0 and not is_pp) else None
        fc_d = "007030" if diff>=0 else "CC0000"
        st(ws,row,1,label,bg=bg_,bold=bold)
        st(ws,row,2,v1,fmt=fmt1,align="right",bg=bg_,bold=bold)
        st(ws,row,3,v2,fmt=fmt1,align="right",bg=bg_,bold=bold)
        st(ws,row,4,diff,fmt=fmt2,align="right",bg=bg_,fc=fc_d,bold=True)
        if diff_pct is not None:
            st(ws,row,5,diff_pct,fmt=F_PCT,align="right",bg=bg_,fc=fc_d)
        else:
            st(ws,row,5,"—",align="center",fc="888888",bg=bg_)
        st(ws,row,6,"",bg=bg_)
        row += 1

    mrow("Qty Sold",          qty_p1, qty_p2, '#,##0', '#,##0',       False, C_WHITE)
    mrow("Goods Value",        B(gv_p1), B(gv_p2), '#,##0',      '#,##0',       False, C_GREY)
    mrow("Total COGS",         B(cogs_p1),B(cogs_p2),'#,##0',    '#,##0',       False, C_WHITE)
    mrow("Gross Profit",       B(gp_p1), B(gp_p2), '#,##0',      '#,##0',       False, C_GREY, bold=True)
    # Margin special
    fc_m = "007030" if m_p2>=m_p1 else "CC0000"
    st(ws,row,1,"Inventory Margin",bg=C_WHITE,bold=True)
    st(ws,row,2,m_p1,fmt=F_PCT,align="right",bg=C_WHITE,bold=True)
    st(ws,row,3,m_p2,fmt=F_PCT,align="right",bg=C_WHITE,bold=True)
    st(ws,row,4,(m_p2-m_p1)*100,fmt='0.00"pp"',align="right",bg=C_WHITE,fc=fc_m,bold=True)
    st(ws,row,5,"—",align="center",fc="888888",bg=C_WHITE)
    st(ws,row,6,"",bg=C_WHITE)
    row += 2

    # Portfolio
    sec(ws, row, 6, "PORTFOLIO")
    row += 1
    for i,h_ in enumerate(["Metric", p1, p2, "Diff", "", ""], 1):
        hdr(ws, row, i, h_)
    ws.row_dimensions[row].height = 24
    row += 1

    for label, v1, v2, bg_ in [
        ("# SKU Active",     n_ex+n_dep, n_ex+n_new, C_WHITE),
        ("# SKU Existing",   n_ex,       n_ex,        C_GREY),
        ("# SKU New",        0,          n_new,       "EBF5D0"),
        ("# SKU Deprecated", n_dep,      0,           C_OR),
    ]:
        diff = v2 - v1
        fc_d = "007030" if diff>=0 else "CC0000"
        st(ws,row,1,label,bg=bg_)
        st(ws,row,2,v1 if v1>0 else "—",fmt='#,##0' if v1>0 else None,align="right",bg=bg_)
        st(ws,row,3,v2 if v2>0 else "—",fmt='#,##0' if v2>0 else None,align="right",bg=bg_)
        st(ws,row,4,diff if diff!=0 else "—",fmt='#,##0' if diff!=0 else None,
           align="right",bg=bg_,fc=fc_d if diff!=0 else "888888")
        st(ws,row,5,"",bg=bg_); st(ws,row,6,"",bg=bg_)
        row += 1
    row += 1

    # ── Section 2: Margin Bridge Summary ─────────────────────────────────────
    sec(ws, row, 6, "MARGIN BRIDGE SUMMARY — Overall")
    row += 1
    for i,h_ in enumerate(["Step","pp Impact","","← neg  |  pos →","",""], 1):
        hdr(ws, row, i, h_)
    ws.row_dimensions[row].height = 24
    row += 1

    # Store pp in actual pp units (× 100) for bar
    pp_existing = (d_tot['pp_cogs'] + d_tot['pp_price'] + d_tot['pp_volmix']) * 100
    bridge_steps = [
        ("1. Churned SKU Effect", d_tot['pp_B']*100,      C_RED),
        ("2. Existing SKU Effect", pp_existing,           "FFFDE7"),
        ("  2.1 COGS Effect",     d_tot['pp_cogs']*100,   "FFF3E0"),
        ("  2.2 Price Effect",    d_tot['pp_price']*100,  "E3F2FD"),
        ("  2.3 Vol/Mix Effect",  d_tot['pp_volmix']*100, "EDE7F6"),
        ("3. New SKU Effect",     d_tot['pp_G']*100,      "E8F5E9"),
    ]
    max_abs = max(abs(v) for _,v,_ in bridge_steps) or 1
    HALF = 10

    def bar(val, max_abs, half=10):
        filled = min(int(round(abs(val)/max_abs * half)), half)
        if val >= 0:
            return "░"*half + "█"*filled + "░"*(half-filled)
        else:
            return "░"*(half-filled) + "█"*filled + "░"*half

    for label, pp_val, bg_ in bridge_steps:
        fc_ = "007030" if pp_val>=0 else "CC0000"
        st(ws,row,1,label,bg=bg_)
        st(ws,row,2,pp_val,fmt='0.0000"pp"',align="right",bg=bg_,fc=fc_,bold=True)
        st(ws,row,3,"",bg=bg_)
        st(ws,row,4,bar(pp_val,max_abs),align="center",bg=bg_,fc=fc_)
        ws.column_dimensions['D'].width = 26
        st(ws,row,5,"",bg=bg_); st(ws,row,6,"",bg=bg_)
        row += 1

    from openpyxl.styles import Border, Side
    for c in range(1,7):
        ws.cell(row=row,column=c).border = Border(top=Side(style='thin'))
    # Exclude aggregate "2. Existing SKU Effect" row to avoid double count with its sub-items
    total_pp = sum(pp_val for label,pp_val,_ in bridge_steps if label != "2. Existing SKU Effect")
    fc_t = "007030" if total_pp>=0 else "CC0000"
    st(ws,row,1,"Total Change",bold=True,bg=C_TOTAL)
    st(ws,row,2,total_pp,fmt='0.0000"pp"',align="right",bold=True,bg=C_TOTAL,fc=fc_t)
    for c in [3,4,5,6]: st(ws,row,c,"",bg=C_TOTAL)
    row += 1
    st(ws,row,1,"Start Margin (P1)",bg=C_GREY)
    st(ws,row,2,m_p1,fmt=F_PCT,align="right",bg=C_GREY)
    for c in [3,4,5,6]: st(ws,row,c,"",bg=C_GREY)
    row += 1
    st(ws,row,1,"End Margin (P2)",bold=True,bg=C_GREY)
    st(ws,row,2,m_p2,fmt=F_PCT,align="right",bold=True,bg=C_GREY)
    for c in [3,4,5,6]: st(ws,row,c,"",bg=C_GREY)
    row += 2

    # ── Section 3: Top Movers ─────────────────────────────────────────────────
    sec(ws, row, 6, "TOP MOVERS — Per Pricing BL")
    row += 1
    for i,h_ in enumerate(["Metric","BL","Value","Note","",""], 1):
        hdr(ws, row, i, h_)
    ws.row_dimensions[row].height = 24
    row += 1

    bl_gp = {bl: pvm[bl]['gp_end']-pvm[bl]['gp_start'] for bl in ['Dry','Fresh','Frozen','PL']}
    bl_m  = {bl: pvm[bl]['m_end'] for bl in ['Dry','Fresh','Frozen','PL']}
    bl_vm = {bl: pvm[bl]['pp_volmix'] for bl in ['Dry','Fresh','Frozen','PL']}
    bl_pr = {bl: pvm[bl]['pp_price']  for bl in ['Dry','Fresh','Frozen','PL']}

    bg_gp = pvm['TOTAL']['gp_start']

    movers = [
        ("Best GP Growth",    max(bl_gp,key=bl_gp.get),
         f"+{bl_gp[max(bl_gp,key=bl_gp.get)]/1e9:.3f}B  ({bl_gp[max(bl_gp,key=bl_gp.get)]/pvm[max(bl_gp,key=bl_gp.get)]['gp_start']:+.1%})",
         "Highest absolute GP contribution", "EBF5D0","007030"),
        ("Worst GP Growth",   min(bl_gp,key=bl_gp.get),
         f"{bl_gp[min(bl_gp,key=bl_gp.get)]/1e9:+.3f}B  ({bl_gp[min(bl_gp,key=bl_gp.get)]/pvm[min(bl_gp,key=bl_gp.get)]['gp_start']:+.1%})",
         "GP drag on portfolio", C_OR,"CC0000"),
        ("Highest Margin",    max(bl_m,key=bl_m.get),
         f"{bl_m[max(bl_m,key=bl_m.get)]:.2%}",
         f"{(bl_m[max(bl_m,key=bl_m.get)]-m_p2)*100:+.2f}pp vs overall avg", "EBF5D0","007030"),
        ("Lowest Margin",     min(bl_m,key=bl_m.get),
         f"{bl_m[min(bl_m,key=bl_m.get)]:.2%}",
         f"{(bl_m[min(bl_m,key=bl_m.get)]-m_p2)*100:+.2f}pp vs overall avg", C_OR,"CC0000"),
        ("Biggest Vol/Mix +", max(bl_vm,key=bl_vm.get),
         f"{bl_vm[max(bl_vm,key=bl_vm.get)]*100:+.4f}pp",
         "Main growth driver within BL", "EBF5D0","007030"),
        ("Biggest Price drag",min(bl_pr,key=bl_pr.get),
         f"{bl_pr[min(bl_pr,key=bl_pr.get)]*100:+.4f}pp",
         "Largest price pressure within BL", C_RED,"CC0000"),
    ]

    for i,(label,bl,val,note,bg_val,fc_) in enumerate(movers):
        bg2_ = C_GREY if i%2==0 else C_WHITE
        st(ws,row,1,label,bg=bg2_)
        st(ws,row,2,bl,bg=bg_val,bold=True,align="center",fc=fc_)
        st(ws,row,3,val,bg=bg2_,align="right",fc=fc_,bold=True)
        st(ws,row,4,note,bg=bg2_,align="left",italic=True,fc="595959")
        st(ws,row,5,"",bg=bg2_); st(ws,row,6,"",bg=bg2_)
        row += 1
    row += 2

    # ── Section 4: Detail per BL ──────────────────────────────────────────────
    sec(ws, row, 6, "DETAIL PER PRICING BL", bg=C_DARK)
    row += 1
    for i,h_ in enumerate(["Metric","Dry","Fresh","Frozen","PL","Overall"], 1):
        hdr(ws, row, i, h_, bg=C_DARK)
    ws.row_dimensions[row].height = 24
    row += 1

    BLS5 = ['Dry','Fresh','Frozen','PL','TOTAL']

    def bl_row(label, vals, fmt='#,##0', is_pp=False, bg_=C_WHITE, bold=False, is_sec=False):
        nonlocal row
        if is_sec:
            ws.merge_cells(f'A{row}:F{row}')
            st(ws,row,1,label,bold=True,bg=C_LIGHT,fc=C_DARK)
            row += 1; return
        st(ws,row,1,label,bg=bg_,bold=bold)
        for ci,v in enumerate(vals,2):
            if v is None:
                st(ws,row,ci,"—",align="center",fc="888888",bg=bg_)
            else:
                fc_v = ("007030" if v>=0 else "CC0000") if is_pp else "000000"
                st(ws,row,ci,v,fmt=fmt,align="right",bg=bg_,fc=fc_v,bold=bold)
        row += 1

    # VOLUME — compute qty from df
    def bl_qty(bl, period='p1'):
        if bl=='TOTAL':
            mask_ex = df['sku_status']=='Existing'
            mask_nd = df['sku_status']==('New' if period=='p2' else 'Deprecated')
            col = 'qty_p2' if period=='p2' else 'qty_p1'
            return df.loc[mask_ex|mask_nd, col].sum()
        mask_ex = (df['sku_status']=='Existing') & (df['pricing_bl']==bl)
        mask_nd = (df['sku_status']==('New' if period=='p2' else 'Deprecated')) & (df['pricing_bl']==bl)
        col = 'qty_p2' if period=='p2' else 'qty_p1'
        return df.loc[mask_ex|mask_nd, col].sum()
    qty1s=[bl_qty(bl,'p1') for bl in BLS5]
    qty2s=[bl_qty(bl,'p2') for bl in BLS5]
    q_diffs=[q2-q1 for q1,q2 in zip(qty1s,qty2s)]
    bl_row("VOLUME", None, is_sec=True)
    bl_row("Qty P1", qty1s,'#,##0',bg_=C_WHITE)
    bl_row("Qty P2", qty2s,'#,##0',bg_=C_GREY)
    bl_row("Qty Diff",  q_diffs,'#,##0',bg_=C_WHITE)
    bl_row("Qty Diff %",[d/q1 if q1>0 else None for d,q1 in zip(q_diffs,qty1s)],F_PCT,bg_=C_GREY)

    # REVENUE
    bl_row("REVENUE", None, is_sec=True)
    bl_row("GV P1",[pvm[bl]['gv_start'] for bl in BLS5],'#,##0',bg_=C_WHITE)
    bl_row("GV P2",[B(pvm[bl]['gv_end'])   for bl in BLS5],'#,##0',bg_=C_GREY)
    gv_d=[pvm[bl]['gv_end']-pvm[bl]['gv_start'] for bl in BLS5]
    bl_row("GV Diff", [v for v in gv_d],'#,##0',bg_=C_WHITE)
    bl_row("GV Diff %",   [v/pvm[bl]['gv_start'] if pvm[bl]['gv_start']>0 else None
                           for v,bl in zip(gv_d,BLS5)],F_PCT,bg_=C_GREY)

    # COST
    bl_row("COST", None, is_sec=True)
    c1=[pvm[bl]['gv_start']-pvm[bl]['gp_start'] for bl in BLS5]
    c2=[pvm[bl]['gv_end']  -pvm[bl]['gp_end']   for bl in BLS5]
    bl_row("COGS P1",[v for v in c1],'#,##0',bg_=C_WHITE)
    bl_row("COGS P2",[v for v in c2],'#,##0',bg_=C_GREY)
    cd=[p2-p1 for p1,p2 in zip(c1,c2)]
    bl_row("COGS Diff %",[d/p1 if p1>0 else None for d,p1 in zip(cd,c1)],F_PCT,bg_=C_WHITE)

    # PROFITABILITY
    bl_row("PROFITABILITY", None, is_sec=True)
    bl_row("GP P1",[pvm[bl]['gp_start'] for bl in BLS5],'#,##0',bg_=C_WHITE,bold=True)
    bl_row("GP P2",[B(pvm[bl]['gp_end'])   for bl in BLS5],'#,##0',bg_=C_GREY,bold=True)
    gp_d=[pvm[bl]['gp_end']-pvm[bl]['gp_start'] for bl in BLS5]
    bl_row("GP Diff", [v for v in gp_d],'#,##0',bg_=C_WHITE,bold=True)
    bl_row("GP Diff %",   [v/pvm[bl]['gp_start'] if pvm[bl]['gp_start']>0 else None
                           for v,bl in zip(gp_d,BLS5)],F_PCT,bg_=C_GREY,bold=True)
    bl_row("Margin P1",[pvm[bl]['m_base'] for bl in BLS5],F_PCT,bg_=C_WHITE)
    bl_row("Margin P2",[pvm[bl]['m_end']   for bl in BLS5],F_PCT,bg_=C_GREY,bold=True)
    bl_row("Margin Change (pp)",[pvm[bl]['pp_total']*100 for bl in BLS5],
           '0.00"pp"',is_pp=True,bg_=C_WHITE)

    # PORTFOLIO
    bl_row("PORTFOLIO", None, is_sec=True)
    bl_row("# Existing",  [pvm[bl]['n_ex']    for bl in BLS5],'#,##0',bg_=C_WHITE)
    bl_row("# New",       [pvm[bl]['new_cnt'] for bl in BLS5],'#,##0',bg_="EBF5D0")
    bl_row("# Deprecated",[pvm[bl]['dep_cnt'] for bl in BLS5],'#,##0',bg_=C_OR)

    # MARGIN BRIDGE pp
    bl_row("MARGIN BRIDGE (pp)", None, is_sec=True)
    # Aggregate Existing SKU Effect = COGS + Price + Vol/Mix
    pp_existing_per_bl = {bl: (pvm[bl]['pp_cogs'] + pvm[bl]['pp_price'] + pvm[bl]['pp_volmix']) for bl in BLS5}
    bridge_rows = [
        ("1. Churned SKU Effect",   'pp_B',      C_RED,    None),
        ("2. Existing SKU Effect",  None,        "FFFDE7", pp_existing_per_bl),
        ("  2.1 COGS Effect",       'pp_cogs',   "FFF3E0", None),
        ("  2.2 Price Effect",      'pp_price',  "E3F2FD", None),
        ("  2.3 Vol/Mix Effect",    'pp_volmix', "EDE7F6", None),
        ("3. New SKU Effect",       'pp_G',      "E8F5E9", None),
    ]
    for label_, key_, bg__, agg_dict in bridge_rows:
        if agg_dict is not None:
            vals = [agg_dict[bl]*100 for bl in BLS5]
        else:
            vals = [pvm[bl][key_]*100 for bl in BLS5]
        bl_row(label_, vals, '0.00"pp"', is_pp=True, bg_=bg__)
    bl_row("Total Change",[pvm[bl]['pp_total']*100 for bl in BLS5],
           '0.00"pp"',is_pp=True,bg_=C_TOTAL,bold=True)


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 1 — Raw Data
# ══════════════════════════════════════════════════════════════════════════════
def write_s1(wb, df, p1, p2):
    progress("Sheet 1: Raw Data...", 1, 8)
    # Remove default empty sheet if it exists, then create fresh
    if 'Sheet' in [s.title for s in wb.worksheets]:
        del wb['Sheet']
    ws = wb.create_sheet("1. Raw Data")

    col_order = [c for c in [
        'period_1','period_2','product_id','product_name','pricing_bl','l1_category',
        'business_line','pareto_class','sku_status',
        'qty_p1','qty_p2','qty_diff','qty_diff_pct',
        'price_p1','price_p2','price_diff','price_diff_pct',
        'cogs_p1','cogs_p2','cogs_diff','cogs_diff_pct',
        'gv_p1','gv_p2','gv_diff','gv_diff_pct',
        'gp_p1','gp_p2','gp_diff','gp_diff_pct',
        'unit_margin_p1','unit_margin_p2','unit_margin_diff','unit_margin_diff_pct',
        'margin_p1','margin_p2','margin_diff',
        'comp_price_p1','comp_price_p2','comp_price_diff','comp_price_diff_pct',
        'pi_p1','pi_p2','pi_diff','pi_diff_pct',
        'avg_stock_p1','avg_stock_p2','avg_stock_diff','avg_stock_diff_pct',
        'cogs_effect_rp','price_effect_rp','vol_mix_effect_rp',
        'oos_flag','cogs_status','price_status','comp_status',
        'framework_check','flag_price',
        'gp_group_p1','gp_group_p2','pi_group_p1','pi_group_p2',
        'cogs_vs_comp_p1','cogs_vs_comp_p2','coverage_status_p1','coverage_status_p2',
    ] if c in df.columns]

    df_out = df[col_order].copy()
    ncols = len(col_order)

    title_row(ws, 1, ncols, f"RAW DATA — {p1} vs {p2}")
    ws.merge_cells(f'A2:{cl(ncols)}2')
    st(ws,2,1,"Green=New  |  Orange=Deprecated  |  period_1/period_2 columns hidden",
       italic=True,fc="595959",bg=C_LIGHT)

    pct_cols={c for c in col_order if 'pct' in c or c in ['margin_p1','margin_p2','margin_diff']}
    pt_cols ={'cogs_vs_comp_p1','cogs_vs_comp_p2'}
    num_cols={c for c in col_order if any(x in c for x in
              ['qty','gv','gp','effect','stock']) and 'pct' not in c}
    price_cols={c for c in col_order if any(x in c for x in
                ['price','cogs','margin_p','unit_margin']) and 'pct' not in c
                and 'status' not in c and 'group' not in c and 'vs' not in c}

    for i,col in enumerate(col_order,1):
        hdr(ws,3,i,col)

    # Bulk write
    df_w = df_out.copy()
    for col in df_w.columns:
        df_w[col] = df_w[col].where(pd.notna(df_w[col]), None)
    stat_idx = col_order.index('sku_status') if 'sku_status' in col_order else None

    for rv in df_w.itertuples(index=False, name=None):
        ws.append(list(rv))

    # Number formats
    for i,col in enumerate(col_order,1):
        fmt=None
        if col in pct_cols:    fmt=F_PCT
        elif col in pt_cols:   fmt='0.00'
        elif col in num_cols:  fmt=F_NUM
        elif col in price_cols:fmt='#,##0.00'
        if fmt:
            for ri in range(4,len(df_w)+4):
                ws.cell(row=ri,column=i).number_format=fmt

    # Row colors new/dep
    if stat_idx is not None:
        for ri in range(4,len(df_w)+4):
            stat=ws.cell(row=ri,column=stat_idx+1).value
            if stat in('New','Deprecated'):
                fill=PatternFill("solid",start_color="EBF5D0" if stat=='New' else C_OR)
                for ci in range(1,ncols+1):
                    ws.cell(row=ri,column=ci).fill=fill

    # Hide period cols
    ws.column_dimensions['A'].hidden=True
    ws.column_dimensions['B'].hidden=True
    ws.column_dimensions['C'].width=12
    ws.column_dimensions['D'].width=38
    ws.column_dimensions['E'].width=10
    ws.column_dimensions['F'].width=22
    ws.column_dimensions['G'].width=16
    ws.column_dimensions['H'].width=16
    ws.column_dimensions['I'].width=12
    for ci in range(10,ncols+1): ws.column_dimensions[cl(ci)].width=14
    ws.freeze_panes="E4"
    ws.row_dimensions[3].height=30  # header row taller so text wraps not diagonal

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 1b — Aggregates (hardcoded values, feeds Sheet 2 via Excel formulas)
# ══════════════════════════════════════════════════════════════════════════════
def write_s1b(wb, pvm, p1, p2):
    progress("Sheet 1b: Aggregates...", 2, 8)
    ws = wb.create_sheet("1b. Aggregates")

    for c,w in [('A',14),('B',20),('C',20),('D',20),('E',20),
                ('F',18),('G',18),('H',18),('I',18),('J',16),('K',16),('L',16),('M',16)]:
        ws.column_dimensions[c].width = w

    row = 1
    title_row(ws, row, 13, f"AGGREGATES — {p1} vs {p2}  |  Hardcoded values from Python. Sheet 2 formulas reference here.")
    row += 1
    ws.merge_cells(f'A{row}:M{row}')
    st(ws, row, 1, "Do not edit manually. All values recomputed when program runs.",
       italic=True, fc="595959", bg=C_LIGHT)
    row += 2

    AGG_ROWS = {}

    def write_section(title, headers, data_rows, section_key):
        nonlocal row
        sec(ws, row, len(headers), title)
        row += 1
        # ALL headers in ONE row — no diagonal
        for i, h_ in enumerate(headers, 1):
            hdr(ws, row, i, h_)
        ws.row_dimensions[row].height = 28
        row += 1
        for bl_key, vals, bg_, bld in data_rows:
            AGG_ROWS[f'{section_key}_{bl_key}'] = row
            for ci, v in enumerate(vals, 1):
                is_pct = ci >= len(headers) - 2 and '%' in headers[ci-1].lower() or                          'margin' in headers[ci-1].lower() or 'weight' in headers[ci-1].lower()
                is_num = ci > 1 and not is_pct
                fmt_ = F_PCT if is_pct else (F_NUM if is_num else None)
                st(ws, row, ci, v, bold=bld, bg=bg_, fmt=fmt_,
                   align="right" if ci > 1 else "left")
            row += 1
        row += 1

    # ── A: BL Summary All SKU ─────────────────────────────────────────────────
    hdrs_A = ["BL","GV P1","COGS P1","GP P1","Qty P1",
              "GV P2","COGS P2","GP P2","Qty P2",
              "Margin P1","Margin P2","GV Weight P2"]
    rows_A = []
    for bi, bl in enumerate(BLS+['TOTAL']):
        d = pvm[bl]
        bg_ = C_TOTAL if bl=='TOTAL' else (C_GREY if bi%2==0 else C_WHITE)
        bld = bl=='TOTAL'
        cs = d['gv_start'] - d['gp_start']
        ce = d['gv_end']   - d['gp_end']
        rows_A.append((bl,
            [bl, d['gv_start'], cs, d['gp_start'], d['n_ex'],
             d['gv_end'], ce, d['gp_end'], d['n_ex'],
             d['m_base'], d['m_end'], d['gv_w2']],
            bg_, bld))
    write_section("A. BL SUMMARY — All SKU (existing + new + dep)", hdrs_A, rows_A, 'A')

    # ── B: Existing Only ─────────────────────────────────────────────────────
    hdrs_B = ["BL","GV P1","COGS P1","GP P1","GV P2","COGS P2","GP P2",
              "Qty P1","Qty P2","Margin P1","Margin P2","GV Weight P2"]
    rows_B = []
    for bi, bl in enumerate(BLS+['TOTAL']):
        d = pvm[bl]
        bg_ = C_TOTAL if bl=='TOTAL' else (C_GREY if bi%2==0 else C_WHITE)
        bld = bl=='TOTAL'
        rows_B.append((bl,
            [bl, d['gv_ex1'], d['gv_ex1']-d['gp_ex1'], d['gp_ex1'],
             d['gv_ex2'], d['gv_ex2']-d['gp_ex2'], d['gp_ex2'],
             d['n_ex'], d['n_ex'], d['m_ex1'], d['m_ex2'], d['gv_w2']],
            bg_, bld))
    write_section("B. EXISTING SKU ONLY", hdrs_B, rows_B, 'B')

    # ── C: Hypotheticals ─────────────────────────────────────────────────────
    hdrs_C = ["BL","GV_hyp1 (q1×p1)","GV_hyp2 (q1×p2)","GP_hyp1","GP_hyp2",
              "M_hyp1","M_hyp2","pp_B","pp_COGS","pp_Price","pp_VolMix","pp_G"]
    rows_C = []
    for bi, bl in enumerate(BLS+['TOTAL']):
        d = pvm[bl]
        bg_ = C_TOTAL if bl=='TOTAL' else (C_GREY if bi%2==0 else C_WHITE)
        bld = bl=='TOTAL'
        gp_h1 = d['gp_ex1'] + d['cogs_rp']
        gp_h2 = gp_h1 + d['price_rp']
        rows_C.append((bl,
            [bl, d['gv_ex1'], d['gv_ex2'], gp_h1, gp_h2,
             d['m_h1'], d['m_h2'],
             d['pp_B'], d['pp_cogs'], d['pp_price'], d['pp_volmix'], d['pp_G']],
            bg_, bld))
    # override format for C — cols 2-5 are NUM, 6+ are PCT
    sec(ws, row, 12, "C. HYPOTHETICAL METRICS (for COGS→Price→Vol/Mix decomposition)")
    row += 1
    for i, h_ in enumerate(hdrs_C, 1):
        hdr(ws, row, i, h_)
    ws.row_dimensions[row].height = 28
    row += 1
    for bl_key, vals, bg_, bld in rows_C:
        AGG_ROWS[f'C_{bl_key}'] = row
        for ci, v in enumerate(vals, 1):
            fmt_ = F_NUM if ci in [2,3,4,5] else (F_PCT if ci >= 6 else None)
            st(ws, row, ci, v, bold=bld, bg=bg_, fmt=fmt_,
               align="right" if ci > 1 else "left")
        row += 1
    row += 1

    # ── D: PVM Components Rp ─────────────────────────────────────────────────
    hdrs_D = ["BL","GP Start","GP End","1. Churned SKU Effect","2. Existing SKU Effect",
              "  2.1 COGS Effect","  2.2 Price Effect","  2.3 Vol/Mix Effect","3. New SKU Effect","Total Change",
              "Churned Rp%","COGS Rp%","Price Rp%","Vol/Mix Rp%"]
    rows_D = []
    for bi, bl in enumerate(BLS+['TOTAL']):
        d = pvm[bl]
        bg_ = C_TOTAL if bl=='TOTAL' else (C_GREY if bi%2==0 else C_WHITE)
        bld = bl=='TOTAL'
        gs = d['gp_start']
        existing_rp = d['cogs_rp'] + d['price_rp'] + d['volmix_rp']
        rows_D.append((bl,
            [bl, gs, d['gp_end'], -d['gp_dep'], existing_rp, d['cogs_rp'], d['price_rp'],
             d['volmix_rp'], d['gp_new'], d['gp_end']-gs,
             -d['gp_dep']/gs if gs else 0,
             d['cogs_rp']/gs if gs else 0,
             d['price_rp']/gs if gs else 0,
             d['volmix_rp']/gs if gs else 0],
            bg_, bld))
    sec(ws, row, 14, "D. PVM COMPONENTS (Rp) — Churned → Existing(COGS→Price→Vol/Mix) → New SKU")
    row += 1
    for i, h_ in enumerate(hdrs_D, 1):
        hdr(ws, row, i, h_)
    ws.row_dimensions[row].height = 28
    row += 1
    for bl_key, vals, bg_, bld in rows_D:
        AGG_ROWS[f'D_{bl_key}'] = row
        for ci, v in enumerate(vals, 1):
            fmt_ = F_NUM if 2 <= ci <= 10 else (F_PCT if ci >= 11 else None)
            st(ws, row, ci, v, bold=bld, bg=bg_, fmt=fmt_,
               align="right" if ci > 1 else "left")
        row += 1
    row += 1

    # ── E: New & Dep Summary ─────────────────────────────────────────────────
    hdrs_E = ["BL","# New","GP New","GV New","Margin New",
              "# Dep","GP Dep","GV Dep","Margin Dep","Net GP"]
    rows_E = []
    for bi, bl in enumerate(BLS+['TOTAL']):
        d = pvm[bl]
        bg_ = C_TOTAL if bl=='TOTAL' else (C_GREY if bi%2==0 else C_WHITE)
        bld = bl=='TOTAL'
        nm = d['gp_new']/d['gv_new'] if d['gv_new'] > 0 else 0
        dm = d['gp_dep']/d['gv_dep'] if d['gv_dep'] > 0 else 0
        rows_E.append((bl,
            [bl, d['new_cnt'], d['gp_new'], d['gv_new'], nm,
             d['dep_cnt'], d['gp_dep'], d['gv_dep'], dm,
             d['gp_new']-d['gp_dep']],
            bg_, bld))
    sec(ws, row, 10, "E. NEW & DEPRECATED SUMMARY")
    row += 1
    for i, h_ in enumerate(hdrs_E, 1):
        hdr(ws, row, i, h_)
    ws.row_dimensions[row].height = 28
    row += 1
    for bl_key, vals, bg_, bld in rows_E:
        AGG_ROWS[f'E_{bl_key}'] = row
        for ci, v in enumerate(vals, 1):
            fmt_ = F_NUM if ci in [2,3,4,6,7,8,10] else (F_PCT if ci in [5,9] else None)
            st(ws, row, ci, v, bold=bld, bg=bg_, fmt=fmt_,
               align="right" if ci > 1 else "left")
        row += 1

    return AGG_ROWS


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 2 — Margin Bridge (formula-driven from 1b)
# ══════════════════════════════════════════════════════════════════════════════
def write_s2(wb, pvm, AGG, p1, p2):
    progress("Sheet 2: Margin Bridge...", 3, 8)
    ws = wb.create_sheet("2. Margin Bridge")
    A = "'1b. Aggregates'"

    # Columns: A=label, B=Dry, C=Fresh, D=Frozen, E=PL, F=Overall,
    #          G=Dry%, H=Fresh%, I=Frozen%, J=PL%, K=Overall%
    NCOLS = 11
    ws.column_dimensions['A'].width = 34
    for c in ['B','C','D','E','F']: ws.column_dimensions[c].width = 18
    for c in ['G','H','I','J','K']: ws.column_dimensions[c].width = 14

    def ref(sheet_name, col, row_num):
        """Clean reference — no leading = inside formulas"""
        return f"'{sheet_name}'!{cl(col)}{row_num}"

    row = 1
    title_row(ws, row, NCOLS,
        f"MARGIN BRIDGE — {p1} vs {p2}  |  1.Churned → 2.Existing (2.1 COGS → 2.2 Price → 2.3 Vol/Mix) → 3.New")
    row += 1
    ws.merge_cells(f'A{row}:{cl(NCOLS)}{row}')
    st(ws, row, 1,
       "GP Start = Existing+Dep  |  COGS & Price: q_P1 basis (cleanest pp)  |  "
       "Vol/Mix = residual  |  GP End = Existing+New  |  % cols = attribution of GP growth",
       italic=True, fc="595959", bg=C_LIGHT)
    row += 2

    rD = {bl: AGG[f'D_{bl}'] for bl in BLS+['TOTAL']}
    rA = {bl: AGG[f'A_{bl}'] for bl in BLS+['TOTAL']}
    rE = {bl: AGG[f'E_{bl}'] for bl in BLS+['TOTAL']}
    rC = {bl: AGG[f'C_{bl}'] for bl in BLS+['TOTAL']}
    rB = {bl: AGG[f'B_{bl}'] for bl in BLS+['TOTAL']}

    # ── Tabel 1A — GP Bridge Rp ───────────────────────────────────────────────
    sec(ws, row, NCOLS,
        "Tabel 1A — GP Bridge (Rp)  |  % cols = attribution of GP growth per BL")
    row += 1
    # Two-row header: row1 = BL names spanning Rp+%, row2 = Rp / %
    # Simpler: single row with explicit labels
    hdrs_1a = ["Komponen","Dry (Rp)","Fresh (Rp)","Frozen (Rp)","PL (Rp)","Overall (Rp)",
               "Dry (% P1 GP)","Fresh (% P1 GP)","Frozen (% P1 GP)","PL (% P1 GP)","Overall (% P1 GP)"]
    for i, h_ in enumerate(hdrs_1a, 1):
        hdr(ws, row, i, h_)
    ws.row_dimensions[row].height = 28
    row += 1

    bridge_rows_1a = [
        ("P1 GP — Starting Point",       "start", C_TOTAL, True),
        ("1. Churned SKU Effect",        "dep",   C_RED,   False),
        ("2. Existing SKU Effect",       "exist", "FFFDE7",False),
        ("  2.1 COGS Effect",             6,       "FFF3E0",False),
        ("  2.2 Price Effect",            7,       "E3F2FD",False),
        ("  2.3 Vol/Mix Effect",          8,       "EDE7F6",False),
        ("3. New SKU Effect",            "new",   "E8F5E9",False),
        ("Total Change",                  "total", C_TOTAL, True),
        ("P2 GP — Ending Point",         "end",   C_TOTAL, True),
    ]

    r_1a = {}
    for label, src, bg_, bld in bridge_rows_1a:
        st(ws, row, 1, label, bold=bld, bg=bg_)

        for ci, bl in enumerate(BLS+['TOTAL'], 2):
            rDbl = rD[bl]; rAbl = rA[bl]; rEbl = rE[bl]
            if   src == 'start': val_ref = f"='{A[1:-1]}'!D{rAbl}"
            elif src == 'dep':   val_ref = f"=-'{A[1:-1]}'!G{rEbl}"
            elif src == 'exist': val_ref = f"='{A[1:-1]}'!E{rDbl}"
            elif src == 'new':   val_ref = f"='{A[1:-1]}'!C{rEbl}"
            elif src == 'total': val_ref = f"='{A[1:-1]}'!J{rDbl}"
            elif src == 'end':   val_ref = f"='{A[1:-1]}'!H{rAbl}"
            else:                val_ref = f"='{A[1:-1]}'!{cl(src)}{rDbl}"
            st(ws, row, ci, val_ref, fmt=F_NUM, align="right", bold=bld, bg=bg_)

        # % of P1 GP columns (cols 7-11, one per BL + overall)
        for ci, bl in enumerate(BLS+['TOTAL'], 7):
            rDbl = rD[bl]; rAbl = rA[bl]; rEbl = rE[bl]
            den_ref = f"'{A[1:-1]}'!D{rAbl}"   # GP P1 for this BL
            if src in ('start', 'end'):
                st(ws, row, ci, "—", align="center", fc="888888", bg=bg_)
            else:
                if   src == 'dep':   num_ref = f"-'{A[1:-1]}'!G{rEbl}"
                elif src == 'exist': num_ref = f"'{A[1:-1]}'!E{rDbl}"
                elif src == 'new':   num_ref = f"'{A[1:-1]}'!C{rEbl}"
                elif src == 'total': num_ref = f"'{A[1:-1]}'!J{rDbl}"
                else:                num_ref = f"'{A[1:-1]}'!{cl(src)}{rDbl}"
                st(ws, row, ci,
                   f"=IF({den_ref}<>0,{num_ref}/{den_ref},0)",
                   fmt=F_PCT, align="right", bg=bg_)

        r_1a[label] = row; row += 1

    note_row(ws, row, NCOLS,
        "✓ P1 GP + Total Change = P2 GP  |  % cols sum to Total% per BL")
    row += 2

    # ── Tabel 1B — pp Bridge ──────────────────────────────────────────────────
    sec(ws, row, NCOLS,
        "Tabel 1B — Margin pp Bridge  |  Start = all P1 margin → End = all P2 margin")
    row += 1
    hdrs_1b = ["Step","Dry","Fresh","Frozen","PL","Overall","GV Reference","","","",""]
    for i, h_ in enumerate(hdrs_1b[:7], 1):
        hdr(ws, row, i, h_)
    ws.row_dimensions[row].height = 28
    row += 1

    pp_rows = [
        ("Margin P1 — Start (all SKU)", 'base',    C_TOTAL, True,  "GV P1 all (existing+dep)"),
        ("1. Churned SKU Effect",       'ppB',     C_RED,   False, "GV existing P1"),
        ("2. Existing SKU Effect",      'ppEx',    "FFFDE7",False, "Sum of 2.1+2.2+2.3"),
        ("  2.1 COGS Effect",           'ppCOGS',  "FFF3E0",False, "GV_hyp1 = q1×p1 (unchanged)"),
        ("  2.2 Price Effect",          'ppPrice', "E3F2FD",False, "GV_hyp2 = q1×p2"),
        ("  2.3 Vol/Mix Effect",        'ppVM',    "EDE7F6",False, "GV actual existing P2"),
        ("3. New SKU Effect",           'ppG',     "E8F5E9",False, "GV P2 all (existing+new)"),
        ("Total Margin Change",         'ppTot',   C_TOTAL, True,  "End - Start"),
        ("Margin P2 — End (all SKU)",   'end',     C_TOTAL, True,  "GV P2 all"),
    ]

    # Map pp key → column index in 1b Section C
    # C cols: 1=BL,2=GV_hyp1,3=GV_hyp2,4=GP_h1,5=GP_h2,6=M_hyp1,7=M_hyp2,
    #         8=pp_B,9=pp_COGS,10=pp_Price,11=pp_VolMix,12=pp_G
    pp_col = {'ppB':8,'ppCOGS':9,'ppPrice':10,'ppVM':11,'ppG':12}

    for label, src, bg_, bld, gv_ref in pp_rows:
        st(ws, row, 1, label, bold=bld, bg=bg_)
        for ci, bl in enumerate(BLS+['TOTAL'], 2):
            rCbl = rC[bl]; rAbl = rA[bl]
            if   src == 'base':   f = f"='{A[1:-1]}'!J{rAbl}"
            elif src == 'ppTot':  f = f"='{A[1:-1]}'!K{rAbl}-'{A[1:-1]}'!J{rAbl}"
            elif src == 'end':    f = f"='{A[1:-1]}'!K{rAbl}"
            elif src == 'ppEx':   f = f"='{A[1:-1]}'!{cl(9)}{rCbl}+'{A[1:-1]}'!{cl(10)}{rCbl}+'{A[1:-1]}'!{cl(11)}{rCbl}"
            else:                 f = f"='{A[1:-1]}'!{cl(pp_col[src])}{rCbl}"
            st(ws, row, ci, f, fmt=F_PCT, align="right", bold=bld, bg=bg_)
        st(ws, row, 7, gv_ref, italic=True, fc="595959", bg=bg_, align="left")
        # blank cols 8-11
        for ci in range(8, 12):
            st(ws, row, ci, None, bg=bg_)
        row += 1

    note_row(ws, row, NCOLS,
        "✓ Start + 1.Churned + 2.Existing (=2.1+2.2+2.3) + 3.New = End  |  "
        "COGS & Price pp cleanest (q_P1 basis)  |  Vol/Mix = residual, absorbs interaction term")
    row += 2

    # ── Tabel 2 — Within & BL Mix Effect ─────────────────────────────────────
    sec(ws, row, NCOLS,
        "Tabel 2 — Within Effect & BL Mix Effect  |  Σ Within + Σ BL Mix = Overall Δmargin")
    row += 1
    hdrs_2 = ["BL","Margin P1","Margin P2","Δ Margin","GV Wt P1","GV Wt P2",
              "Δ Weight","Within Effect","BL Mix Effect","",""]
    for i, h_ in enumerate(hdrs_2[:9], 1):
        hdr(ws, row, i, h_)
    ws.row_dimensions[row].height = 28
    row += 1

    r_t2_start = row
    for bi, bl in enumerate(BLS):
        d = pvm[bl]; rAbl = rA[bl]
        bg_ = C_GREY if bi%2==0 else C_WHITE
        st(ws, row, 1, bl, bg=bg_)
        st(ws, row, 2, f"='{A[1:-1]}'!J{rAbl}", fmt=F_PCT, align="right", bg=bg_)
        st(ws, row, 3, f"='{A[1:-1]}'!K{rAbl}", fmt=F_PCT, align="right", bg=bg_)
        st(ws, row, 4, f"=C{row}-B{row}", fmt=F_PCT, align="right", bg=bg_)
        st(ws, row, 5, d['gv_w1'], fmt=F_PCT, align="right", bg=bg_)
        st(ws, row, 6, d['gv_w2'], fmt=F_PCT, align="right", bg=bg_)
        st(ws, row, 7, f"=F{row}-E{row}", fmt=F_PCT, align="right", bg=bg_)
        st(ws, row, 8, f"=D{row}*F{row}", fmt=F_PCT, align="right", bg=bg_)
        st(ws, row, 9, f"=G{row}*B{row}", fmt=F_PCT, align="right", bg=bg_)
        row += 1

    rATot = rA['TOTAL']
    st(ws, row, 1, "TOTAL", bold=True, bg=C_TOTAL, bt=True)
    st(ws, row, 2, f"='{A[1:-1]}'!J{rATot}", fmt=F_PCT, align="right", bold=True, bg=C_TOTAL, bt=True)
    st(ws, row, 3, f"='{A[1:-1]}'!K{rATot}", fmt=F_PCT, align="right", bold=True, bg=C_TOTAL, bt=True)
    st(ws, row, 4, f"=C{row}-B{row}", fmt=F_PCT, align="right", bold=True, bg=C_TOTAL, bt=True)
    st(ws, row, 5, 1.0, fmt=F_PCT, align="right", bg=C_TOTAL, bt=True)
    st(ws, row, 6, 1.0, fmt=F_PCT, align="right", bg=C_TOTAL, bt=True)
    st(ws, row, 7, "—", align="center", fc="888888", bg=C_TOTAL, bt=True)
    st(ws, row, 8, f"=SUM(H{r_t2_start}:H{row-1})", fmt=F_PCT, align="right", bold=True, bg=C_TOTAL, bt=True)
    st(ws, row, 9, f"=SUM(I{r_t2_start}:I{row-1})", fmt=F_PCT, align="right", bold=True, bg=C_TOTAL, bt=True)
    row += 1
    note_row(ws, row, NCOLS,
        "✓ Σ Within (H total) + Σ BL Mix (I total) = Overall Δmargin (D total)")
    row += 2

    # ── Tabel 3 — Contribution per BL per Effect to Overall ──────────────────
    sec(ws, row, NCOLS,
        "Tabel 3 — Kontribusi per BL per Effect ke Overall pp  |  Effect × GV Weight P2")
    row += 1
    hdrs_3 = ["Effect","Dry × wt","Fresh × wt","Frozen × wt","PL × wt",
              "Sum via components","BL Mix Effect","Total Overall","","",""]
    for i, h_ in enumerate(hdrs_3[:8], 1):
        hdr(ws, row, i, h_)
    ws.row_dimensions[row].height = 28
    row += 1

    bl_rows_t2 = {BLS[i]: r_t2_start + i for i in range(len(BLS))}
    total_row_t2 = r_t2_start + len(BLS)

    effects_t3 = [
        ("Start Overall P1 Margin",     None,      C_TOTAL, True),
        ("1. Churned SKU Effect",       'ppB',     C_RED,   False),
        ("2. Existing SKU Effect",      'ppEx',    "FFFDE7",False),
        ("  2.1 COGS Effect",           'ppCOGS',  "FFF3E0",False),
        ("  2.2 Price Effect",          'ppPrice', "E3F2FD",False),
        ("  2.3 Vol/Mix Effect",        'ppVM',    "EDE7F6",False),
        ("3. New SKU Effect",           'ppG',     "E8F5E9",False),
        ("Σ Within (via components)",   "sum",     C_LIGHT, True),
        ("Within Effect (direct)",      "direct",  C_LIGHT, True),
        ("Residual (via comp - direct)","res",     "F9F9F9",False),
        ("BL Mix Effect",               "blmix",   C_AMBER, False),
        ("Total Overall",               "tot",     C_TOTAL, True),
        ("End Overall P2 Margin",       None,      C_TOTAL, True),
    ]

    r_t3 = {}
    for label, src, bg_, bld in effects_t3:
        st(ws, row, 1, label, bold=bld, bg=bg_)

        if src is None:
            for ci in range(2, 8):
                st(ws, row, ci, "—", align="center", fc="888888", bg=bg_)
            if label.startswith("Start"):
                st(ws, row, 8, f"='{A[1:-1]}'!J{rA['TOTAL']}",
                   fmt=F_PCT, align="right", bold=True, bg=bg_)
            else:
                st(ws, row, 8, f"='{A[1:-1]}'!K{rA['TOTAL']}",
                   fmt=F_PCT, align="right", bold=True, bg=bg_)

        elif src == 'sum':
            # Sum components excluding the aggregate "2. Existing SKU Effect" row to avoid double-count
            effect_rows = [r_t3.get(k) for k in
                ["1. Churned SKU Effect","  2.1 COGS Effect","  2.2 Price Effect",
                 "  2.3 Vol/Mix Effect","3. New SKU Effect"] if r_t3.get(k)]
            for ci in range(2, 6):
                if effect_rows:
                    parts = "+".join([f"{cl(ci)}{r}" for r in effect_rows])
                    st(ws, row, ci, f"={parts}", fmt=F_PCT, align="right", bold=True, bg=bg_)
            st(ws, row, 6, f"=SUM(B{row}:E{row})", fmt=F_PCT, align="right", bold=True, bg=bg_)
            st(ws, row, 7, "—", align="center", fc="888888", bg=bg_)
            st(ws, row, 8, f"=F{row}", fmt=F_PCT, align="right", bold=True, bg=bg_)

        elif src == 'direct':
            for ci, bl in enumerate(BLS, 2):
                bl_r = bl_rows_t2[bl]
                st(ws, row, ci, f"=H{bl_r}", fmt=F_PCT, align="right", bold=True, bg=bg_)
            st(ws, row, 6, f"=SUM(B{row}:E{row})", fmt=F_PCT, align="right", bold=True, bg=bg_)
            st(ws, row, 7, "—", align="center", fc="888888", bg=bg_)
            st(ws, row, 8, f"=F{row}", fmt=F_PCT, align="right", bold=True, bg=bg_)
            r_t3['direct'] = row

        elif src == 'res':
            sum_row = r_t3.get('Σ Within (via components)', row-2)
            dir_row = r_t3.get('direct', row-1)
            for ci in range(2, 7):
                st(ws, row, ci,
                   f"={cl(ci)}{sum_row}-{cl(ci)}{dir_row}",
                   fmt=F_PCT, align="right", italic=True, fc="888888", bg=bg_)
            st(ws, row, 7, "—", align="center", fc="888888", bg=bg_)
            st(ws, row, 8, f"=F{row}", fmt=F_PCT, align="right", italic=True, fc="888888", bg=bg_)

        elif src == 'blmix':
            for ci, bl in enumerate(BLS, 2):
                bl_r = bl_rows_t2[bl]
                st(ws, row, ci, f"=I{bl_r}", fmt=F_PCT, align="right", bg=bg_)
            st(ws, row, 6, "—", align="center", fc="888888", bg=bg_)
            st(ws, row, 7, f"=SUM(B{row}:E{row})", fmt=F_PCT, align="right", bg=bg_)
            st(ws, row, 8, f"=G{row}", fmt=F_PCT, align="right", bg=bg_)

        elif src == 'tot':
            dir_row  = r_t3.get('direct', row-3)
            blmix_row = row - 1
            for ci in range(2, 7):
                st(ws, row, ci, "—", align="center", fc="888888", bg=bg_)
            st(ws, row, 7, "—", align="center", fc="888888", bg=bg_)
            st(ws, row, 8,
               f"=H{dir_row}+G{blmix_row}",
               fmt=F_PCT, align="right", bold=True, bg=bg_)

        elif src == 'ppEx':
            # Aggregate Existing SKU Effect = COGS + Price + Vol/Mix (each × GV weight P2)
            for ci, bl in enumerate(BLS, 2):
                rCbl = rC[bl]; d = pvm[bl]
                st(ws, row, ci,
                   f"=('{A[1:-1]}'!{cl(pp_col['ppCOGS'])}{rCbl}+'{A[1:-1]}'!{cl(pp_col['ppPrice'])}{rCbl}+'{A[1:-1]}'!{cl(pp_col['ppVM'])}{rCbl})*{d['gv_w2']}",
                   fmt=F_PCT, align="right", bg=bg_)
            st(ws, row, 6, f"=SUM(B{row}:E{row})", fmt=F_PCT, align="right", bg=bg_)
            st(ws, row, 7, "—", align="center", fc="888888", bg=bg_)
            st(ws, row, 8, f"=F{row}", fmt=F_PCT, align="right", bg=bg_)

        else:
            # Effect rows: pp × weight per BL
            for ci, bl in enumerate(BLS, 2):
                rCbl = rC[bl]; d = pvm[bl]
                st(ws, row, ci,
                   f"='{A[1:-1]}'!{cl(pp_col[src])}{rCbl}*{d['gv_w2']}",
                   fmt=F_PCT, align="right", bg=bg_)
            st(ws, row, 6, f"=SUM(B{row}:E{row})", fmt=F_PCT, align="right", bg=bg_)
            st(ws, row, 7, "—", align="center", fc="888888", bg=bg_)
            st(ws, row, 8, f"=F{row}", fmt=F_PCT, align="right", bg=bg_)

        r_t3[label] = row; row += 1

    note_row(ws, row, NCOLS,
        "✓ Within (direct) + BL Mix = Total Overall = Actual Δmargin  |  "
        "Residual = aggregation artifact from sequential decomposition (expected, disclose)")
    row += 1


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 2b — Margin Bridge Split (+) and (-)
# ══════════════════════════════════════════════════════════════════════════════
def write_s2b(wb, df, pvm, AGG, p1, p2):
    progress("Sheet 2b: Margin Bridge (+/-) Split...", 0, 8)
    ws = wb.create_sheet("2b. Margin Bridge (+-)")
    A = "'1b. Aggregates'"
    NCOLS = 11

    ws.column_dimensions['A'].width = 34
    for c in ['B','C','D','E','F','G','H','I','J','K']:
        ws.column_dimensions[c].width = 16

    row = 1
    title_row(ws, row, NCOLS,
        f"MARGIN BRIDGE (pos/neg) SPLIT — {p1} vs {p2}  |  Each effect split into (+) positive and (-) negative contributors")
    row += 1
    ws.merge_cells(f'A{row}:{cl(NCOLS)}{row}')
    st(ws, row, 1,
       "(+) = SKUs contributing positively to each effect  |  (-) = SKUs contributing negatively  |  (+) + (-) = Net = Sheet 2 value",
       italic=True, fc="595959", bg=C_LIGHT)
    row += 2

    rA = {bl: AGG[f'A_{bl}'] for bl in BLS+['TOTAL']}
    rD = {bl: AGG[f'D_{bl}'] for bl in BLS+['TOTAL']}
    rE = {bl: AGG[f'E_{bl}'] for bl in BLS+['TOTAL']}
    rC = {bl: AGG[f'C_{bl}'] for bl in BLS+['TOTAL']}

    # ── Compute per-SKU splits ────────────────────────────────────────────────
    ex      = df[df['sku_status']=='Existing'].copy()
    new_sku = df[df['sku_status']=='New'].copy()
    dep_sku = df[df['sku_status']=='Deprecated'].copy()

    for d_ in [new_sku]:
        if 'gp_p2' not in d_.columns:
            d_['gp_p2'] = d_['qty_p2'] * (d_['price_p2'] - d_['cogs_p2'])
    for d_ in [dep_sku]:
        if 'gp_p1' not in d_.columns:
            d_['gp_p1'] = d_['qty_p1'] * (d_['price_p1'] - d_['cogs_p1'])

    ex['_cogs_eff']   = -(ex['qty_p1'] * (ex['cogs_p2']  - ex['cogs_p1']))
    ex['_price_eff']  =   ex['qty_p1'] * (ex['price_p2'] - ex['price_p1'])
    ex['_gp_hyp2']    =   ex['qty_p1'] * (ex['price_p2'] - ex['cogs_p2'])
    if 'gp_p2' not in ex.columns:
        ex['gp_p2'] = ex['qty_p2'] * (ex['price_p2'] - ex['cogs_p2'])
    ex['_volmix_eff'] =   ex['gp_p2'] - ex['_gp_hyp2']
    dep_sku['_dep_eff'] = -(dep_sku['qty_p1'] * (dep_sku['price_p1'] - dep_sku['cogs_p1']))
    new_sku['_new_eff'] =   new_sku['gp_p2']

    def spl(series):
        return series[series > 0].sum(), series[series < 0].sum()

    def bl_spl(bl, col, src='ex'):
        if bl == 'TOTAL':
            if src == 'ex':  return spl(ex[col])
            if src == 'dep': return spl(dep_sku['_dep_eff'])
            if src == 'new': return spl(new_sku['_new_eff'])
        else:
            if src == 'ex':  return spl(ex[ex['pricing_bl']==bl][col])
            if src == 'dep': return spl(dep_sku[dep_sku['pricing_bl']==bl]['_dep_eff'])
            if src == 'new': return spl(new_sku[new_sku['pricing_bl']==bl]['_new_eff'])

    sdata = {}
    for bl in BLS + ['TOTAL']:
        sdata[bl] = {
            'dep':    bl_spl(bl, None, 'dep'),
            'cogs':   bl_spl(bl, '_cogs_eff'),
            'price':  bl_spl(bl, '_price_eff'),
            'volmix': bl_spl(bl, '_volmix_eff'),
            'new':    bl_spl(bl, None, 'new'),
        }

    # ── Tabel 1A — GP Bridge Split ────────────────────────────────────────────
    sec(ws, row, NCOLS, "Tabel 1A — GP Bridge (Rp) Split (pos / neg)")
    row += 1

    # Row 1: BL headers merged over (+)(-)
    st(ws, row, 1, "Komponen", bold=True, bg=C_MID, fc="FFFFFF", align="center")
    ws.cell(row=row, column=1).border = Border(bottom=Side(style='thin'))
    for ci, bl in enumerate(BLS + ['Overall'], 0):
        col = 2 + ci*2
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+1)
        bg_h = C_DARK if bl == 'Overall' else C_MID
        c = ws.cell(row=row, column=col, value=bl)
        c.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        c.fill = PatternFill("solid", start_color=bg_h)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = Border(bottom=Side(style='thin'))
    ws.row_dimensions[row].height = 20
    row += 1

    # Row 2: (+) (-) sub-headers
    hdr(ws, row, 1, "Komponen", bg=C_MID)
    for ci in range(5):
        col = 2 + ci*2
        hdr(ws, row, col,   "(+)", bg="375623")
        hdr(ws, row, col+1, "(-)", bg="CC0000", fc="FFFFFF")
    ws.row_dimensions[row].height = 20
    row += 1

    bridge_defs = [
        ("P1 GP — Starting Point",       None,     C_TOTAL, True),
        ("1. Churned SKU Effect",        'dep',    C_RED,   False),
        ("2. Existing SKU Effect",       'exist',  "FFFDE7",False),
        ("  2.1 COGS Effect",            'cogs',   "FFF3E0",False),
        ("  2.2 Price Effect",           'price',  "E3F2FD",False),
        ("  2.3 Vol/Mix Effect",         'volmix', "EDE7F6",False),
        ("3. New SKU Effect",            'new',    "E8F5E9",False),
        ("Total Change",                 'total',  C_TOTAL, True),
        ("P2 GP — Ending Point",         None,     C_TOTAL, True),
    ]

    step_rows = []
    aggregate_rows = []  # rows that aggregate sub-items (excluded from total sum)
    for label, eff, bg_, bld in bridge_defs:
        st(ws, row, 1, label, bold=bld, bg=bg_)
        is_step = eff not in (None, 'total', 'exist')
        is_aggregate = eff == 'exist'

        for ci, bl in enumerate(BLS + ['TOTAL'], 0):
            col_pos = 2 + ci*2
            col_neg = col_pos + 1
            rAbl = rA[bl]

            if label.startswith("P1"):
                st(ws, row, col_pos, f"='{A[1:-1]}'!D{rAbl}", fmt=F_NUM, align="right", bold=True, bg=bg_)
                st(ws, row, col_neg, "—", align="center", fc="888888", bg=bg_)
            elif label.startswith("P2"):
                st(ws, row, col_pos, f"='{A[1:-1]}'!H{rAbl}", fmt=F_NUM, align="right", bold=True, bg=bg_)
                st(ws, row, col_neg, "—", align="center", fc="888888", bg=bg_)
            elif eff == 'exist':
                # Aggregate: sum of COGS + Price + Vol/Mix per SKU contributions
                p_c, n_c = sdata[bl]['cogs']
                p_p, n_p = sdata[bl]['price']
                p_v, n_v = sdata[bl]['volmix']
                st(ws, row, col_pos, p_c + p_p + p_v, fmt=F_NUM, align="right", bg=bg_, fc="007030")
                st(ws, row, col_neg, n_c + n_p + n_v, fmt=F_NUM, align="right", bg=bg_, fc="CC0000")
            elif eff == 'total':
                # Sum only step rows (exclude aggregate rows to avoid double-count)
                non_agg_steps = [r for r in step_rows if r not in aggregate_rows]
                if non_agg_steps:
                    pos_refs = "+".join([f"{cl(col_pos)}{r}" for r in non_agg_steps])
                    neg_refs = "+".join([f"{cl(col_neg)}{r}" for r in non_agg_steps])
                    st(ws, row, col_pos, f"={pos_refs}", fmt=F_NUM, align="right", bold=True, bg=bg_)
                    st(ws, row, col_neg, f"={neg_refs}", fmt=F_NUM, align="right", bold=True, bg=bg_)
            else:
                pos_v, neg_v = sdata[bl][eff]
                st(ws, row, col_pos, pos_v, fmt=F_NUM, align="right", bg=bg_, fc="007030")
                st(ws, row, col_neg, neg_v, fmt=F_NUM, align="right", bg=bg_, fc="CC0000")

        if is_step: step_rows.append(row)
        if is_aggregate:
            step_rows.append(row)
            aggregate_rows.append(row)
        row += 1

    note_row(ws, row, NCOLS,
        "✓ (+) + (-) per row = Net value in Sheet 2  |  Green = positive contributors  |  Red = negative contributors")
    row += 2

    # ── Tabel 1B — pp Bridge (same as Sheet 2, reference 1b) ─────────────────
    sec(ws, row, 7, "Tabel 1B — Margin pp Bridge (net — same as Sheet 2)")
    row += 1
    for i, h_ in enumerate(["Step","Dry","Fresh","Frozen","PL","Overall","GV Reference"], 1):
        hdr(ws, row, i, h_)
    ws.row_dimensions[row].height = 24
    row += 1

    pp_col_map = {'ppB': 8, 'ppCOGS': 9, 'ppPrice': 10, 'ppVM': 11, 'ppG': 12}
    pp_rows_2b = [
        ("Margin P1 — Start (all SKU)", 'base',    C_TOTAL, True,  "GV P1 all"),
        ("1. Churned SKU Effect",       'ppB',     C_RED,   False, "GV existing P1"),
        ("2. Existing SKU Effect",      'ppEx',    "FFFDE7",False, "Sum of 2.1+2.2+2.3"),
        ("  2.1 COGS Effect",           'ppCOGS',  "FFF3E0",False, "GV_hyp1 = q1×p1"),
        ("  2.2 Price Effect",          'ppPrice', "E3F2FD",False, "GV_hyp2 = q1×p2"),
        ("  2.3 Vol/Mix Effect",        'ppVM',    "EDE7F6",False, "GV actual existing P2"),
        ("3. New SKU Effect",           'ppG',     "E8F5E9",False, "GV P2 all"),
        ("Total Margin Change",         'ppTot',   C_TOTAL, True,  "End - Start"),
        ("Margin P2 — End (all SKU)",   'end',     C_TOTAL, True,  "GV P2 all"),
    ]

    for label, src, bg_, bld, gv_ref in pp_rows_2b:
        st(ws, row, 1, label, bold=bld, bg=bg_)
        for ci, bl in enumerate(BLS+['TOTAL'], 2):
            rCbl = rC[bl]; rAbl = rA[bl]
            if   src == 'base':   f = f"='{A[1:-1]}'!J{rAbl}"
            elif src == 'ppTot':  f = f"='{A[1:-1]}'!K{rAbl}-'{A[1:-1]}'!J{rAbl}"
            elif src == 'end':    f = f"='{A[1:-1]}'!K{rAbl}"
            elif src == 'ppEx':   f = f"='{A[1:-1]}'!{cl(9)}{rCbl}+'{A[1:-1]}'!{cl(10)}{rCbl}+'{A[1:-1]}'!{cl(11)}{rCbl}"
            else:                 f = f"='{A[1:-1]}'!{cl(pp_col_map[src])}{rCbl}"
            st(ws, row, ci, f, fmt=F_PCT, align="right", bold=bld, bg=bg_)
        st(ws, row, 7, gv_ref, italic=True, fc="595959", bg=bg_, align="left")
        row += 1

    note_row(ws, row, 7, "pp is a net margin ratio — splitting (+)/(-) per SKU would not be additive at portfolio level")


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 3 — GV Tier Analysis
# ══════════════════════════════════════════════════════════════════════════════
def write_s3(wb, df, p1, p2):
    progress("Sheet 3: GV Tier Analysis...", 4, 8)
    ws = wb.create_sheet("3. GV Tier Analysis")
    NCOLS=16
    for c,w in [(cl(i),14) for i in range(1,NCOLS+1)]:
        ws.column_dimensions[c].width=w
    ws.column_dimensions['A'].width=14

    row=1
    title_row(ws,row,NCOLS,f"GV TIER ANALYSIS — {p1} vs {p2}  |  Existing SKU Only"); row+=1
    ws.merge_cells(f'A{row}:{cl(NCOLS)}{row}')
    st(ws,row,1,"Ranked by GV P1 descending per BL. Tiers mutually exclusive (cumulative). COGS & Price: q_P1 basis. Vol/Mix = residual.",
       italic=True,fc="595959",bg=C_LIGHT); row+=2

    ex=df[df['sku_status']=='Existing'].copy()
    tier_order=['Top 10%','10-25%','25-50%','50-80%','Bottom 20%']

    def get_tiers(e):
        e=e.copy().sort_values('gv_p1',ascending=False)
        tot=e['gv_p1'].sum()
        if tot==0: return e.assign(tier='N/A')
        e['cum']=e['gv_p1'].cumsum()/tot
        def t(x):
            if x<=0.10: return 'Top 10%'
            if x<=0.25: return '10-25%'
            if x<=0.50: return '25-50%'
            if x<=0.80: return '50-80%'
            return 'Bottom 20%'
        e['tier']=e['cum'].apply(t)
        return e

    for bl in BLS:
        e_bl=get_tiers(ex[ex['pricing_bl']==bl].copy())
        if len(e_bl)==0: continue
        tot_gv=e_bl['gv_p1'].sum(); tot_n=len(e_bl)

        sec(ws,row,NCOLS,f"{bl} — {tot_n:,} Existing SKUs",bg=C_MID); row+=1

        # Section A
        ws.merge_cells(f'A{row}:{cl(NCOLS)}{row}')
        st(ws,row,1,"A. PVM by GV Tier",bold=True,bg=C_LIGHT,fc=C_DARK); row+=1
        for i,h_ in enumerate(["Tier","# SKU","GV Share",
                                 "GP P1","GP P2","Diff GP","Diff GP%",
                                 "COGS Rp","Price Rp","Vol/Mix Rp",
                                 "Margin P1","Margin P2","Diff Margin",
                                 "COGS pp","Price pp","Vol/Mix pp"],1):
            hdr(ws,row,i,h_)
        row+=1; ws.row_dimensions[row-1].height=24

        r_data_start=row
        n_tiers_written=0
        for ti,tier in enumerate(tier_order):
            t=e_bl[e_bl['tier']==tier]
            if len(t)==0: continue
            bg_=C_GREY if ti%2==0 else C_WHITE; n=len(t)
            gv1=t['gv_p1'].sum(); gv2=t['gv_p2'].sum()
            gp1=t['gp_p1'].sum(); gp2=t['gp_p2'].sum()
            gv_share=gv1/tot_gv if tot_gv>0 else 0
            # Effects q_P1 basis
            cogs_rp=-(t['qty_p1']*(t['cogs_p2']-t['cogs_p1'])).sum()
            price_rp=(t['qty_p1']*(t['price_p2']-t['price_p1'])).sum()
            gv_h1=gv1; cogs_h1=(t['qty_p1']*t['cogs_p2']).sum()
            gp_h1=gv_h1-cogs_h1; m_h1=gp_h1/gv_h1 if gv_h1>0 else 0
            gv_h2=(t['qty_p1']*t['price_p2']).sum()
            gp_h2=gv_h2-cogs_h1; m_h2=gp_h2/gv_h2 if gv_h2>0 else 0
            volmix_rp=gp2-gp_h2
            m1=gp1/gv1 if gv1>0 else 0; m2=gp2/gv2 if gv2>0 else 0
            pp_cogs=m_h1-m1; pp_price=m_h2-m_h1; pp_volmix=m2-m_h2

            st(ws,row,1,tier,bg=bg_)
            for ci,v in enumerate([n,gv_share,gp1,gp2,gp2-gp1,
                                    (gp2-gp1)/gp1 if gp1 else 0,
                                    cogs_rp,price_rp,volmix_rp,m1,m2,m2-m1,
                                    pp_cogs,pp_price,pp_volmix],2):
                fmt_=F_PCT if ci in[3,7,11,12,13,14,15,16] else F_NUM
                st(ws,row,ci,v,fmt=fmt_,align="right",bg=bg_)
            n_tiers_written+=1; row+=1

        # Total row with SUM formula
        r_end=row-1
        st(ws,row,1,"TOTAL",bold=True,bg=C_TOTAL,bt=True)
        st(ws,row,2,f"=SUM(B{r_data_start}:B{r_end})",fmt=F_NUM,align="right",bold=True,bg=C_TOTAL,bt=True)
        st(ws,row,3,f"=1",fmt=F_PCT,align="right",bg=C_TOTAL,bt=True)
        for ci in range(4,17):
            if ci not in[7,11,12,13,14,15,16]:
                st(ws,row,ci,f"=SUM({cl(ci)}{r_data_start}:{cl(ci)}{r_end})",
                   fmt=F_NUM,align="right",bold=True,bg=C_TOTAL,bt=True)
            elif ci==7:
                gp1_tot=e_bl['gp_p1'].sum(); gp2_tot=e_bl['gp_p2'].sum()
                st(ws,row,ci,(gp2_tot-gp1_tot)/gp1_tot if gp1_tot else 0,fmt=F_PCT,align="right",bold=True,bg=C_TOTAL,bt=True)
            else:
                st(ws,row,ci,"—",align="center",fc="888888",bg=C_TOTAL,bt=True)
        row+=2

        # Section B — COGS Movement
        ws.merge_cells(f'A{row}:{cl(11)}{row}')
        st(ws,row,1,"B. COGS Movement by GV Tier",bold=True,bg=C_LIGHT,fc=C_DARK); row+=1
        for i,h_ in enumerate(["Tier","# SKU","# Up","% Up","# Down","% Down","# Flat","% Flat","Avg COGS Chg%","GV-Wtd COGS%","Concentration"],1):
            hdr(ws,row,i,h_)
        row+=1; ws.row_dimensions[row-1].height=24

        for ti,tier in enumerate(tier_order):
            t=e_bl[e_bl['tier']==tier]
            if len(t)==0: continue
            bg_=C_GREY if ti%2==0 else C_WHITE; nt=len(t)
            nu=(t['cogs_p2']>t['cogs_p1']).sum()
            nd=(t['cogs_p2']<t['cogs_p1']).sum()
            nf=(t['cogs_p2']==t['cogs_p1']).sum()
            ac=t['cogs_diff_pct'].mean()
            wc=(t['cogs_diff_pct']*t['gv_p1']).sum()/t['gv_p1'].sum() if t['gv_p1'].sum()>0 else 0
            top20=t.nlargest(max(1,int(nt*0.2)),'gv_p1')
            ti_=abs((t['qty_p1']*t['cogs_diff']).sum())
            t20=abs((top20['qty_p1']*top20['cogs_diff']).sum())
            conc="Concentrated" if ti_>0 and t20/ti_>0.6 else "Widespread"
            st(ws,row,1,tier,bg=bg_)
            for ci,v in enumerate([nt,nu,nu/nt if nt>0 else 0,nd,nd/nt if nt>0 else 0,
                                    nf,nf/nt if nt>0 else 0,ac,wc],2):
                st(ws,row,ci,v,fmt=F_PCT if ci in[4,6,8,9,10] else F_NUM,align="right",bg=bg_)
            st(ws,row,11,conc,align="center",bg=bg_,
               fc="CC0000" if conc=="Concentrated" else "007030")
            row+=1
        row+=2

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 4 — L1 Category Analysis
# ══════════════════════════════════════════════════════════════════════════════
def write_s4(wb, df, p1, p2):
    progress("Sheet 4: L1 Category Analysis...", 5, 8)
    ws = wb.create_sheet("4. L1 Category Analysis")
    NCOLS=17
    ws.column_dimensions['A'].width=28
    for c in [cl(i) for i in range(2,NCOLS+1)]: ws.column_dimensions[c].width=14

    row=1
    title_row(ws,row,NCOLS,f"L1 CATEGORY ANALYSIS — {p1} vs {p2}  |  Existing SKU Only"); row+=1
    ws.merge_cells(f'A{row}:{cl(NCOLS)}{row}')
    st(ws,row,1,"Existing SKUs only. COGS & Price pp: q_P1 basis. Sum COGS pp + Price pp + Vol/Mix pp = Diff Margin ✓",
       italic=True,fc="595959",bg=C_LIGHT); row+=2

    ex=df[df['sku_status']=='Existing'].copy()

    def l1_calc(grp):
        gv1=grp['gv_p1'].sum(); gv2=grp['gv_p2'].sum()
        gp1=grp['gp_p1'].sum(); gp2=grp['gp_p2'].sum()
        cogs_rp=-(grp['qty_p1']*(grp['cogs_p2']-grp['cogs_p1'])).sum()
        price_rp=(grp['qty_p1']*(grp['price_p2']-grp['price_p1'])).sum()
        gv_h1=gv1; cogs_h1=(grp['qty_p1']*grp['cogs_p2']).sum()
        gp_h1=gv_h1-cogs_h1; m_h1=gp_h1/gv_h1 if gv_h1>0 else 0
        gv_h2=(grp['qty_p1']*grp['price_p2']).sum()
        gp_h2=gv_h2-cogs_h1; m_h2=gp_h2/gv_h2 if gv_h2>0 else 0
        volmix_rp=gp2-gp_h2
        m1=gp1/gv1 if gv1>0 else 0; m2=gp2/gv2 if gv2>0 else 0
        return pd.Series({
            'n':len(grp),'gp1':gp1,'gp2':gp2,'gv1':gv1,'gv2':gv2,
            'diff_gp':gp2-gp1,'diff_gp_pct':(gp2-gp1)/gp1 if gp1 else 0,
            'm1':m1,'m2':m2,'diff_m':m2-m1,
            'cogs_rp':cogs_rp,'price_rp':price_rp,'volmix_rp':volmix_rp,
            'pp_cogs':m_h1-m1,'pp_price':m_h2-m_h1,'pp_volmix':m2-m_h2,
        })

    g=ex.groupby('l1_category').apply(l1_calc,include_groups=False).reset_index()
    g['driver']=g.apply(lambda r:'COGS' if abs(r['pp_cogs'])>=abs(r['pp_price']) and abs(r['pp_cogs'])>=abs(r['pp_volmix']) else('Price' if abs(r['pp_price'])>=abs(r['pp_volmix']) else 'Vol/Mix'),axis=1)
    g=g.sort_values('diff_gp',ascending=False)

    # Section A
    sec(ws,row,NCOLS,"A. L1 CATEGORY SCORECARD — Sorted by GP Impact Descending"); row+=1
    for i,h_ in enumerate(["L1 Category","# SKU","GP P1","GP P2","Diff GP","Diff GP%",
                             "Margin P1","Margin P2","Diff Margin",
                             "COGS Rp","COGS pp","Price Rp","Price pp",
                             "Vol/Mix Rp","Vol/Mix pp","Driver","Validation"],1):
        hdr(ws,row,i,h_)
    row+=1; ws.row_dimensions[row-1].height=24

    r_data=row
    for i,(_,r) in enumerate(g.iterrows()):
        bg_=C_GREY if i%2==0 else C_WHITE
        st(ws,row,1,r['l1_category'],bg=bg_)
        st(ws,row,2,int(r['n']),fmt=F_NUM,align="right",bg=bg_)
        st(ws,row,3,r['gp1'],fmt=F_NUM,align="right",bg=bg_)
        st(ws,row,4,r['gp2'],fmt=F_NUM,align="right",bg=bg_)
        fc_=("007030" if r['diff_gp']>=0 else "CC0000")
        st(ws,row,5,r['diff_gp'],fmt=F_NUM,align="right",bg=bg_,fc=fc_)
        st(ws,row,6,r['diff_gp_pct'],fmt=F_PCT,align="right",bg=bg_,fc=fc_)
        st(ws,row,7,r['m1'],fmt=F_PCT,align="right",bg=bg_)
        st(ws,row,8,r['m2'],fmt=F_PCT,align="right",bg=bg_)
        fc_m=("007030" if r['diff_m']>=0 else "CC0000")
        st(ws,row,9,r['diff_m'],fmt=F_PCT,align="right",bg=bg_,fc=fc_m)
        st(ws,row,10,r['cogs_rp'],fmt=F_NUM,align="right",bg=bg_)
        st(ws,row,11,r['pp_cogs'],fmt=F_PCT,align="right",bg=bg_)
        st(ws,row,12,r['price_rp'],fmt=F_NUM,align="right",bg=bg_)
        st(ws,row,13,r['pp_price'],fmt=F_PCT,align="right",bg=bg_)
        st(ws,row,14,r['volmix_rp'],fmt=F_NUM,align="right",bg=bg_)
        st(ws,row,15,r['pp_volmix'],fmt=F_PCT,align="right",bg=bg_)
        st(ws,row,16,r['driver'],align="center",bg=bg_,
           fc="2F5496" if r['driver']=='Vol/Mix' else("CC0000" if r['driver']=='COGS' else "C55A11"),bold=True)
        # Validation: sum pp = diff_m
        check=abs(r['pp_cogs']+r['pp_price']+r['pp_volmix']-r['diff_m'])
        st(ws,row,17,"✓" if check<1e-8 else "✗",align="center",bg=bg_,
           fc="007030" if check<1e-8 else "CC0000",bold=True)
        row+=1

    st(ws,row,1,"TOTAL",bold=True,bg=C_TOTAL,bt=True)
    for ci in [2,3,4,5,10,12,14]:
        st(ws,row,ci,f"=SUM({cl(ci)}{r_data}:{cl(ci)}{row-1})",
           fmt=F_NUM,align="right",bold=True,bg=C_TOTAL,bt=True)
    row+=2

    # Section B — COGS Pressure
    sec(ws,row,12,"B. COGS PRESSURE PER L1 CATEGORY"); row+=1
    for i,h_ in enumerate(["L1 Category","# SKU","# Up","% Up","# Down","% Down",
                             "# Flat","% Flat","Avg COGS Chg%","GV-Wtd COGS%","COGS pp","Concentration"],1):
        hdr(ws,row,i,h_)
    row+=1; ws.row_dimensions[row-1].height=24

    # Sort by GV-Wtd COGS Chg% descending (most impactful first)
    l1_cogs_stats = []
    for cat, grp in ex.groupby('l1_category'):
        nt=len(grp)
        wc=(grp['cogs_diff_pct']*grp['gv_p1']).sum()/grp['gv_p1'].sum() if grp['gv_p1'].sum()>0 else 0
        l1_cogs_stats.append((cat, grp, wc))
    l1_cogs_stats.sort(key=lambda x: x[2], reverse=True)
    for i,(cat,grp,_) in enumerate(l1_cogs_stats):
        bg_=C_GREY if i%2==0 else C_WHITE; nt=len(grp)
        nu=(grp['cogs_p2']>grp['cogs_p1']).sum()
        nd=(grp['cogs_p2']<grp['cogs_p1']).sum()
        nf=(grp['cogs_p2']==grp['cogs_p1']).sum()
        ac=grp['cogs_diff_pct'].mean()
        wc=(grp['cogs_diff_pct']*grp['gv_p1']).sum()/grp['gv_p1'].sum() if grp['gv_p1'].sum()>0 else 0
        top20=grp.nlargest(max(1,int(nt*0.2)),'gv_p1')
        ti_=abs((grp['qty_p1']*grp['cogs_diff']).sum())
        t20=abs((top20['qty_p1']*top20['cogs_diff']).sum())
        conc="Concentrated" if ti_>0 and t20/ti_>0.6 else "Widespread"
        pp_c=g[g['l1_category']==cat]['pp_cogs'].values[0] if cat in g['l1_category'].values else 0
        st(ws,row,1,cat,bg=bg_)
        for ci,v in enumerate([nt,nu,nu/nt if nt else 0,nd,nd/nt if nt else 0,
                                nf,nf/nt if nt else 0,ac,wc,pp_c],2):
            st(ws,row,ci,v,fmt=F_PCT if ci in[4,6,8,9,10,11] else F_NUM,align="right",bg=bg_)
        st(ws,row,12,conc,align="center",bg=bg_,
           fc="CC0000" if conc=="Concentrated" else "007030")
        row+=1
    row+=2

    # Section C — Growth Quality
    sec(ws,row,7,"C. GROWTH QUALITY"); row+=1
    for i,h_ in enumerate(["L1 Category","Vol Growth%","Diff Margin","COGS pp","Price pp","Vol/Mix pp","Growth Quality"],1):
        hdr(ws,row,i,h_)
    row+=1; ws.row_dimensions[row-1].height=24

    for i,(_,r) in enumerate(g.iterrows()):
        bg_=C_GREY if i%2==0 else C_WHITE
        vol_g=ex[ex['l1_category']==r['l1_category']]['qty_diff_pct'].mean()
        if pd.isna(vol_g): vol_g=0
        if r['diff_m']>=-0.005 and vol_g>=0:
            if abs(r['pp_volmix'])>abs(r['pp_cogs']) and abs(r['pp_volmix'])>abs(r['pp_price']):
                q,qfc="Healthy — Mix Driven","007030"
            else: q,qfc="Healthy","007030"
        elif vol_g<0: q,qfc="Shrinking","CC0000"
        else:
            if r['pp_cogs']<-0.005 and r['pp_price']<-0.005: q,qfc="Dilutive — Cost+Price","C55A11"
            elif r['pp_cogs']<-0.005: q,qfc="Dilutive — Cost Driven","C55A11"
            elif r['pp_price']<-0.005: q,qfc="Dilutive — Price Driven","C55A11"
            else: q,qfc="Dilutive","C55A11"
        st(ws,row,1,r['l1_category'],bg=bg_)
        st(ws,row,2,vol_g,fmt=F_PCT,align="right",bg=bg_)
        st(ws,row,3,r['diff_m'],fmt=F_PCT,align="right",bg=bg_)
        st(ws,row,4,r['pp_cogs'],fmt=F_PCT,align="right",bg=bg_)
        st(ws,row,5,r['pp_price'],fmt=F_PCT,align="right",bg=bg_)
        st(ws,row,6,r['pp_volmix'],fmt=F_PCT,align="right",bg=bg_)
        st(ws,row,7,q,align="center",bg=bg_,fc=qfc,bold=True)
        row+=1
    row+=2

    # Section D — Top 10 COGS pp most negative
    sec(ws,row,17,"D. TOP 10 — COGS pp MOST NEGATIVE",bg="375623"); row+=1
    for i,h_ in enumerate(["L1 Category","# SKU","GP P1","GP P2","Diff GP","Diff GP%",
                             "Margin P1","Margin P2","Diff Margin",
                             "COGS Rp","COGS pp","Price Rp","Price pp",
                             "Vol/Mix Rp","Vol/Mix pp","# COGS Up","% COGS Up"],1):
        hdr(ws,row,i,h_,bg="375623")
    row+=1; ws.row_dimensions[row-1].height=24

    top10c=g.nsmallest(10,'pp_cogs')
    for i,(_,r) in enumerate(top10c.iterrows()):
        bg_=C_GREY if i%2==0 else C_WHITE
        grp=ex[ex['l1_category']==r['l1_category']]
        nu=(grp['cogs_p2']>grp['cogs_p1']).sum(); nt=len(grp)
        st(ws,row,1,r['l1_category'],bg=bg_)
        for ci,v in enumerate([int(r['n']),r['gp1'],r['gp2'],r['diff_gp'],r['diff_gp_pct'],
                                r['m1'],r['m2'],r['diff_m'],
                                r['cogs_rp'],r['pp_cogs'],r['price_rp'],r['pp_price'],
                                r['volmix_rp'],r['pp_volmix'],nu,nu/nt if nt else 0],2):
            fmt_=F_PCT if ci in[6,7,8,9,11,13,15,17] else F_NUM
            st(ws,row,ci,v,fmt=fmt_,align="right",bg=bg_)
        row+=1
    row+=2

    # Section E — Top 10 Price pp most negative
    sec(ws,row,17,"E. TOP 10 — PRICE pp MOST NEGATIVE",bg="7F3F00"); row+=1
    for i,h_ in enumerate(["L1 Category","# SKU","GP P1","GP P2","Diff GP","Diff GP%",
                             "Margin P1","Margin P2","Diff Margin",
                             "COGS Rp","COGS pp","Price Rp","Price pp",
                             "Vol/Mix Rp","Vol/Mix pp","Vol Growth%","Context Flag"],1):
        hdr(ws,row,i,h_,bg="C55A11")
    row+=1; ws.row_dimensions[row-1].height=24

    top10p=g.nsmallest(10,'pp_price')
    for i,(_,r) in enumerate(top10p.iterrows()):
        bg_=C_GREY if i%2==0 else C_WHITE
        grp=ex[ex['l1_category']==r['l1_category']]
        vol_g=grp['qty_diff_pct'].mean()
        pi_dn=(grp['pi_p2']<grp['pi_p1']).mean() if 'pi_p2' in grp.columns else 0
        if not pd.isna(vol_g) and vol_g>0.10: ctx="Intentional Promo"
        elif pi_dn>0.5: ctx="Comp Pressure"
        else: ctx="Price Erosion"
        st(ws,row,1,r['l1_category'],bg=bg_)
        for ci,v in enumerate([int(r['n']),r['gp1'],r['gp2'],r['diff_gp'],r['diff_gp_pct'],
                                r['m1'],r['m2'],r['diff_m'],
                                r['cogs_rp'],r['pp_cogs'],r['price_rp'],r['pp_price'],
                                r['volmix_rp'],r['pp_volmix'],
                                vol_g if not pd.isna(vol_g) else 0],2):
            fmt_=F_PCT if ci in[6,7,8,9,11,13,15,16] else F_NUM
            st(ws,row,ci,v,fmt=fmt_,align="right",bg=bg_)
        ctx_colors={"Intentional Promo":"0C447C","Price Erosion":"7F4000","Comp Pressure":"4B0082"}
        ctx_bgs={"Intentional Promo":"E6F0FB","Price Erosion":"FFF4E6","Comp Pressure":"F5F0FF"}
        st(ws,row,17,ctx,align="center",bg=ctx_bgs.get(ctx,C_WHITE),
           fc=ctx_colors.get(ctx,"000000"),bold=True)
        row+=1

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 5 — New & Deprecated Analysis
# ══════════════════════════════════════════════════════════════════════════════
def write_s5(wb, df, p1, p2):
    progress("Sheet 5: New & Dep Analysis...", 6, 8)
    ws = wb.create_sheet("5. New & Dep Analysis")
    ws.column_dimensions['A'].width=12
    ws.column_dimensions['B'].width=38
    for c in [cl(i) for i in range(3,15)]: ws.column_dimensions[c].width=14

    row=1
    title_row(ws,row,14,f"NEW & DEPRECATED PRODUCTS — {p1} vs {p2}"); row+=1

    new=df[df['sku_status']=='New'].copy()
    dep=df[df['sku_status']=='Deprecated'].copy()

    new['gv_new']=(new['qty_p2']*new['price_p2'])
    new['gp_new']=(new['qty_p2']*(new['price_p2']-new['cogs_p2']))
    dep['gv_dep']=(dep['qty_p1']*dep['price_p1'])
    dep['gp_dep']=(dep['qty_p1']*(dep['price_p1']-dep['cogs_p1']))

    # Summary
    sec(ws,row,8,"SUMMARY — Overall"); row+=1
    for i,h_ in enumerate(["","# SKU","GV","GP","Margin"],1):
        hdr(ws,row,i,h_)
    row+=1

    new_gv=new['gv_new'].sum(); new_gp=new['gp_new'].sum()
    dep_gv=dep['gv_dep'].sum(); dep_gp=dep['gp_dep'].sum()
    new_m=new_gp/new_gv if new_gv>0 else 0
    dep_m=dep_gp/dep_gv if dep_gv>0 else 0

    for label,n,gv,gp,m,bg_ in [
        ("New SKU",len(new),new_gv,new_gp,new_m,"EBF5D0"),
        ("Deprecated SKU",len(dep),dep_gv,dep_gp,dep_m,C_OR),
        ("Net (New − Dep)",len(new)-len(dep),new_gv-dep_gv,new_gp-dep_gp,None,C_TOTAL),
    ]:
        st(ws,row,1,label,bold=label.startswith("Net"),bg=bg_)
        st(ws,row,2,n,fmt=F_NUM,align="right",bg=bg_,bold=label.startswith("Net"))
        st(ws,row,3,gv,fmt=F_NUM,align="right",bg=bg_,bold=label.startswith("Net"))
        st(ws,row,4,gp,fmt=F_NUM,align="right",bg=bg_,bold=label.startswith("Net"))
        st(ws,row,5,m if m else "—",fmt=F_PCT if m else None,align="right",bg=bg_)
        row+=1
    row+=2

    # New SKU detail
    sec(ws,row,11,"NEW PRODUCTS (P2 only)",bg="375623"); row+=1
    nd_cols=['product_id','product_name','pricing_bl','l1_category',
             'qty_p2','price_p2','cogs_p2','gv_new','gp_new','margin_p2']
    new['margin_p2']=new['gp_new']/new['gv_new']
    new['Potential GP Contribution']=new['gp_new']
    nd_cols_ext=nd_cols+['Potential GP Contribution']
    for i,h_ in enumerate(nd_cols_ext,1):
        hdr(ws,row,i,h_,bg="375623")
    row+=1; ws.row_dimensions[row-1].height=24
    for i,(_,r) in enumerate(new.sort_values('gp_new',ascending=False).iterrows()):
        bg_="EBF5D0" if i%2==0 else C_WHITE
        st(ws,row,1,r.get('product_id',''),bg=bg_)
        st(ws,row,2,r.get('product_name',''),bg=bg_,align="left")
        st(ws,row,3,r.get('pricing_bl',''),bg=bg_,align="center")
        st(ws,row,4,r.get('l1_category',''),bg=bg_)
        for ci,col in enumerate(['qty_p2','price_p2','cogs_p2','gv_new','gp_new'],5):
            st(ws,row,ci,r.get(col,0),fmt=F_NUM,align="right",bg=bg_)
        st(ws,row,10,r.get('margin_p2',0),fmt=F_PCT,align="right",bg=bg_)
        st(ws,row,11,r.get('gp_new',0),fmt=F_NUM,align="right",bg=bg_)
        row+=1
    row+=2

    # Deprecated SKU detail
    sec(ws,row,11,"DEPRECATED PRODUCTS (P1 only)",bg="7F3F00"); row+=1
    dep['margin_p1_dep']=dep['gp_dep']/dep['gv_dep']
    dep['Potential GP Loss']=dep['gp_dep']
    dep_cols=['product_id','product_name','pricing_bl','l1_category',
              'qty_p1','price_p1','cogs_p1','gv_dep','gp_dep','margin_p1_dep','Potential GP Loss']
    for i,h_ in enumerate(dep_cols,1):
        hdr(ws,row,i,h_,bg="C55A11")
    row+=1; ws.row_dimensions[row-1].height=24
    for i,(_,r) in enumerate(dep.sort_values('gp_dep',ascending=False).iterrows()):
        bg_=C_OR if i%2==0 else "FFF0E6"
        st(ws,row,1,r.get('product_id',''),bg=bg_)
        st(ws,row,2,r.get('product_name',''),bg=bg_,align="left")
        st(ws,row,3,r.get('pricing_bl',''),bg=bg_,align="center")
        st(ws,row,4,r.get('l1_category',''),bg=bg_)
        for ci,col in enumerate(['qty_p1','price_p1','cogs_p1','gv_dep','gp_dep'],5):
            st(ws,row,ci,r.get(col,0),fmt=F_NUM,align="right",bg=bg_)
        st(ws,row,10,r.get('margin_p1_dep',0),fmt=F_PCT,align="right",bg=bg_)
        st(ws,row,11,r.get('gp_dep',0),fmt=F_NUM,align="right",bg=bg_)
        row+=1

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 6 — COGS vs Comp Price Analysis  (summary only — no SKU list)
# SHEET 6b — COGS vs Comp Price SKU Detail (bermasalah: coverage D or E)
# ══════════════════════════════════════════════════════════════════════════════
def write_s6(wb, df, p1, p2):
    progress("Sheet 6: COGS vs Comp Price...", 7, 8)

    ex = df[(df['sku_status']=='Existing') & df['cogs_vs_comp_p2'].notna()].copy()
    cov_buckets = ["(A) < 70","(B) 70-85","(C) 85-95","(D) 95-105","(E) > 105"]
    gp_buckets  = ["(A) < -20%","(B) -20% to -10%","(C) -10% to 0%","(D) 0% to 10%",
                   "(E) 10% to 20%","(F) 20% to 30%","(G) 30% to 50%","(H) > 50%"]
    BLS_ALL = BLS + ['TOTAL']

    # ── Sheet 6: Summary ─────────────────────────────────────────────────────
    ws = wb.create_sheet("6. COGS vs Comp Price")
    ws.column_dimensions['A'].width = 16
    for c in ['B','C','D','E','F','G','H','I','J','K','L']:
        ws.column_dimensions[c].width = 13

    row = 1
    title_row(ws, row, 12,
        f"COGS vs COMP PRICE ANALYSIS — {p1} vs {p2}  |  Coverage = COGS × 100 / Comp Price")
    row += 1
    ws.merge_cells(f'A{row}:L{row}')
    st(ws, row, 1,
       "(A)<70=room besar  |  (B)70-85=sehat  |  (C)85-95=watch  |  (D)95-105=at risk  |  (E)>105=critical  |  SKU bermasalah (D+E) → Sheet 6b",
       italic=True, fc="595959", bg=C_LIGHT)
    row += 2

    # Section A — Summary per BL
    sec(ws, row, 10, "A. SUMMARY PER BL  |  # SKU per Coverage Bucket (P2)")
    row += 1
    for i,h_ in enumerate(["pricing_bl","# SKU","(A) < 70","(B) 70-85","(C) 85-95",
                             "(D) 95-105","(E) > 105","Avg Cov P1","Avg Cov P2","Diff (pts)"], 1):
        hdr(ws, row, i, h_)
    ws.row_dimensions[row].height = 24
    row += 1

    r_secA_start = row
    for bi, bl in enumerate(BLS_ALL):
        bg_ = C_TOTAL if bl=='TOTAL' else (C_GREY if bi%2==0 else C_WHITE)
        bld = bl=='TOTAL'
        sub = ex if bl=='TOTAL' else ex[ex['pricing_bl']==bl]
        n = len(sub)
        counts = [(sub['coverage_status_p2']==b).sum() for b in cov_buckets]
        avg1 = sub['cogs_vs_comp_p1'].mean() if n>0 else 0
        avg2 = sub['cogs_vs_comp_p2'].mean() if n>0 else 0
        diff = avg2 - avg1
        st(ws, row, 1, bl, bold=bld, bg=bg_)
        st(ws, row, 2, n,  fmt=F_NUM, align="right", bold=bld, bg=bg_)
        for ci, c in enumerate(counts, 3):
            bg_c = C_RED if ci in [6,7] and c>0 else bg_
            st(ws, row, ci, c, fmt=F_NUM, align="right", bg=bg_c,
               fc="CC0000" if ci in [6,7] and c>0 else "000000")
        st(ws, row, 8,  avg1, fmt='0.00', align="right", bg=bg_)
        st(ws, row, 9,  avg2, fmt='0.00', align="right", bg=bg_)
        fc_ = ("CC0000" if diff>0 else "007030") if not bld else "000000"
        st(ws, row, 10, diff, fmt='0.00', align="right", bg=bg_, fc=fc_, bold=bld)
        row += 1
    row += 1

    # Section B — Heatmap count + % (merged), with totals, per BL
    for bi_bl, bl in enumerate(BLS_ALL):
        sub_bl = ex if bl=='TOTAL' else ex[ex['pricing_bl']==bl]
        sub_bl = sub_bl.dropna(subset=['coverage_status_p2','gp_group_p2'])

        ncols_h = len(gp_buckets) + 2  # label + 8 gp groups + total
        bg_t = C_DARK if bl=='TOTAL' else C_MID

        sec(ws, row, ncols_h,
            f"B. HEATMAP — {bl}  |  Coverage Bucket vs GP% Group (P2 current)")
        row += 1

        # Count table
        ws.merge_cells(f'A{row}:{cl(ncols_h)}{row}')
        st(ws, row, 1, "Count (# SKU)", bold=True, bg=C_LIGHT, fc=C_DARK)
        row += 1

        for i, h_ in enumerate(["Coverage P2"] + gp_buckets + ["TOTAL"], 1):
            hdr(ws, row, i, h_)
        ws.row_dimensions[row].height = 28
        row += 1

        r_heat_start = row
        for ti, bkt in enumerate(cov_buckets):
            bg_r = C_GREY if ti%2==0 else C_WHITE
            st(ws, row, 1, bkt, bg=bg_r,
               fc="CC0000" if bkt in ["(D) 95-105","(E) > 105"] else "000000",
               bold=bkt in ["(D) 95-105","(E) > 105"])
            sub_b = sub_bl[sub_bl['coverage_status_p2']==bkt]
            for ci, gpb in enumerate(gp_buckets, 2):
                cnt = (sub_b['gp_group_p2']==gpb).sum()
                st(ws, row, ci, cnt, fmt=F_NUM, align="right", bg=bg_r,
                   fc="CC0000" if cnt>0 and bkt in ["(D) 95-105","(E) > 105"] else "000000")
            st(ws, row, 10, f"=SUM(B{row}:I{row})", fmt=F_NUM, align="right",
               bold=True, bg=bg_r)
            row += 1

        # Total row count
        st(ws, row, 1, "TOTAL", bold=True, bg=C_TOTAL, bt=True)
        for ci in range(2, 11):
            st(ws, row, ci, f"=SUM({cl(ci)}{r_heat_start}:{cl(ci)}{row-1})",
               fmt=F_NUM, align="right", bold=True, bg=C_TOTAL, bt=True)
        tot_row_count = row; row += 2

        # % table
        ws.merge_cells(f'A{row}:{cl(ncols_h)}{row}')
        st(ws, row, 1, "% of Total SKU in BL  |  Row + Col totals sum to 100%", bold=True, bg=C_LIGHT, fc=C_DARK)
        row += 1

        for i, h_ in enumerate(["Coverage P2"] + gp_buckets + ["Row Total"], 1):
            hdr(ws, row, i, h_)
        ws.row_dimensions[row].height = 28
        row += 1

        r_pct_start = row
        total_n = len(sub_bl)
        for ti, bkt in enumerate(cov_buckets):
            bg_r = C_GREY if ti%2==0 else C_WHITE
            cnt_row = r_heat_start + ti
            st(ws, row, 1, bkt, bg=bg_r)
            for ci in range(2, 10):
                st(ws, row, ci,
                   f"=IF({cl(ci)}{tot_row_count}>0,{cl(ci)}{cnt_row}/{cl(ci)}{tot_row_count},0)",
                   fmt=F_PCT, align="right", bg=bg_r)
            st(ws, row, 10, f"=SUM(B{row}:I{row})", fmt=F_PCT, align="right", bold=True, bg=bg_r)
            row += 1

        # Col total % row
        st(ws, row, 1, "Col Total", bold=True, bg=C_TOTAL, bt=True)
        for ci in range(2, 11):
            st(ws, row, ci, f"=SUM({cl(ci)}{r_pct_start}:{cl(ci)}{row-1})",
               fmt=F_PCT, align="right", bold=True, bg=C_TOTAL, bt=True)
        row += 2

    # ── Sheet 6b: SKU bermasalah (coverage D or E) ────────────────────────────
    ws2 = wb.create_sheet("6b. SKU Bermasalah")
    ws2.column_dimensions['A'].width = 12
    ws2.column_dimensions['B'].width = 38
    for c in ['C','D','E','F','G','H','I','J','K','L','M','N','O','P']:
        ws2.column_dimensions[c].width = 14

    row2 = 1
    bermasalah = ex[ex['coverage_status_p2'].isin(["(D) 95-105","(E) > 105"])].copy()
    bermasalah = bermasalah.sort_values('cogs_vs_comp_p2', ascending=False)

    title_row(ws2, row2, 16,
        f"SKU BERMASALAH — {p1} vs {p2}  |  Coverage Status P2 = (D) 95-105 atau (E) > 105")
    row2 += 1
    ws2.merge_cells(f'A{row2}:P{row2}')
    st(ws2, row2, 1,
       f"Total {len(bermasalah):,} SKU bermasalah  |  "
       f"(D) 95-105 = COGS hampir sama dengan comp price  |  "
       f"(E) >105 = COGS melebihi comp price  |  Sorted by coverage P2 descending",
       italic=True, fc="595959", bg=C_LIGHT)
    row2 += 2

    # Summary per BL
    sec(ws2, row2, 6, "RINGKASAN PER BL")
    row2 += 1
    for i, h_ in enumerate(["pricing_bl","# SKU (D)","# SKU (E)","Total Bermasalah",
                              "% of BL SKU","Avg Coverage P2"], 1):
        hdr(ws2, row2, i, h_)
    ws2.row_dimensions[row2].height = 24
    row2 += 1

    for bi, bl in enumerate(BLS_ALL):
        sub = ex if bl=='TOTAL' else ex[ex['pricing_bl']==bl]
        n_d = (sub['coverage_status_p2']=="(D) 95-105").sum()
        n_e = (sub['coverage_status_p2']=="(E) > 105").sum()
        tot = n_d + n_e
        pct = tot/len(sub) if len(sub)>0 else 0
        avg = sub[sub['coverage_status_p2'].isin(["(D) 95-105","(E) > 105"])]['cogs_vs_comp_p2'].mean()
        bg_ = C_TOTAL if bl=='TOTAL' else (C_GREY if bi%2==0 else C_WHITE)
        bld = bl=='TOTAL'
        st(ws2, row2, 1, bl, bold=bld, bg=bg_)
        st(ws2, row2, 2, n_d, fmt=F_NUM, align="right", bg=bg_, fc="C55A11" if n_d>0 else "000000")
        st(ws2, row2, 3, n_e, fmt=F_NUM, align="right", bg=bg_, fc="CC0000" if n_e>0 else "000000")
        st(ws2, row2, 4, tot, fmt=F_NUM, align="right", bold=bld, bg=bg_)
        st(ws2, row2, 5, pct, fmt=F_PCT, align="right", bg=bg_, fc="CC0000" if pct>0.1 else "000000")
        st(ws2, row2, 6, avg if not np.isnan(avg) else 0, fmt='0.00', align="right", bg=bg_)
        row2 += 1
    row2 += 1

    # Full SKU list
    sec(ws2, row2, 16, f"DAFTAR SKU BERMASALAH ({len(bermasalah):,} SKU)  |  Sorted by Coverage P2 descending")
    row2 += 1

    c_cols = [c for c in [
        'product_id','product_name','pricing_bl','l1_category',
        'price_p1','price_p2','price_diff_pct',
        'cogs_p1','cogs_p2','cogs_diff_pct',
        'comp_price_p1','comp_price_p2',
        'cogs_vs_comp_p1','cogs_vs_comp_p2',
        'coverage_status_p1','coverage_status_p2',
        'margin_p1','margin_p2','margin_diff',
        'gp_p1','gp_p2','gp_diff',
        'flag_price',
    ] if c in bermasalah.columns]

    for i, h_ in enumerate(c_cols, 1):
        hdr(ws2, row2, i, h_)
        if i <= len(c_cols):
            ws2.column_dimensions[cl(i)].width = 14
    ws2.column_dimensions['A'].width = 12
    ws2.column_dimensions['B'].width = 38
    ws2.row_dimensions[row2].height = 24
    row2 += 1

    pct_c = {'price_diff_pct','cogs_diff_pct','margin_p1','margin_p2','margin_diff'}
    pt_c  = {'cogs_vs_comp_p1','cogs_vs_comp_p2'}
    for idx, (_, rd) in enumerate(bermasalah.iterrows()):
        r = row2 + idx
        cov2 = rd.get('coverage_status_p2','')
        bg_ = C_RED if cov2=="(E) > 105" else C_OR
        for i, col in enumerate(c_cols, 1):
            v = rd.get(col, None)
            if isinstance(v, float) and np.isnan(v): v = None
            fmt_ = F_PCT if col in pct_c else ('0.00' if col in pt_c else (F_NUM if i>4 else None))
            st(ws2, r, i, v, bg=bg_, fmt=fmt_, align="right" if i>4 else "left")

    ws2.freeze_panes = "C4"


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 7 — SKU Watch List (combined: Flag Price + Framework Check)
# ══════════════════════════════════════════════════════════════════════════════
def write_s7(wb, df, p1, p2):
    progress("Sheet 7: SKU Watch List...", 8, 8)
    ws = wb.create_sheet("7. SKU Watch List")

    ex = df[df['sku_status']=='Existing'].copy()

    # Three tiers of concern
    priority  = ex[ex['flag_price']=='Priority'].copy()
    review    = ex[ex['flag_price'].isin(['Review','Adjust'])].copy()
    framework = ex[ex['framework_check']==True].copy()

    # Sort each by gp_diff ascending (most negative first)
    for d_ in [priority, review, framework]:
        if 'gp_diff' not in d_.columns:
            d_['gp_diff'] = d_['gp_p2'] - d_['gp_p1'] if 'gp_p1' in d_.columns else 0

    priority  = priority.sort_values('gp_diff',  ascending=True)
    review    = review.sort_values('gp_diff',    ascending=True)
    framework = framework.sort_values('gp_diff', ascending=True)

    COL_ORDER = [c for c in [
        'product_id','product_name','pricing_bl','l1_category',
        'pareto_class','sku_status',
        'price_p1','price_p2','price_diff_pct',
        'cogs_p1','cogs_p2','cogs_diff_pct',
        'comp_price_p1','comp_price_p2','comp_price_diff_pct',
        'pi_p1','pi_p2',
        'margin_p1','margin_p2','margin_diff',
        'gp_p1','gp_p2','gp_diff',
        'cogs_effect_rp','price_effect_rp','vol_mix_effect_rp',
        'cogs_status','price_status','comp_status',
        'coverage_status_p2','cogs_vs_comp_p2',
        'framework_check','flag_price',
    ] if c in ex.columns]

    pct_cols = {c for c in COL_ORDER if 'pct' in c or c in ['margin_p1','margin_p2','margin_diff']}
    pt_cols  = {'cogs_vs_comp_p2'}
    ncols = len(COL_ORDER)

    row = 1
    title_row(ws, row, ncols,
        f"SKU WATCH LIST — {p1} vs {p2}  |  Existing SKU Only  |  Sorted by GP Diff ascending")
    row += 1
    ws.merge_cells(f'A{row}:{cl(ncols)}{row}')
    st(ws, row, 1,
       f"Priority: {len(priority):,} SKU  |  Review/Adjust: {len(review):,} SKU  |  "
       f"Framework Check: {len(framework):,} SKU  |  "
       f"Red=Priority, Amber=Review/Adjust, Purple=Framework Only",
       italic=True, fc="595959", bg=C_LIGHT)
    row += 2

    def write_section(title, data, bg_row, bg_sec):
        nonlocal row
        if len(data) == 0:
            return
        sec(ws, row, ncols, f"{title}  ({len(data):,} SKU)", bg=bg_sec)
        row += 1
        for i, col in enumerate(COL_ORDER, 1):
            hdr(ws, row, i, col)
        ws.row_dimensions[row].height = 24
        row += 1
        for idx, (_, rd) in enumerate(data.iterrows()):
            bg_ = bg_row if idx%2==0 else C_WHITE
            for i, col in enumerate(COL_ORDER, 1):
                v = rd.get(col, None)
                if isinstance(v, float) and np.isnan(v): v = None
                if isinstance(v, bool): v = True if v else None
                fmt_ = F_PCT if col in pct_cols else \
                       '0.00' if col in pt_cols else \
                       (F_NUM if i > 6 else None)
                st(ws, row, i, v, bg=bg_, fmt=fmt_,
                   align="right" if i > 6 else "left")
            row += 1
        row += 1

    write_section("PRIORITY — COGS naik + Comp naik + Price flat  |  ATAU  COGS turun + Comp turun + Price flat",
                  priority, C_RED, "8B0000")

    write_section("REVIEW / ADJUST — Perlu perhatian pricing",
                  review, C_AMBER, "7F4000")

    # Framework only = framework=True but not in Priority/Review/Adjust
    fw_only = framework[~framework.index.isin(priority.index) &
                        ~framework.index.isin(review.index)].copy()
    write_section("FRAMEWORK CHECK — PI atau Margin di luar threshold",
                  fw_only, C_PURPLE, "4B0082")

    # Column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 38
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 22
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 12
    for ci in range(7, ncols+1):
        ws.column_dimensions[cl(ci)].width = 14
    ws.freeze_panes = "E4"


# ══════════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════════
def main():
    print("\n" + "="*60)
    print("  PVM ANALYZER v3.0")
    print("="*60)

    fp=select_file()
    print(f"\n  Loading: {os.path.basename(fp)}")
    df=load_data(fp)
    print(f"  Rows: {len(df):,}  |  Cols: {len(df.columns)}")

    p1c,p2c=detect_period(df)
    if p1c is None:
        print("  Period columns not found.")
        p1=input("  Period 1 date (YYYY-MM-DD): ").strip() or "P1"
        p2=input("  Period 2 date (YYYY-MM-DD): ").strip() or "P2"
        df['week_key']=p1; df['next_key']=p2
        p1c,p2c='week_key','next_key'
    else:
        p1,p2=get_dates(df,p1c,p2c)
    print(f"  Period: {p1} vs {p2}\n")

    df=ensure_cols(df)

    print("  [1/3] ENRICHING RAW DATA")
    df=enrich(df,p1c,p2c)

    print("\n  [2/3] COMPUTING PVM BRIDGE")
    pvm=compute_pvm(df)
    progress("PVM computed.", 10, 10)

    print("\n  [3/3] WRITING EXCEL OUTPUT")
    wb=Workbook()
    write_s1(wb,df,p1,p2)
    write_exec(wb,df,pvm,p1,p2)
    AGG=write_s1b(wb,pvm,p1,p2)
    write_s2(wb,pvm,AGG,p1,p2)
    write_s2b(wb,df,pvm,AGG,p1,p2)
    write_s3(wb,df,p1,p2)
    write_s4(wb,df,p1,p2)
    write_s5(wb,df,p1,p2)
    write_s6(wb,df,p1,p2)
    write_s7(wb,df,p1,p2)

    out=f"{p1}_vs_{p2}_enriched.xlsx"
    out_path=os.path.join(os.path.dirname(os.path.abspath(fp)),out)
    write_s0(wb,p1,p2)
    wb.save(out_path)
    print(f"\n{'='*60}")
    print(f"  OUTPUT: {out}")
    print(f"  Path:   {out_path}")
    print(f"{'='*60}\n")


# ══════════════════════════════════════════════════════════════════════════════
# MODULE API — for Streamlit / other callers
# ══════════════════════════════════════════════════════════════════════════════
def compute(df_input, p1_label=None, p2_label=None):
    """
    Fast analytical computation only — NO Excel workbook generation.
    Returns dict for Streamlit display (fast).

    Args:
        df_input: raw input DataFrame
        p1_label, p2_label: optional period label overrides

    Returns:
        dict with keys: df, pvm, p1, p2, meta, p1c, p2c
    """
    df = df_input.copy()

    # Period detection
    p1c, p2c = detect_period(df)
    if p1c is None:
        p1 = p1_label or "P1"
        p2 = p2_label or "P2"
        df['week_key'] = p1
        df['next_key'] = p2
        p1c, p2c = 'week_key', 'next_key'
    else:
        p1, p2 = get_dates(df, p1c, p2c)

    # Detect missing optional columns
    optional_cols = ['comp_price', 'comp_price1', 'pi', 'pi1',
                     'avg_stock', 'avg_stock1', 'pareto_classification',
                     'margin_pct', 'margin1_pct']
    missing_optional = [c for c in optional_cols if c not in df.columns]

    df = ensure_cols(df)
    df = enrich(df, p1c, p2c)
    pvm = compute_pvm(df)

    meta = {
        'n_rows': len(df),
        'n_cols': len(df.columns),
        'missing_optional': missing_optional,
        'period_detected': p1c is not None and p1c != 'week_key',
        'p1_col': p1c,
        'p2_col': p2c,
    }

    return {
        'df': df,
        'pvm': pvm,
        'p1': p1,
        'p2': p2,
        'p1c': p1c,
        'p2c': p2c,
        'meta': meta,
    }


def generate_excel(result, progress_callback=None):
    """
    Generate Excel workbook from compute() result.
    SLOW (~20s for 12K rows) — only call when user requests download.

    Returns:
        bytes of Excel workbook
    """
    import io
    df = result['df']
    pvm = result['pvm']
    p1 = result['p1']
    p2 = result['p2']

    # Optional progress callback
    global progress
    original_progress = progress
    if progress_callback is not None:
        def _pcb(msg, cur, tot):
            try: progress_callback(msg, cur, tot)
            except Exception: pass
        progress = _pcb

    try:
        wb = Workbook()
        write_s1(wb, df, p1, p2)
        write_exec(wb, df, pvm, p1, p2)
        AGG = write_s1b(wb, pvm, p1, p2)
        write_s2(wb, pvm, AGG, p1, p2)
        write_s2b(wb, df, pvm, AGG, p1, p2)
        write_s3(wb, df, p1, p2)
        write_s4(wb, df, p1, p2)
        write_s5(wb, df, p1, p2)
        write_s6(wb, df, p1, p2)
        write_s7(wb, df, p1, p2)
        write_s0(wb, p1, p2)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.getvalue()
    finally:
        progress = original_progress


def analyze(df_input, p1_label=None, p2_label=None, progress_callback=None):
    """
    Legacy wrapper for backwards compatibility — runs compute() + generate_excel().
    For Streamlit, prefer using compute() + generate_excel() separately.
    """
    result = compute(df_input, p1_label, p2_label)
    result['excel_bytes'] = generate_excel(result, progress_callback)
    return result


if __name__=="__main__":
    main()
