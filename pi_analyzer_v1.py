import os, sys, glob, time
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────
# COLORS
# ─────────────────────────────────────────────
NAVY       = "1C4587"
NAVY_LT    = "2A5BA8"
WHITE      = "FFFFFF"
LT_BLUE    = "EEF2FB"
RED_TXT    = "C0392B"
GREEN_TXT  = "1A7A4A"
MUTED      = "A0A8C0"
SUB_BG     = "F7F8FC"

SEG_COLORS = {
    "Overall": ("1C4587", "EEF2FB"),
    "Dry":     ("145A32", "E8F5EE"),
    "Fresh":   ("1A5276", "EBF5FB"),
    "Frozen":  ("6C3483", "F5EEF8"),
}

def _s(style="thin", color="CCCCCC"): return Side(style=style, color=color)
def bb(color="CCCCCC"):               return Border(bottom=_s(color=color))
def thick_l(bottom=True):
    return Border(left=_s("medium","1C4587"), bottom=_s() if bottom else None)

def sc(ws, r, c, v=None, bold=False, italic=False, bg=None, fc="000000",
       align="left", fmt=None, fs=10, wrap=False, bdr=None, ind=0):
    cell = ws.cell(row=r, column=c)
    if v is not None: cell.value = v
    cell.font = Font(name="Calibri", bold=bold, italic=italic, size=fs, color=fc)
    if bg: cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal=align, vertical="center",
                                wrap_text=wrap, indent=ind)
    if fmt: cell.number_format = fmt
    if bdr: cell.border = bdr
    return cell

def hdr(ws, r, labels, bg=NAVY, fc=WHITE, h=22, sc0=1):
    for i, l in enumerate(labels):
        sc(ws, r, sc0+i, l, bold=True, bg=bg, fc=fc,
           align="center", fs=9, wrap=True, bdr=bb("2A5BA8"))
    ws.row_dimensions[r].height = h

def sec(ws, r, text, ncols=10, seg=None):
    bg, _ = SEG_COLORS.get(seg or "Overall", SEG_COLORS["Overall"])
    label  = f"  {seg.upper()}  —  {text}" if seg else f"  {text}"
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
    sc(ws, r, 1, label, bold=True, bg=bg, fc=WHITE, fs=12)
    ws.row_dimensions[r].height = 26

def cw(ws, widths):
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col) if isinstance(col,int) else col].width = w

# ─────────────────────────────────────────────
# PROGRESS
# ─────────────────────────────────────────────
def progress(msg, current=None, total=None):
    if current is not None and total is not None:
        pct  = current / total
        done = int(pct * 20)
        bar  = "█" * done + "░" * (20 - done)
        print(f"\r  [{bar}] {pct*100:5.1f}%  {msg}", end="", flush=True)
    else:
        print(f"  {msg}", flush=True)

# ─────────────────────────────────────────────
# FILE SELECTOR
# ─────────────────────────────────────────────
def select_file():
    sd = os.path.dirname(os.path.abspath(__file__))
    found = sorted(set(
        f for pat in ["*.csv","*.xlsx","*.xls"]
        for f in glob.glob(os.path.join(sd, pat))
    ))
    if not found:
        print("Tidak ada file CSV/Excel di folder ini."); sys.exit(1)
    print("\n📂  File yang ditemukan:")
    for i, f in enumerate(found):
        print(f"  [{i+1}] {os.path.basename(f)}  ({os.path.getsize(f)/1024:.1f} KB)")
    while True:
        try:
            c = int(input("\nPilih nomor file: "))
            if 1 <= c <= len(found): return found[c-1]
        except ValueError: pass
        print("  Input tidak valid.")

# ─────────────────────────────────────────────
# LOAD
# ─────────────────────────────────────────────
def load_data(fp):
    ext = os.path.splitext(fp)[1].lower()
    print(f"\n⏳  Membaca {os.path.basename(fp)} ...")
    t0 = time.time()
    df = pd.read_csv(fp) if ext == ".csv" else pd.read_excel(fp, engine="openpyxl")
    print(f"✅  {len(df):,} baris dimuat ({time.time()-t0:.1f}s)")
    return df

def detect_period(df):
    if "week_key"  in df.columns: return "week",  "week_key",  "next_week"
    if "month_key" in df.columns: return "month", "month_key", "next_month"
    raise ValueError("Kolom period tidak ditemukan.")

# ─────────────────────────────────────────────
# BINS & LABELS
# ─────────────────────────────────────────────
PI_BINS = [0,95,100,105,110,120,9999]
PI_LBL  = ["A.<95","B.95-<100","C.100-105","D.105-110","E.110-120","F.>120"]
CI_BINS = [0,70,85,95,105,9999]
CI_LBL  = ["A.<70","B.70-85","C.85-95","D.95-105","E.>105"]
MG_BINS = [-9999,-0.20,-0.10,0,0.10,0.20,0.30,0.50,9999]
MG_LBL  = ["A.<-20%","B.-20to-10%","C.-10to0%","D.0to10%",
           "E.10to20%","F.20to30%","G.30to50%","H.>50%"]

# ─────────────────────────────────────────────
# ENRICH  (vectorised)
# ─────────────────────────────────────────────
def enrich(df):
    print("⏳  Enriching data ...")
    t0 = time.time()
    d = df.copy()
    print(repr(d.columns.tolist()))
    print(d[["pi","next_pi"]].dtypes)

    # SKU type
    d["sku_type"] = "Existing"
    d.loc[d["pi"].isna()  & d["next_pi"].notna(), "sku_type"] = "New SKU"
    d.loc[d["pi"].notna() & d["next_pi"].isna(),  "sku_type"] = "Departing SKU"

    # rename: treat "prev" = original columns, "current" = next_* columns
    # We keep original col names internally but rename for display in Sheet 1

    # Margins
    d["margin_prev"]    = d["price"]      - d["cogs"]
    d["margin_pct_prev"]= d["margin_prev"]/ d["price"]
    d["margin_cur"]     = d["next_price"] - d["next_cogs"]
    d["margin_pct_cur"] = d["margin_cur"] / d["next_price"]

    # Diffs (prev → current)
    for prev, cur, pfx in [
        ("price",      "next_price",      "price"),
        ("cogs",       "next_cogs",       "cogs"),
        ("comp_price", "next_comp_price", "comp"),
        ("margin_prev","margin_cur",      "margin"),
    ]:
        d[f"diff_{pfx}"]     = d[cur] - d[prev]
        d[f"diff_{pfx}_pct"] = d[f"diff_{pfx}"] / d[prev].replace(0, np.nan)

    d["diff_pi"]     = d["next_pi"] - d["pi"]
    d["diff_pi_pct"] = d["diff_pi"] / d["pi"].replace(0, np.nan)

    # COGS Index
    d["ci_prev"] = d["cogs"]      * 100 / d["comp_price"].replace(0, np.nan)
    d["ci_cur"]  = d["next_cogs"] * 100 / d["next_comp_price"].replace(0, np.nan)

    # Groups — prev and current
    for col, bins, lbls, out in [
        ("pi",          PI_BINS, PI_LBL, "pi_group_prev"),
        ("next_pi",     PI_BINS, PI_LBL, "pi_group_cur"),
        ("ci_prev",     CI_BINS, CI_LBL, "ci_group_prev"),
        ("ci_cur",      CI_BINS, CI_LBL, "ci_group_cur"),
        ("margin_pct_prev", MG_BINS, MG_LBL, "margin_group_prev"),
        ("margin_pct_cur",  MG_BINS, MG_LBL, "margin_group_cur"),
    ]:
        d[out] = pd.cut(d[col], bins=bins, labels=lbls, right=False).astype(str)
        d.loc[d[col].isna(), out] = None

    # Tags (vectorised)
    for prev, cur, tag in [
        ("price",      "next_price",      "price_tag"),
        ("cogs",       "next_cogs",       "cogs_tag"),
        ("comp_price", "next_comp_price", "comp_tag"),
    ]:
        da = d[cur] - d[prev]
        dp = da / d[prev].replace(0, np.nan)
        d[tag] = "Stay"
        d.loc[(da >= 5000) | (dp >= 0.05),  tag] = "Up"
        d.loc[(da <= -5000) | (dp <= -0.05), tag] = "Down"

    # Status labels
    d["price_status"] = d["price_tag"].map(
        {"Up":"Price Increase","Stay":"Price Stable","Down":"Price Reduction"})
    d["cogs_status"]  = d["cogs_tag"].map(
        {"Up":"Cost Pressure","Stay":"Cost Stable","Down":"Cost Improvement"})
    d["comp_status"]  = d["comp_tag"].map(
        {"Up":"Competitor Retreat","Stay":"Market Stable","Down":"Competitor Aggressive"})

    # Framework check (uses current = next_pi, next_margin_pct)
    pi_c  = d["next_pi"]
    mg_c  = d["margin_pct_cur"]
    bl    = d["pricing_bl_25"]
    cond1 = (bl=="Fresh")  & (pi_c > 110) & (mg_c <= 0.15)
    cond2 = (bl=="Frozen") & (pi_c > 100) & (mg_c <= 0.15)
    cond3 = (bl=="Fresh")  & (pi_c > 120) & (mg_c >= 0.70)
    cond4 = (bl=="Dry")    & (pi_c < 105) & (mg_c <= 0.00)
    cond5 = (bl=="Dry")    & (pi_c > 120) & (mg_c > 0.40)
    d["framework_check"] = None
    d.loc[cond1|cond2|cond3|cond4|cond5, "framework_check"] = "TRUE"

    # Normal comp price columns (input dari raw data)
    # discount_comp = normal_comp - comp_price  (potongan diskon kompetitor)
    d["discount_comp_price"]      = d["normal_comp_price"]      - d["comp_price"]
    d["next_discount_comp_price"] = d["next_normal_comp_price"] - d["next_comp_price"]
    d["diff_normal_comp"]         = d["next_normal_comp_price"] - d["normal_comp_price"]
    d["diff_discount_comp"]       = d["next_discount_comp_price"] - d["discount_comp_price"]

    # Effect columns — existing SKU only (Shapley Value decomposition)
    # Shapley = rata-rata dua urutan midpoint (B: price dulu, D: comp dulu)
    # Zero residual by construction — tidak ada interaction term
    ex_mask = d["pi"].notna() & d["next_pi"].notna()
    comp0   = d["comp_price"].replace(0, np.nan)
    comp1   = d["next_comp_price"].replace(0, np.nan)
    price0  = d["price"].replace(0, np.nan)

    # Metode B: PI_mid = price_cur / comp_prev * 100
    pi_mid_B = d["next_price"] / comp0 * 100
    ep_B = pi_mid_B - d["pi"]
    ec_B = d["next_pi"] - pi_mid_B

    # Metode D: PI_mid = price_prev / comp_cur * 100
    pi_mid_D = d["price"] / comp1 * 100
    ec_D = pi_mid_D - d["pi"]
    ep_D = d["next_pi"] - pi_mid_D

    # Shapley = avg(B, D)
    d["eff_price"] = np.where(ex_mask, (ep_B + ep_D) / 2, np.nan)
    d["eff_comp"]  = np.where(ex_mask, (ec_B + ec_D) / 2, np.nan)

    # Proportional split: eff_comp → eff_normal_comp + eff_discount_comp
    # eff_normal_comp   = eff_comp × (Δnormal_comp / Δcomp)
    # eff_discount_comp = eff_comp × (Δdiscount_comp / Δcomp)
    # Jika Δcomp = 0 maka kedua efek = 0 (comp net tidak berubah)
    delta_comp         = d["next_comp_price"] - d["comp_price"]
    delta_normal_comp  = d["next_normal_comp_price"] - d["normal_comp_price"]
    delta_discount_comp= d["next_discount_comp_price"] - d["discount_comp_price"]
    ratio_normal  = np.where(delta_comp != 0, delta_normal_comp   / delta_comp, 0.0)
    ratio_discount= np.where(delta_comp != 0, -delta_discount_comp / delta_comp, 0.0)
    d["eff_normal_comp"]  = np.where(ex_mask, d["eff_comp"] * ratio_normal,  np.nan)
    d["eff_discount_comp"]= np.where(ex_mask, d["eff_comp"] * ratio_discount, np.nan)

    print(f"✅  Data enriched ({time.time()-t0:.1f}s)")
    return d

# ─────────────────────────────────────────────
# DECOMPOSITION  (returns rich dict)
# ─────────────────────────────────────────────
def decompose(d):
    dep = d[d["pi"].notna() & d["next_pi"].isna()]
    new = d[d["pi"].isna()  & d["next_pi"].notna()]
    ex  = d[d["pi"].notna() & d["next_pi"].notna()]

    n_ex  = len(ex);  n_dep = len(dep);  n_new = len(new)
    n_cur = n_ex + n_dep;                n_nxt = n_ex + n_new

    A = pd.concat([ex, dep])["pi"].mean()      if n_cur > 0 else np.nan
    B = ex["pi"].mean()                         if n_ex  > 0 else np.nan
    C = ex["next_pi"].mean()                    if n_ex  > 0 else np.nan
    D = new["next_pi"].mean()                   if n_new > 0 else np.nan
    E = (n_ex*C + n_new*D)/(n_nxt)              if n_new > 0 and not np.isnan(D) else C

    eff_dep = (B - A) if not (np.isnan(A) or np.isnan(B)) else 0.0

    s = ex.copy()
    comp0  = s["comp_price"].replace(0, np.nan)
    comp1  = s["next_comp_price"].replace(0, np.nan)

    # Metode B
    pi_mid_B = s["next_price"] / comp0 * 100
    ep_B = pi_mid_B - s["pi"]
    ec_B = s["next_pi"] - pi_mid_B

    # Metode D
    pi_mid_D = s["price"] / comp1 * 100
    ec_D = pi_mid_D - s["pi"]
    ep_D = s["next_pi"] - pi_mid_D

    # Shapley
    s["ep"] = (ep_B + ep_D) / 2
    s["ec"] = (ec_B + ec_D) / 2

    ep = s["ep"].mean() if n_ex > 0 else 0.0
    ec = s["ec"].mean() if n_ex > 0 else 0.0
    en = (E - C)        if not (np.isnan(E) or np.isnan(C)) else 0.0

    # Proportional split of eff_comp → eff_normal_comp + eff_discount_comp
    delta_comp          = s["next_comp_price"] - s["comp_price"]
    delta_normal_comp   = s["next_normal_comp_price"] - s["normal_comp_price"]
    delta_discount_comp = s["next_discount_comp_price"] - s["discount_comp_price"]
    ratio_n = np.where(delta_comp != 0, delta_normal_comp    / delta_comp, 0.0)
    ratio_d = np.where(delta_comp != 0, -delta_discount_comp / delta_comp, 0.0)
    s["enc"] = s["ec"] * ratio_n   # eff_normal_comp per SKU
    s["edc"] = s["ec"] * ratio_d   # eff_discount_comp per SKU
    enc = s["enc"].mean() if n_ex > 0 else 0.0
    edc = s["edc"].mean() if n_ex > 0 else 0.0

    # sums needed for exact contribution calc
    sum_pi_ex  = ex["pi"].sum()           if n_ex  > 0 else 0.0
    sum_pi_cur = pd.concat([ex,dep])["pi"].sum() if n_cur > 0 else 0.0
    sum_npi_ex = ex["next_pi"].sum()      if n_ex  > 0 else 0.0
    sum_npi_new= new["next_pi"].sum()     if n_new > 0 else 0.0

    return {
        "A":A,"B":B,"C":C,"D":D,"E":E,
        "eff_dep":eff_dep,"eff_price":ep,"eff_comp":ec,
        "eff_normal_comp":enc,"eff_discount_comp":edc,
        "eff_new":en,
        "total":(E-A) if not (np.isnan(E) or np.isnan(A)) else 0.0,
        "n_ex":n_ex,"n_dep":n_dep,"n_new":n_new,"n_cur":n_cur,"n_next":n_nxt,
        "sum_pi_ex":sum_pi_ex,"sum_pi_cur":sum_pi_cur,
        "sum_npi_ex":sum_npi_ex,"sum_npi_new":sum_npi_new,
        "sum_ep":s["ep"].sum() if n_ex>0 else 0.0,
        "sum_ec":s["ec"].sum() if n_ex>0 else 0.0,
        "sum_enc":s["enc"].sum() if n_ex>0 else 0.0,
        "sum_edc":s["edc"].sum() if n_ex>0 else 0.0,
    }

# ─────────────────────────────────────────────
# EXACT CONTRIBUTION  (fixed formula)
# ─────────────────────────────────────────────
def exact_contributions(seg_res, overall):
    """
    For price/comp/int: contrib_seg = sum_effect_seg / n_ex_total
    For departing:      contrib_seg = sum_pi_ex_seg/n_ex_total - sum_pi_cur_seg/n_cur_total
    For new:            contrib_seg = (sum_npi_ex_seg + sum_npi_new_seg)/n_next_total
                                      - sum_npi_ex_seg/n_ex_total
    All three sum exactly to overall effect.
    """
    segs   = ["Dry","Fresh","Frozen"]
    n_ex   = overall["n_ex"]
    n_cur  = overall["n_cur"]
    n_nxt  = overall["n_next"]

    contribs = {}
    for seg in segs:
        r = seg_res[seg]
        # price / comp
        cp  = r["sum_ep"]  / n_ex  if n_ex > 0 else 0
        cc  = r["sum_ec"]  / n_ex  if n_ex > 0 else 0
        cnc = r["sum_enc"] / n_ex  if n_ex > 0 else 0
        cdc = r["sum_edc"] / n_ex  if n_ex > 0 else 0

        # departing  (exact)
        cd = (r["sum_pi_ex"]  / n_ex  if n_ex  > 0 else 0) \
           - (r["sum_pi_cur"] / n_cur if n_cur > 0 else 0)

        # new SKU  (exact)
        cn = ((r["sum_npi_ex"] + r["sum_npi_new"]) / n_nxt if n_nxt > 0 else 0) \
           - (r["sum_npi_ex"] / n_ex if n_ex > 0 else 0)

        contribs[seg] = {
            "eff_dep":            cd,
            "eff_price":          cp,
            "eff_comp":           cc,
            "eff_normal_comp":    cnc,
            "eff_discount_comp":  cdc,
            "eff_new":            cn,
        }

    # Verify sums
    for eff in ["eff_dep","eff_price","eff_comp","eff_normal_comp","eff_discount_comp","eff_new"]:
        s = sum(contribs[seg][eff] for seg in segs)
        ov = overall[eff]
        residual = abs(s - ov)
        if residual > 0.0001:
            print(f"  ⚠️  Residual {eff}: {residual:.6f}")

    return contribs

# ─────────────────────────────────────────────
# PRE-COMPUTE ALL AGGREGATES FOR SHEET 1B
# ─────────────────────────────────────────────
def precompute(d):
    print("⏳  Pre-computing aggregates ...")
    t0     = time.time()
    segs   = ["Dry","Fresh","Frozen"]
    ov     = decompose(d)
    sr     = {s: decompose(d[d["pricing_bl_25"]==s]) for s in segs}
    contribs = exact_contributions(sr, ov)

    # weights
    for s in segs:
        r = sr[s]
        r["w_ex"]  = r["n_ex"]  / ov["n_ex"]  if ov["n_ex"]  > 0 else 0
        r["w_cur"] = r["n_cur"] / ov["n_cur"]  if ov["n_cur"] > 0 else 0
        r["w_nxt"] = r["n_ex"]  / ov["n_next"] if ov["n_next"]> 0 else 0

    print(f"✅  Aggregates ready ({time.time()-t0:.1f}s)")
    return ov, sr, contribs

# ─────────────────────────────────────────────
# SHEET 1 — RAW DATA  (batch write)
# ─────────────────────────────────────────────
def build_s1(wb, d, period_col, next_col):
    print("\n⏳  Sheet 1 — Raw Data ...")
    t0 = time.time()
    ws = wb.active
    ws.title = "1. Raw Data"

    # Column order — renamed to prev/current
    cols_order = [
        period_col, next_col,
        "product_id","product_name","l1_category_name",
        "pricing_bl_25","pareto_classification","source_status","sku_type",
        "framework_check",
        # Price
        "price","next_price","diff_price","diff_price_pct","price_status",
        # COGS
        "cogs","next_cogs","diff_cogs","diff_cogs_pct","cogs_status",
        # Margin
        "margin_prev","margin_cur","diff_margin","diff_margin_pct",
        "margin_pct_prev","margin_pct_cur",
        # Normal Comp Price
        "normal_comp_price","next_normal_comp_price","diff_normal_comp",
        # Discount Comp Price
        "discount_comp_price","next_discount_comp_price","diff_discount_comp",
        # Comp (effective / discounted)
        "comp_price","next_comp_price","diff_comp","diff_comp_pct","comp_status",
        # PI
        "pi","next_pi","diff_pi","diff_pi_pct",
        # Effects (existing SKU only)
        "eff_price","eff_comp","eff_normal_comp","eff_discount_comp",
        # COGS Index
        "ci_prev","ci_cur",
        # Groups
        "pi_group_prev","pi_group_cur",
        "ci_group_prev","ci_group_cur",
        "margin_group_prev","margin_group_cur",
        # Tags
        "price_tag","cogs_tag","comp_tag",
    ]
    cols_order = [c for c in cols_order if c in d.columns]

    # Display labels
    col_labels = {
        period_col:"Period (Prev)", next_col:"Period (Current)",
        "product_id":"Product ID","product_name":"Product Name",
        "l1_category_name":"L1 Category","pricing_bl_25":"Pricing BL 25",
        "pareto_classification":"Pareto","source_status":"Source Status",
        "sku_type":"SKU Type","framework_check":"Framework Check",
        "price":"Prev","next_price":"Current","diff_price":"Diff","diff_price_pct":"Diff %",
        "price_status":"Price Status",
        "cogs":"Prev","next_cogs":"Current","diff_cogs":"Diff","diff_cogs_pct":"Diff %",
        "cogs_status":"COGS Status",
        "margin_prev":"Prev","margin_cur":"Current",
        "diff_margin":"Diff","diff_margin_pct":"Diff %",
        "margin_pct_prev":"Margin % Prev","margin_pct_cur":"Margin % Cur",
        "normal_comp_price":"Prev","next_normal_comp_price":"Current","diff_normal_comp":"Diff",
        "discount_comp_price":"Prev","next_discount_comp_price":"Current","diff_discount_comp":"Diff",
        "comp_price":"Prev","next_comp_price":"Current",
        "diff_comp":"Diff","diff_comp_pct":"Diff %","comp_status":"Comp Status",
        "pi":"Prev","next_pi":"Current","diff_pi":"Diff","diff_pi_pct":"Diff %",
        "eff_price":"Eff Price","eff_comp":"Eff Comp (Total)",
        "eff_normal_comp":"Eff Normal Comp","eff_discount_comp":"Eff Discount Comp",
        "ci_prev":"COGS Idx Prev","ci_cur":"COGS Idx Cur",
        "pi_group_prev":"PI Grp Prev","pi_group_cur":"PI Grp Cur",
        "ci_group_prev":"CI Grp Prev","ci_group_cur":"CI Grp Cur",
        "margin_group_prev":"Margin Grp Prev","margin_group_cur":"Margin Grp Cur",
        "price_tag":"Price Tag","cogs_tag":"COGS Tag","comp_tag":"Comp Tag",
    }

    groups = {
        "Price":             ["price","next_price","diff_price","diff_price_pct","price_status"],
        "COGS":              ["cogs","next_cogs","diff_cogs","diff_cogs_pct","cogs_status"],
        "Margin":            ["margin_prev","margin_cur","diff_margin","diff_margin_pct",
                              "margin_pct_prev","margin_pct_cur"],
        "Normal Comp Price": ["normal_comp_price","next_normal_comp_price","diff_normal_comp"],
        "Discount Comp Price":["discount_comp_price","next_discount_comp_price","diff_discount_comp"],
        "Comp Price (Eff)":  ["comp_price","next_comp_price","diff_comp","diff_comp_pct","comp_status"],
        "PI":                ["pi","next_pi","diff_pi","diff_pi_pct"],
        "Effects":           ["eff_price","eff_comp","eff_normal_comp","eff_discount_comp"],
        "COGS Index":        ["ci_prev","ci_cur"],
        "Groups":            ["pi_group_prev","pi_group_cur","ci_group_prev","ci_group_cur",
                              "margin_group_prev","margin_group_cur"],
        "Tags":              ["price_tag","cogs_tag","comp_tag"],
    }

    col_idx = {c: i+1 for i,c in enumerate(cols_order)}
    data    = d[cols_order]

    # ── Row 1: group headers ──
    info_cols = [period_col,next_col,"product_id","product_name","l1_category_name",
                 "pricing_bl_25","pareto_classification","source_status","sku_type",
                 "framework_check"]
    for c in info_cols:
        if c in col_idx:
            ws.cell(row=1,column=col_idx[c]).fill = PatternFill("solid",fgColor="E8EDF8")

    for grp, gcols in groups.items():
        valid = [c for c in gcols if c in col_idx]
        if not valid: continue
        s_col = min(col_idx[c] for c in valid)
        e_col = max(col_idx[c] for c in valid)
        if s_col < e_col:
            ws.merge_cells(start_row=1,start_column=s_col,end_row=1,end_column=e_col)
        cell = ws.cell(row=1,column=s_col,value=grp)
        cell.font = Font(name="Calibri",bold=True,size=10,color=WHITE)
        cell.fill = PatternFill("solid",fgColor=NAVY)
        cell.alignment = Alignment(horizontal="center",vertical="center")

    # ── Row 2: column headers ──
    for c, idx in col_idx.items():
        cell = ws.cell(row=2,column=idx,value=col_labels.get(c,c))
        cell.font = Font(name="Calibri",bold=True,size=9,color=WHITE)
        cell.fill = PatternFill("solid",fgColor=NAVY_LT)
        cell.alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)

    ws.row_dimensions[1].height = 18
    ws.row_dimensions[2].height = 30
    ws.freeze_panes = "A3"

    pct_cols = {"diff_price_pct","diff_cogs_pct","diff_margin_pct","diff_comp_pct",
                "diff_pi_pct","margin_pct_prev","margin_pct_cur"}
    num_cols = {"price","next_price","diff_price","cogs","next_cogs","diff_cogs",
                "margin_prev","margin_cur","diff_margin",
                "normal_comp_price","next_normal_comp_price","diff_normal_comp",
                "discount_comp_price","next_discount_comp_price","diff_discount_comp",
                "comp_price","next_comp_price","diff_comp",
                "pi","next_pi","diff_pi","ci_prev","ci_cur",
                "eff_price","eff_comp","eff_normal_comp","eff_discount_comp"}

    # ── Batch write data rows ──
    total = len(data)
    for r_idx, row in enumerate(data.itertuples(index=False), 3):
        if (r_idx-3) % 2000 == 0:
            progress("Writing rows...", r_idx-3, total)
        bg = "FFFFFF" if r_idx % 2 == 1 else "F5F7FF"
        for c_name, c_idx in col_idx.items():
            val = getattr(row, c_name, None)
            if isinstance(val, float) and np.isnan(val): val = None
            if isinstance(val, str) and val == "nan":    val = None
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.font = Font(name="Calibri", size=9)
            cell.fill = PatternFill("solid", fgColor=bg)
            cell.alignment = Alignment(
                horizontal="right" if c_name in pct_cols or c_name in num_cols else "left",
                vertical="center")
            if c_name in pct_cols:    cell.number_format = "0.0%"
            elif c_name in num_cols:  cell.number_format = "#,##0.00"
            # Framework check highlight
            if c_name == "framework_check" and val == "TRUE":
                cell.fill = PatternFill("solid", fgColor="FFF3CD")
                cell.font = Font(name="Calibri", size=9, bold=True, color="856404")

    progress("Writing rows...", total, total)
    print()

    # widths
    widths = {"product_name":28,"l1_category_name":22,period_col:14,next_col:14,
              "product_id":12,"pricing_bl_25":12,"pareto_classification":14,
              "sku_type":14,"framework_check":16,"price_status":16,
              "cogs_status":16,"comp_status":20}
    for c, idx in col_idx.items():
        ws.column_dimensions[get_column_letter(idx)].width = widths.get(c, 11)

    print(f"✅  Sheet 1 selesai ({time.time()-t0:.1f}s) — {total:,} baris")

# ─────────────────────────────────────────────
# SHEET 1B — AGGREGATES (single source of truth)
# ─────────────────────────────────────────────
def build_s1b(wb, ov, sr, contribs):
    print("⏳  Sheet 1B — Aggregates ...")
    t0   = time.time()
    ws   = wb.create_sheet("1B. Aggregates")
    segs = ["Dry","Fresh","Frozen"]

    def _sec(r, text):
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        sc(ws, r, 1, text, bold=True, bg=NAVY, fc=WHITE, fs=11, ind=1)
        ws.row_dimensions[r].height = 22

    def _hdr(r, labels):
        hdr(ws, r, labels, h=20)

    def _row(r, label, vals, bold=False, bg_label="FFFFFF"):
        sc(ws, r, 1, label, bold=bold, bg=bg_label, ind=1, fs=9)
        for i, v in enumerate(vals):          # fixed: was enumerate(vals, 2)
            fmt = "#,##0" if isinstance(v, int) else \
                  "0.0%"  if isinstance(v, float) and abs(v) <= 2 and "w_" in label.lower() else \
                  "#,##0.000000"
            sc(ws, r, 2+i, v, bold=bold, align="right", fmt=fmt, fs=9)  # fixed: was 1+i
        ws.row_dimensions[r].height = 16

    r = 1
    # ── Section 1: SKU Count ──
    _sec(r, "Section 1 — SKU Count per Segment"); r += 1
    _hdr(r, ["Metric","Overall","Dry","Fresh","Frozen"]); r += 1

    addr = {}

    def _store(key, row, segs_list=["Overall","Dry","Fresh","Frozen"]):
        addr[key] = {}
        for i, s in enumerate(segs_list):
            addr[key][s] = f"'1B. Aggregates'!{get_column_letter(2+i)}{row}"

    # SKU type col in Sheet 1 = col I (col 9), BL25 col = col F (col 6)
    sku_type_col = "'1. Raw Data'!I:I"
    bl_col       = "'1. Raw Data'!F:F"

    sku_count_formulas = {
        "n_ex":  ("Existing",      "COUNTIFS"),
        "n_dep": ("Departing SKU", "COUNTIFS"),
        "n_new": ("New SKU",       "COUNTIFS"),
    }

    for lbl, (sku_type, _) in [
        ("n Existing",     ("Existing",      None)),
        ("n Departing",    ("Departing SKU", None)),
        ("n New SKU",      ("New SKU",       None)),
        ("n Current Pool", (None,            "cur")),
        ("n Next Pool",    (None,            "nxt")),
    ]:
        key = {"n Existing":"n_ex","n Departing":"n_dep","n New SKU":"n_new",
               "n Current Pool":"n_cur","n Next Pool":"n_next"}[lbl]

        row_vals = []
        for seg in ["Overall","Dry","Fresh","Frozen"]:
            bl_filter = f',{bl_col},"{seg}"' if seg != "Overall" else ""
            if sku_type:
                f = f'=COUNTIFS({sku_type_col},"{sku_type}"{bl_filter})'
            elif key == "n_cur":
                f1 = f'=COUNTIFS({sku_type_col},"Existing"{bl_filter})'
                f2 = f'COUNTIFS({sku_type_col},"Departing SKU"{bl_filter.replace("=","",1) if bl_filter else ""})'
                f  = f'={f1[1:]}+COUNTIFS({sku_type_col},"Departing SKU"{bl_filter})'
            else:  # n_next
                f  = f'=COUNTIFS({sku_type_col},"Existing"{bl_filter})+COUNTIFS({sku_type_col},"New SKU"{bl_filter})'
            row_vals.append(f)

        # Write formulas directly
        sc(ws, r, 1, lbl, ind=1, fs=9)
        for i, fval in enumerate(row_vals):
            cell = ws.cell(row=r, column=2+i, value=fval)
            cell.number_format = "#,##0"
            cell.font = Font(name="Calibri", size=9)
            cell.alignment = Alignment(horizontal="right", vertical="center")
        ws.row_dimensions[r].height = 16
        _store(key, r)
        r += 1

    r += 1
    # ── Section 2: Avg PI ──
    _sec(r, "Section 2 — Avg PI per Segment"); r += 1
    _hdr(r, ["Metric","Overall","Dry","Fresh","Frozen"]); r += 1

    pi_prev_col = "'1. Raw Data'!AF:AF"   # pi col (col 32)
    pi_cur_col  = "'1. Raw Data'!AG:AG"   # next_pi col (col 33)

    for lbl, key, pi_col, type_filter in [
        ("A — Avg PI Prev (existing+dep)", "A",   pi_prev_col, ["Existing","Departing SKU"]),
        ("B — Avg PI Prev (existing only)", "B",  pi_prev_col, ["Existing"]),
        ("C — Avg PI Current (existing only)","C",pi_cur_col,  ["Existing"]),
        ("D — Avg PI Current (new SKU only)","D", pi_cur_col,  ["New SKU"]),
        ("E — Avg PI Current (existing+new)","E", None,        None),
    ]:
        sc(ws, r, 1, lbl, ind=1, fs=9)

        for i, seg in enumerate(["Overall","Dry","Fresh","Frozen"]):
            bl_filter = f',{bl_col},"{seg}"' if seg != "Overall" else ""
            if key == "E":
                # E = (n_ex * C + n_new * D) / n_next
                c_ref = addr["C"][seg]
                d_ref = addr["D"][seg]
                n_ex_ref  = addr["n_ex"][seg]
                n_new_ref = addr["n_new"][seg]
                n_nxt_ref = addr["n_next"][seg]
                fval = f"=({n_ex_ref}*{c_ref}+{n_new_ref}*{d_ref})/{n_nxt_ref}"
            elif len(type_filter) == 1:
                t = type_filter[0]
                fval = f"=AVERAGEIFS({pi_col},{sku_type_col},{chr(34)}{t}{chr(34)}{bl_filter})"
            else:
                # Two types — weighted average
                t1, t2 = type_filter
                n1 = f"COUNTIFS({sku_type_col},{chr(34)}{t1}{chr(34)}{bl_filter})"
                n2 = f"COUNTIFS({sku_type_col},{chr(34)}{t2}{chr(34)}{bl_filter})"
                s1 = f"SUMIFS({pi_col},{sku_type_col},{chr(34)}{t1}{chr(34)}{bl_filter})"
                s2 = f"SUMIFS({pi_col},{sku_type_col},{chr(34)}{t2}{chr(34)}{bl_filter})"
                fval = f"=({s1}+{s2})/({n1}+{n2})"

            cell = ws.cell(row=r, column=2+i, value=fval)
            cell.number_format = "#,##0.000000"
            cell.font = Font(name="Calibri", size=9)
            cell.alignment = Alignment(horizontal="right", vertical="center")

        ws.row_dimensions[r].height = 16
        _store(key, r)
        r += 1

    r += 1
    # ── Section 3: Effect per Segment ──
    _sec(r, "Section 3 — Effect per Segment (avg per SKU)"); r += 1
    _hdr(r, ["Effect","Overall","Dry","Fresh","Frozen"]); r += 1

    for lbl, key in [("1. Churned SKU Effect","eff_dep"),
                     ("2. Price Change Effect","eff_price"),
                     ("3. Comp Price Effect","eff_comp"),
                     ("  3.1 Normal Comp Price Effect","eff_normal_comp"),
                     ("  3.2 Discount (Blended) Comp Price Effect","eff_discount_comp"),
                     ("4. New SKU Effect","eff_new"),("Total Delta","total")]:
        vals = [ov[key]] + [sr[s][key] for s in segs]
        _row(r, lbl, vals)
        _store(key, r)
        r += 1

    r += 1
    # ── Section 4: Exact Contributions ──
    _sec(r, "Section 4 — Contribution to Overall (Exact Formula)"); r += 1
    _hdr(r, ["Effect","Overall","Dry","Fresh","Frozen"]); r += 1

    for lbl, key in [("1. Churned SKU Contrib","eff_dep"),
                     ("2. Price Change Contrib","eff_price"),
                     ("3. Comp Price Contrib","eff_comp"),
                     ("  3.1 Normal Comp Price Contrib","eff_normal_comp"),
                     ("  3.2 Discount (Blended) Comp Price Contrib","eff_discount_comp"),
                     ("4. New SKU Contrib","eff_new")]:
        ov_val = ov[key]  # overall effect = sum of exact contribs
        vals = [ov_val] + [contribs[s][key] for s in segs]
        _row(r, lbl, vals)
        _store("contrib_"+key, r)
        # Verify
        s_sum = sum(contribs[s][key] for s in segs)
        if abs(s_sum - ov_val) > 0.0001:
            print(f"  ⚠️  Contrib mismatch {key}: sum={s_sum:.6f} ov={ov_val:.6f}")
        r += 1

    r += 1
    # ── Section 5: Weights ──
    _sec(r, "Section 5 — Weights per Segment"); r += 1
    _hdr(r, ["Weight Type","Overall","Dry","Fresh","Frozen"]); r += 1

    for lbl, key in [("w_existing (for price/comp)","w_ex"),
                     ("w_current (for departing)","w_cur"),
                     ("w_next (for new SKU)","w_nxt")]:
        vals = [1.0] + [sr[s][key] for s in segs]
        _row(r, lbl, vals)
        r += 1

    r += 1
    # ── Section 6: Movement Summary ──
    s1      = "'1. Raw Data'"
    c_type  = f"{s1}!I:I"
    c_bl    = f"{s1}!F:F"
    c_dp    = f"{s1}!M:M"
    c_dppct = f"{s1}!N:N"
    c_dc    = f"{s1}!AC:AC"
    c_dcpct = f"{s1}!AD:AD"
    c_dg    = f"{s1}!R:R"
    c_dgpct = f"{s1}!S:S"

    def _s6_hdr(r):
        hdr(ws, r, ["Metric","Overall","Dry","Fresh","Frozen"], h=20)

    def _s6_write(r, label, formulas, fmt="#,##0", ind=2):
        sc(ws, r, 1, label, ind=ind, fs=9)
        for i, fv in enumerate(formulas):
            cell = ws.cell(row=r, column=2+i, value=fv)
            cell.number_format = fmt
            cell.font = Font(name="Calibri", size=9)
            cell.alignment = Alignment(horizontal="right", vertical="center")
        ws.row_dimensions[r].height = 16

    def _s6_sub(r, title):
        sc(ws, r, 1, title, bold=True, bg=LT_BLUE, ind=1, fs=9)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
        ws.row_dimensions[r].height = 14

    def mkf(metric, sv):
        bl2 = f',{c_bl},"{sv}"' if sv != "Overall" else ""
        ne  = f"COUNTIFS({c_type},\"Existing\"{bl2})"
        nd  = f"COUNTIFS({c_type},\"Departing SKU\"{bl2})"
        nn  = f"COUNTIFS({c_type},\"New SKU\"{bl2})"
        if   metric == "n_cur":    return f"=({ne}+{nd})"
        elif metric == "n_nxt":    return f"=({ne}+{nn})"
        elif metric == "n_ex":     return f"={ne}"
        elif metric == "n_dep":    return f"={nd}"
        elif metric == "n_new":    return f"={nn}"
        elif metric == "pct_dep":  return f"=IFERROR({nd}/({ne}+{nd}),0)"
        elif metric == "pct_new":  return f"=IFERROR({nn}/({ne}+{nn}),0)"

    def mkm(metric, sv, abs_col, pct_col):
        bl2 = f',{c_bl},"{sv}"' if sv != "Overall" else ""
        ne  = f"COUNTIFS({c_type},\"Existing\"{bl2})"
        nm  = f"(COUNTIFS({c_type},\"Existing\"{bl2},{abs_col},\">0\")+COUNTIFS({c_type},\"Existing\"{bl2},{abs_col},\"<0\"))"
        if   metric == "n":    return f"={nm}"
        elif metric == "pct":  return f"=IFERROR({nm}/{ne},0)"
        elif metric == "avg":  return f"=AVERAGEIFS({pct_col},{c_type},\"Existing\"{bl2},{abs_col},\"<>0\")"

    segs4 = ["Overall","Dry","Fresh","Frozen"]

    # Overall — 4 cols
    bg_map = {"Overall":NAVY,"Dry":"145A32","Fresh":"1A5276","Frozen":"6C3483"}
    ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=5)
    sc(ws,r,1,"  Section 6 — Movement Summary",bold=True,bg=NAVY,fc=WHITE,fs=11)
    ws.row_dimensions[r].height=24; r+=1
    _s6_hdr(r); r+=1

    _s6_sub(r, "SKU Composition"); r+=1
    for lbl,metric,fmt in [
        ("n Current Pool (ex+dep)","n_cur","#,##0"),
        ("n Existing","n_ex","#,##0"),
        ("n Departing","n_dep","#,##0"),
        ("  % dari current","pct_dep","0.0%"),
        ("n Next Pool (ex+new)","n_nxt","#,##0"),
        ("n New SKU","n_new","#,##0"),
        ("  % dari next","pct_new","0.0%"),
    ]:
        _s6_write(r, lbl, [mkf(metric,sv) for sv in segs4], fmt,
                  ind=3 if lbl.startswith("  ") else 2)
        r+=1

    for title, ac, pc in [
        ("Price Movement (existing only)", c_dp, c_dppct),
        ("Comp Movement (existing only)",  c_dc, c_dcpct),
        ("COGS Movement (existing only)",  c_dg, c_dgpct),
    ]:
        _s6_sub(r, title); r+=1
        for lbl,metric in [
            ("n SKU berubah","n"),
            ("  % dari existing","pct"),
            ("  avg change","avg"),
        ]:
            fmt = "0.00%" if metric=="avg" else ("0.0%" if metric=="pct" else "#,##0")
            _s6_write(r, lbl, [mkm(metric,sv,ac,pc) for sv in segs4], fmt,
                      ind=3 if lbl.startswith("  ") else 2)
            r+=1

    r+=1

    cw(ws, {"A":38,"B":14,"C":14,"D":14,"E":14})
    print(f"✅  Sheet 1B selesai ({time.time()-t0:.1f}s)")
    return addr

# ─────────────────────────────────────────────
# SHEET 2 — PI DECOMPOSITION  (hardcoded + ref col)
# ─────────────────────────────────────────────
def build_s2(wb, addr, ov, sr, contribs):
    print("⏳  Sheet 2 — PI Decomposition ...")
    t0   = time.time()
    ws   = wb.create_sheet("2. PI Decomposition")
    segs = ["Dry","Fresh","Frozen"]

    # Reference col starts at col 8 (gap after col 5)
    REF_START = 8

    def _val_cell(ws, r, col, val, fmt, bold=False, bg="FFFFFF", ftc=None):
        cell = ws.cell(row=r, column=col, value=val)
        cell.number_format = fmt
        # Color: positive=green, negative=red, zero/baseline/result=default
        if ftc:
            color = ftc
        elif val is not None and not bold and isinstance(val, (int, float)):
            color = GREEN_TXT if val > 0.0001 else (RED_TXT if val < -0.0001 else "000000")
        else:
            color = WHITE if bg == NAVY else "000000"
        cell.font  = Font(name="Calibri", bold=bold, size=10, color=color)
        cell.fill  = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(horizontal="right", vertical="center")
        cell.border = bb()
        return cell

    def _data_row(r, label, key, is_base=False, is_res=False, is_delta=False, is_contrib=False):
        bg  = LT_BLUE if is_base else (NAVY if is_res else ("F0F3FA" if is_delta else "FFFFFF"))
        bld = is_base or is_res or is_delta
        fmt = "#,##0.000" if (is_base or is_res) else "+#,##0.000;-#,##0.000;-"
        ftc_fixed = (WHITE if is_res else None)

        sc(ws, r, 1, label, bold=bld, bg=bg,
           fc=WHITE if is_res else "000000", ind=1, bdr=bb())
        ws.row_dimensions[r].height = 20

        ref_key = ("contrib_"+key) if is_contrib else key
        vals_map = {}
        if ref_key in addr:
            vals_map = {seg: addr[ref_key][seg] for seg in ["Overall"]+segs if seg in addr[ref_key]}

        # Hardcoded values
        if is_contrib:
            data_vals = [ov[key]] + [contribs[s][key] for s in segs]
        else:
            data_vals = [ov[key]] + [sr[s][key] for s in segs]

        for i, (seg, val) in enumerate(zip(["Overall"]+segs, data_vals)):
            _val_cell(ws, r, 2+i, val, fmt, bold=bld, bg=bg, ftc=ftc_fixed)

        # Reference col — label + source cells (greyed out, far right)
        ref_label = ws.cell(row=r, column=REF_START, value=f"← source 1B")
        ref_label.font = Font(name="Calibri", italic=True, size=7, color="C0C0C0")
        ref_label.alignment = Alignment(horizontal="left", vertical="center")
        for i, seg in enumerate(["Overall"]+segs):
            if seg in vals_map:
                rc = ws.cell(row=r, column=REF_START+1+i, value=f"={vals_map[seg]}")
                rc.number_format = fmt
                rc.font = Font(name="Calibri", italic=True, size=7, color="C0C0C0")
                rc.alignment = Alignment(horizontal="right", vertical="center")

    def _weight_subrow(r, key_w):
        sc(ws, r, 1, "  \u21b3 SKU weight", italic=True, fs=8, fc=MUTED, bg=SUB_BG)
        c_ov = ws.cell(row=r, column=2, value=1.0)
        c_ov.number_format = "0.0%"
        c_ov.font = Font(name="Calibri", italic=True, size=8, color=MUTED)
        c_ov.fill = PatternFill("solid", fgColor=SUB_BG)
        c_ov.alignment = Alignment(horizontal="right", vertical="center")
        for i, seg in enumerate(segs):
            val = sr[seg].get(key_w, 0.0)
            c = ws.cell(row=r, column=3+i, value=val)
            c.number_format = "0.0%"
            c.font = Font(name="Calibri", italic=True, size=8, color=MUTED)
            c.fill = PatternFill("solid", fgColor=SUB_BG)
            c.alignment = Alignment(horizontal="right", vertical="center")
        ws.row_dimensions[r].height = 14

    # ── TABLE 1: Waterfall ──
    r = 2
    sec(ws, r, "Tabel 1 — Waterfall Avg PI per Pricing BL 25", 5)
    r += 1
    hdr(ws, r, ["Effect","Overall","Dry","Fresh","Frozen"])
    # Reference header
    sc(ws, r, REF_START, "Source (1B ref)", italic=True, fs=7, fc="C0C0C0", align="center")
    for i, seg in enumerate(["Overall","Dry","Fresh","Frozen"]):
        sc(ws, r, REF_START+1+i, seg, italic=True, fs=7, fc="C0C0C0", align="center")
    r += 1

    for label, key, ib, ir, id_ in [
        ("Baseline — Avg PI Prev (A)",       "A",                  True,  False, False),
        ("+ 1. Churned SKU Effect",          "eff_dep",            False, False, False),
        ("+ 2. Price Change Effect",         "eff_price",          False, False, False),
        ("+ 3. Comp Price Effect",  "eff_comp",           False, False, False),
        ("    3.1 Normal Comp Price Effect",       "eff_normal_comp",    False, False, False),
        ("    3.2 Discount (Blended) Comp Price Effect", "eff_discount_comp", False, False, False),
        ("+ 4. New SKU Effect",              "eff_new",            False, False, False),
        ("Result — Avg PI Current (E)",      "E",                  False, True,  False),
        ("Total Delta",                       "total",              False, False, True),
    ]:
        _data_row(r, label, key, ib, ir, id_)
        r += 1

    r += 1
    # ── TABLE 2: Contributions ──
    sec(ws, r, "Tabel 2 — Kontribusi per Segment ke Overall Effect (Exact)", 5)
    r += 1
    hdr(ws, r, ["Effect","Overall","Dry","Fresh","Frozen"])
    r += 1
    _weight_subrow(r, "w_ex")
    r += 1

    for label, key, wkey in [
        ("1. Churned SKU Effect",                  "eff_dep",            "w_cur"),
        ("2. Price Change Effect",                 "eff_price",          "w_ex"),
        ("3. Comp Price Effect",          "eff_comp",           "w_ex"),
        ("  3.1 Normal Comp Price Effect",               "eff_normal_comp",    "w_ex"),
        ("  3.2 Discount (Blended) Comp Price Effect",   "eff_discount_comp",  "w_ex"),
        ("4. New SKU Effect",                      "eff_new",            "w_nxt"),
    ]:
        _data_row(r, label, key, is_contrib=True)
        r += 1
        _weight_subrow(r, wkey)
        r += 1

    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    c = ws.cell(row=r, column=1,
        value="* Kontribusi exact: price/comp = sum_effect_seg/n_ex_total | "
              "departing = sum_pi_ex_seg/n_ex - sum_pi_cur_seg/n_cur | "
              "new = (sum_npi_ex_seg+sum_npi_new_seg)/n_next - sum_npi_ex_seg/n_ex")
    c.font = Font(name="Calibri", italic=True, size=8, color=MUTED)
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[r].height = 28

    # Column widths
    cw(ws, {"A":40,"B":14,"C":14,"D":14,"E":14,"F":4,"G":4,
            "H":14,"I":14,"J":14,"K":14,"L":14})
    ws.freeze_panes = "B1"
    print(f"✅  Sheet 2 selesai ({time.time()-t0:.1f}s)")

# ─────────────────────────────────────────────
# HELPER: DIMENSION SHEET (S3/S4)
# ─────────────────────────────────────────────
def _dim_sheet(ws, d, dim_col, dim_label):
    dims = sorted(d[dim_col].dropna().unique(), key=str)
    hdrs = [dim_label,"n Existing","n Departing","n New",
            "n Comp Down","n Comp Stay","n Comp Up",
            "Avg PI Prev","Avg PI Cur","Total Delta",
            "1. Churned Eff","2. Price Eff","3. Comp Price Eff",
            "3.1 Normal Comp Price Eff","3.2 Discount (Blended) Comp Price Eff","4. New SKU Eff"]
    hdr(ws, 1, hdrs, h=30)
    ws.freeze_panes = "B2"

    rows_data = []
    total = len(dims)
    for i, dv in enumerate(dims):
        if i % 10 == 0: progress(f"Computing {dim_label}...", i, total)
        sub = d[d[dim_col]==dv]
        res = decompose(sub)
        ex_sub = sub[sub["sku_type"]=="Existing"]
        rows_data.append((dv, res,
                          (ex_sub["comp_tag"]=="Down").sum(),
                          (ex_sub["comp_tag"]=="Stay").sum(),
                          (ex_sub["comp_tag"]=="Up").sum()))
    rows_data.append(("OVERALL", decompose(d),
                      (d[d["sku_type"]=="Existing"]["comp_tag"]=="Down").sum(),
                      (d[d["sku_type"]=="Existing"]["comp_tag"]=="Stay").sum(),
                      (d[d["sku_type"]=="Existing"]["comp_tag"]=="Up").sum()))
    progress(f"Writing {dim_label}...", total, total); print()

    for r_idx, (dv, res, ncd, ncs, ncu) in enumerate(rows_data, 2):
        is_ov = dv == "OVERALL"
        bg  = NAVY if is_ov else ("FFFFFF" if r_idx%2==0 else "F5F7FF")
        ftc = WHITE if is_ov else "000000"
        vals = [dv, res["n_ex"], res["n_dep"], res["n_new"], ncd, ncs, ncu,
                res["A"], res["E"], res["total"],
                res["eff_dep"], res["eff_price"], res["eff_comp"],
                res["eff_normal_comp"], res["eff_discount_comp"],
                res["eff_new"]]
        for ci, val in enumerate(vals, 1):
            cell = ws.cell(row=r_idx, column=ci, value=val)
            cell.font = Font(name="Calibri", bold=is_ov, size=9, color=ftc)
            cell.fill = PatternFill("solid", fgColor=bg)
            cell.alignment = Alignment(
                horizontal="left" if ci==1 else "right",
                vertical="center", indent=1 if ci==1 else 0)
            if ci in {2,3,4,5,6,7}: cell.number_format = "#,##0"
            elif ci >= 8:
                cell.number_format = "#,##0.000"
                if not is_ov and ci >= 10 and isinstance(val, float):
                    if val < -0.0001: cell.font = Font(name="Calibri",size=9,color=RED_TXT)
                    elif val > 0.0001: cell.font = Font(name="Calibri",size=9,color=GREEN_TXT)
        ws.row_dimensions[r_idx].height = 16

    ws.column_dimensions["A"].width = 28
    for i in range(2, len(hdrs)+1):
        ws.column_dimensions[get_column_letter(i)].width = 12

def build_s3(wb, d):
    print("⏳  Sheet 3 — L1 Category ...")
    t0 = time.time()
    ws = wb.create_sheet("3. L1 Category")
    _dim_sheet(ws, d, "l1_category_name", "L1 Category")
    print(f"✅  Sheet 3 selesai ({time.time()-t0:.1f}s)")

def build_s4(wb, d):
    print("⏳  Sheet 4 — Pareto ...")
    t0 = time.time()
    ws = wb.create_sheet("4. Pareto")
    _dim_sheet(ws, d, "pareto_classification", "Pareto")
    print(f"✅  Sheet 4 selesai ({time.time()-t0:.1f}s)")

# ─────────────────────────────────────────────
# HELPER: 3x3 MATRIX
# ─────────────────────────────────────────────
def _matrix_3x3(ws, d, title_text, start_row=1, seg=None):
    DIRS = ["Down","Stay","Up"]
    ncols = 10
    sec(ws, start_row, title_text, ncols, seg=seg)
    r = start_row + 1

    ex = d[d["sku_type"]=="Existing"].copy()
    total = len(ex)
    _, seg_lt = SEG_COLORS.get(seg or "Overall", SEG_COLORS["Overall"])
    tl = Border(left=_s("medium","1C4587"), bottom=_s())
    tlo = Border(left=_s("medium","1C4587"))

    matrix = np.zeros((3,3), dtype=int)
    for ri, pr in enumerate(DIRS):
        for ci, cr in enumerate(DIRS):
            matrix[ri,ci] = int(((ex["price_tag"]==pr)&(ex["comp_tag"]==cr)).sum())
    row_tots = [int(matrix[ri,:].sum()) for ri in range(3)]
    col_tots = [int(matrix[:,ci].sum()) for ci in range(3)]

    # header — col1=blank(row label), col2=blank(dir label), col3-5=Down/Stay/Up, col6=Total
    #           col7=Down%, col8=Stay%, col9=Up%, col10=Total%
    for i, h in enumerate(["",""] + DIRS + ["Total"]):
        sc(ws,r,1+i,h,bold=True,bg=NAVY_LT,fc=WHITE,align="center",wrap=True,fs=9)
    for i, h in enumerate(DIRS+["Total %"]):
        c = sc(ws,r,7+i,h,bold=True,bg=NAVY_LT,fc=WHITE,align="center",wrap=True,fs=9)
        if i==0: c.border = tlo
    ws.row_dimensions[r].height = 22; r += 1

    for ri, pr in enumerate(DIRS):
        if ri==0:
            ws.merge_cells(start_row=r,start_column=1,end_row=r+2,end_column=1)
            sc(ws,r,1,"PRICE KITA",bold=True,bg=seg_lt,align="center",wrap=True)
        sc(ws,r,2,pr,bold=True,bg=seg_lt,align="center",fs=9)
        for ci in range(3):
            v = int(matrix[ri,ci])
            bg = "FFF0F0" if (ri==0 and ci==0) else (seg_lt if ri==ci else "FFFFFF")
            sc(ws,r,3+ci,v,bg=bg,align="right",fmt="#,##0",bdr=bb(),fs=9)
        sc(ws,r,6,row_tots[ri],bold=True,bg="F0F3FA",align="right",fmt="#,##0",bdr=bb(),fs=9)
        for ci in range(3):
            pv = matrix[ri,ci]/total if total>0 else 0
            bg = "FFF0F0" if (ri==0 and ci==0) else (seg_lt if ri==ci else "FFFFFF")
            c = sc(ws,r,7+ci,pv,bg=bg,align="right",fmt="0.0%",bdr=bb(),fs=9)
            if ci==0: c.border = tl
        sc(ws,r,10,row_tots[ri]/total if total>0 else 0,
           bold=True,bg="F0F3FA",align="right",fmt="0.0%",bdr=bb(),fs=9)
        r += 1

    sc(ws,r,1,"Total",bold=True,bg=NAVY,fc=WHITE,fs=9)
    sc(ws,r,2,"",bg=NAVY)
    for ci in range(3):
        sc(ws,r,3+ci,col_tots[ci],bold=True,bg=NAVY,fc=WHITE,align="right",fmt="#,##0",fs=9)
    sc(ws,r,6,total,bold=True,bg=NAVY,fc=WHITE,align="right",fmt="#,##0",fs=9)
    for ci in range(3):
        c = sc(ws,r,7+ci,col_tots[ci]/total if total>0 else 0,
               bold=True,bg=NAVY,fc=WHITE,align="right",fmt="0.0%",fs=9)
        if ci==0: c.border = Border(left=_s("medium",WHITE))
    sc(ws,r,10,1.0,bold=True,bg=NAVY,fc=WHITE,align="right",fmt="0.0%",fs=9)
    r += 1
    return r + 2

def build_s5(wb, d):
    print("⏳  Sheet 5 — Price vs Comp Matrix ...")
    t0 = time.time()
    ws = wb.create_sheet("5. Price vs Comp Matrix")
    r = _matrix_3x3(ws, d, "Price vs Comp Movement Matrix", seg="Overall")
    for seg in ["Dry","Fresh","Frozen"]:
        r = _matrix_3x3(ws, d[d["pricing_bl_25"]==seg],
                        "Price vs Comp Movement Matrix", r, seg=seg)
    cw(ws, {"A":14,"B":10,"C":12,"D":12,"E":12,"F":12,
            "G":12,"H":12,"I":12,"J":12})
    print(f"✅  Sheet 5 selesai ({time.time()-t0:.1f}s)")

# ─────────────────────────────────────────────
# HELPER: CROSS TABLE  (side-by-side count + %)
# ─────────────────────────────────────────────
def _cross_table(ws, d_sub, row_labels, col_lbls, row_col, col_col,
                 title_text, start_row, n_total=None, seg=None):
    n_cols = len(col_lbls)
    ncols  = 1 + n_cols + 1 + n_cols + 1
    sec(ws, start_row, title_text, ncols, seg=seg)
    r = start_row + 1

    _, seg_lt = SEG_COLORS.get(seg or "Overall", SEG_COLORS["Overall"])
    tl  = Border(left=_s("medium","1C4587"), bottom=_s())
    tlo = Border(left=_s("medium","1C4587"))

    # compute
    rc_sub = d_sub[row_col].astype(str)
    cc_sub = d_sub[col_col].astype(str)
    counts = {}
    for rl in row_labels:
        counts[rl] = {cl: int(((rc_sub==rl)&(cc_sub==cl)).sum()) for cl in col_lbls}
    row_tots = {rl: sum(counts[rl].values()) for rl in row_labels}
    col_tots = {cl: sum(counts[rl][cl] for rl in row_labels) for cl in col_lbls}
    grand    = sum(row_tots.values())
    if n_total is None: n_total = grand
    off = n_cols + 2  # start col of % block

    # header
    for i, h in enumerate([""]+col_lbls+["Total"]):
        sc(ws,r,1+i,h,bold=True,bg=NAVY_LT,fc=WHITE,align="center",wrap=True,fs=9)
    for i, h in enumerate(col_lbls+["Total %"]):
        c = sc(ws,r,off+1+i,h,bold=True,bg=NAVY_LT,fc=WHITE,align="center",wrap=True,fs=9)
        if i==0: c.border = tlo
    ws.row_dimensions[r].height = 28; r += 1

    for ri, rl in enumerate(row_labels):
        bg_row = "FFFFFF" if ri%2==0 else "FAFBFF"
        sc(ws,r,1,rl,bg=bg_row,ind=1,bdr=bb(),fs=9)
        max_v = max(counts[rl].values()) if counts[rl] else 0
        for ci, cl in enumerate(col_lbls):
            v  = counts[rl][cl]
            bg = seg_lt if (v==max_v and v>0) else bg_row
            sc(ws,r,2+ci,v,bg=bg,align="right",fmt="#,##0",bdr=bb(),fs=9)
        sc(ws,r,1+n_cols+1,row_tots[rl],bold=True,bg="F0F3FA",
           align="right",fmt="#,##0",bdr=bb(),fs=9)
        for ci, cl in enumerate(col_lbls):
            v  = counts[rl][cl]/n_total if n_total>0 else 0
            bg = seg_lt if (counts[rl][cl]==max_v and counts[rl][cl]>0) else bg_row
            c  = sc(ws,r,off+1+ci,v,bg=bg,align="right",fmt="0.0%",bdr=bb(),fs=9)
            if ci==0: c.border = tl
        sc(ws,r,off+1+n_cols,row_tots[rl]/n_total if n_total>0 else 0,
           bold=True,bg="F0F3FA",align="right",fmt="0.0%",bdr=bb(),fs=9)
        r += 1

    sc(ws,r,1,"Total",bold=True,bg=NAVY,fc=WHITE,fs=9)
    for ci,cl in enumerate(col_lbls):
        sc(ws,r,2+ci,col_tots[cl],bold=True,bg=NAVY,fc=WHITE,align="right",fmt="#,##0",fs=9)
    sc(ws,r,1+n_cols+1,grand,bold=True,bg=NAVY,fc=WHITE,align="right",fmt="#,##0",fs=9)
    for ci,cl in enumerate(col_lbls):
        c = sc(ws,r,off+1+ci,col_tots[cl]/n_total if n_total>0 else 0,
               bold=True,bg=NAVY,fc=WHITE,align="right",fmt="0.0%",fs=9)
        if ci==0: c.border = Border(left=_s("medium",WHITE))
    sc(ws,r,off+1+n_cols,1.0,bold=True,bg=NAVY,fc=WHITE,align="right",fmt="0.0%",fs=9)
    r += 1
    return r + 2

def build_s7(wb, d):
    print("⏳  Sheet 7 — PI vs Margin ...")
    t0 = time.time()
    ws = wb.create_sheet("7. PI vs Margin")
    ex = d[d["next_pi"].notna() & d["margin_pct_cur"].notna()].copy()
    r  = 1
    for seg_label, sf in [("Overall",None),("Dry","Dry"),("Fresh","Fresh"),("Frozen","Frozen")]:
        sub = ex if sf is None else ex[ex["pricing_bl_25"]==sf]
        r = _cross_table(ws, sub, MG_LBL, PI_LBL,
                         "margin_group_cur","pi_group_cur",
                         "PI Group (current) vs Margin Group (current)",
                         r, n_total=len(sub), seg=seg_label)
    ws.column_dimensions["A"].width = 16
    for i in range(2, len(PI_LBL)*2+4):
        ws.column_dimensions[get_column_letter(i)].width = 10
    print(f"✅  Sheet 7 selesai ({time.time()-t0:.1f}s)")

def build_s8(wb, d):
    print("⏳  Sheet 8 — COGS Index vs PI ...")
    t0 = time.time()
    ws = wb.create_sheet("8. COGS Index vs PI")
    ex = d[d["next_pi"].notna() & d["ci_cur"].notna()].copy()
    r  = 1
    for seg_label, sf in [("Overall",None),("Dry","Dry"),("Fresh","Fresh"),("Frozen","Frozen")]:
        sub = ex if sf is None else ex[ex["pricing_bl_25"]==sf]
        r = _cross_table(ws, sub, CI_LBL, PI_LBL,
                         "ci_group_cur","pi_group_cur",
                         "COGS Index Group (current) vs PI Group (current)",
                         r, n_total=len(sub), seg=seg_label)
    ws.column_dimensions["A"].width = 14
    for i in range(2, len(PI_LBL)*2+4):
        ws.column_dimensions[get_column_letter(i)].width = 10
    print(f"✅  Sheet 8 selesai ({time.time()-t0:.1f}s)")

def build_s9(wb, d):
    print("⏳  Sheet 9 — COGS Index vs Margin ...")
    t0 = time.time()
    ws = wb.create_sheet("9. COGS Index vs Margin")
    ex = d[d["margin_pct_cur"].notna() & d["ci_cur"].notna()].copy()
    r  = 1
    for seg_label, sf in [("Overall",None),("Dry","Dry"),("Fresh","Fresh"),("Frozen","Frozen")]:
        sub = ex if sf is None else ex[ex["pricing_bl_25"]==sf]
        r = _cross_table(ws, sub, CI_LBL, MG_LBL,
                         "ci_group_cur","margin_group_cur",
                         "COGS Index Group (current) vs Margin Group (current)",
                         r, n_total=len(sub), seg=seg_label)
    ws.column_dimensions["A"].width = 14
    for i in range(2, len(MG_LBL)*2+4):
        ws.column_dimensions[get_column_letter(i)].width = 11
    print(f"✅  Sheet 9 selesai ({time.time()-t0:.1f}s)")

# ─────────────────────────────────────────────
# SHEET 6 — PI DISTRIBUTION & MOVEMENT
# ─────────────────────────────────────────────
def build_s6(wb, d):
    print("⏳  Sheet 6 — PI Distribution & Movement ...")
    t0 = time.time()
    ws = wb.create_sheet("6. PI Distribution")
    ex = d[d["pi"].notna() & d["next_pi"].notna()].copy()
    tl  = Border(left=_s("medium","1C4587"), bottom=_s())
    tlo = Border(left=_s("medium","1C4587"))

    def _section(sub, seg_key, start_r):
        _, seg_lt = SEG_COLORS.get(seg_key, SEG_COLORS["Overall"])
        sec(ws, start_r, "PI Distribution & Movement", 16, seg=seg_key)
        r = start_r + 1
        n = len(sub)

        # Tabel 1: distribution side by side
        for i, h in enumerate(["PI Group","n Prev","n Cur","Delta"]):
            sc(ws,r,1+i,h,bold=True,bg=NAVY_LT,fc=WHITE,align="center",wrap=True,fs=9)
        for i, h in enumerate(["% Prev","% Cur","Delta %"]):
            c = sc(ws,r,5+i,h,bold=True,bg=NAVY_LT,fc=WHITE,align="center",wrap=True,fs=9)
            if i==0: c.border = tlo
        ws.row_dimensions[r].height = 22; r += 1

        for lbl in PI_LBL:
            nc = int((sub["pi_group_prev"].astype(str)==lbl).sum())
            nn = int((sub["pi_group_cur"].astype(str)==lbl).sum())
            dlt = nn - nc
            sc(ws,r,1,lbl,ind=1,bdr=bb(),fs=9)
            sc(ws,r,2,nc,align="right",fmt="#,##0",bdr=bb(),fs=9)
            sc(ws,r,3,nn,align="right",fmt="#,##0",bdr=bb(),fs=9)
            cd = ws.cell(row=r,column=4,value=dlt)
            cd.number_format="+#,##0;-#,##0;-"
            cd.font=Font(name="Calibri",size=9,
                         color=GREEN_TXT if dlt>0 else (RED_TXT if dlt<0 else "000000"))
            cd.alignment=Alignment(horizontal="right",vertical="center")
            cd.border=bb()
            cp1=sc(ws,r,5,nc/n if n>0 else 0,align="right",fmt="0.0%",
                   bdr=Border(left=_s("medium","1C4587"),bottom=_s()),fs=9)
            sc(ws,r,6,nn/n if n>0 else 0,align="right",fmt="0.0%",bdr=bb(),fs=9)
            dpct = (nn-nc)/n if n>0 else 0
            cp2=ws.cell(row=r,column=7,value=dpct)
            cp2.number_format="+0.0%;-0.0%;-"
            cp2.font=Font(name="Calibri",size=9,
                          color=GREEN_TXT if dpct>0 else (RED_TXT if dpct<0 else "000000"))
            cp2.alignment=Alignment(horizontal="right",vertical="center")
            cp2.border=bb()
            r += 1

        sc(ws,r,1,"Total",bold=True,bg=NAVY,fc=WHITE,fs=9)
        sc(ws,r,2,n,bold=True,bg=NAVY,fc=WHITE,align="right",fmt="#,##0",fs=9)
        sc(ws,r,3,n,bold=True,bg=NAVY,fc=WHITE,align="right",fmt="#,##0",fs=9)
        sc(ws,r,4,0,bold=True,bg=NAVY,fc=WHITE,align="right",fs=9)
        c5=sc(ws,r,5,1.0,bold=True,bg=NAVY,fc=WHITE,align="right",fmt="0.0%",fs=9)
        c5.border=Border(left=_s("medium",WHITE))
        sc(ws,r,6,1.0,bold=True,bg=NAVY,fc=WHITE,align="right",fmt="0.0%",fs=9)
        sc(ws,r,7,0.0,bold=True,bg=NAVY,fc=WHITE,align="right",fmt="0.0%",fs=9)
        r += 2

        # Tabel 2: movement matrix side by side
        npl = len(PI_LBL)
        off = npl + 2
        for i, h in enumerate(["PI Prev \\ PI Cur"]+PI_LBL+["Total"]):
            sc(ws,r,1+i,h,bold=True,bg=NAVY_LT,fc=WHITE,align="center",wrap=True,fs=9)
        for i, h in enumerate(PI_LBL+["Total %"]):
            c=sc(ws,r,off+1+i,h,bold=True,bg=NAVY_LT,fc=WHITE,align="center",wrap=True,fs=9)
            if i==0: c.border=tlo
        ws.row_dimensions[r].height=28; r+=1

        mc = {}
        for rl in PI_LBL:
            mc[rl] = {cl: int(((sub["pi_group_prev"].astype(str)==rl)&
                                (sub["pi_group_cur"].astype(str)==cl)).sum())
                      for cl in PI_LBL}
        ct_mv = {cl: sum(mc[rl][cl] for rl in PI_LBL) for cl in PI_LBL}

        for rl in PI_LBL:
            rt = sum(mc[rl].values())
            sc(ws,r,1,rl,ind=1,bdr=bb(),fs=9)
            for ci,cl in enumerate(PI_LBL):
                v = mc[rl][cl]
                od = (rl==cl)
                md = PI_LBL.index(cl)<PI_LBL.index(rl)
                bg = seg_lt if od else ("FFF0F0" if md and v>0 else "FFFFFF")
                sc(ws,r,2+ci,v,bg=bg,align="right",fmt="#,##0",bdr=bb(),fs=9)
            sc(ws,r,1+npl+1,rt,bold=True,bg="F0F3FA",align="right",fmt="#,##0",bdr=bb(),fs=9)
            for ci,cl in enumerate(PI_LBL):
                pv=mc[rl][cl]/n if n>0 else 0
                od=(rl==cl); md=PI_LBL.index(cl)<PI_LBL.index(rl)
                bg=seg_lt if od else ("FFF0F0" if md and mc[rl][cl]>0 else "FFFFFF")
                c=sc(ws,r,off+1+ci,pv,bg=bg,align="right",fmt="0.0%",bdr=bb(),fs=9)
                if ci==0: c.border=tl
            sc(ws,r,off+1+npl,rt/n if n>0 else 0,bold=True,bg="F0F3FA",
               align="right",fmt="0.0%",bdr=bb(),fs=9)
            r+=1

        sc(ws,r,1,"Total",bold=True,bg=NAVY,fc=WHITE,fs=9)
        for ci,cl in enumerate(PI_LBL):
            sc(ws,r,2+ci,ct_mv[cl],bold=True,bg=NAVY,fc=WHITE,align="right",fmt="#,##0",fs=9)
        sc(ws,r,1+npl+1,n,bold=True,bg=NAVY,fc=WHITE,align="right",fmt="#,##0",fs=9)
        for ci,cl in enumerate(PI_LBL):
            c=sc(ws,r,off+1+ci,ct_mv[cl]/n if n>0 else 0,
                 bold=True,bg=NAVY,fc=WHITE,align="right",fmt="0.0%",fs=9)
            if ci==0: c.border=Border(left=_s("medium",WHITE))
        sc(ws,r,off+1+npl,1.0,bold=True,bg=NAVY,fc=WHITE,align="right",fmt="0.0%",fs=9)
        r+=2

        # Tabel 3: why moved
        for i,h in enumerate(["Movement","n SKU","Avg Eff Price","Avg Eff Comp",
                               "Avg Delta PI"]):
            sc(ws,r,1+i,h,bold=True,bg=NAVY_LT,fc=WHITE,align="center",wrap=True,fs=9)
        ws.row_dimensions[r].height=22; r+=1

        sub2 = sub.copy()
        sub2["pci"] = sub2["pi_group_prev"].astype(str).apply(
            lambda x: PI_LBL.index(x) if x in PI_LBL else -1)
        sub2["nci"] = sub2["pi_group_cur"].astype(str).apply(
            lambda x: PI_LBL.index(x) if x in PI_LBL else -1)
        sub2["mv"]  = sub2["nci"].astype(int) - sub2["pci"].astype(int)
        ex2 = sub2[sub2["pi"].notna()&sub2["next_pi"].notna()].copy()

        # Shapley
        _comp0  = ex2["comp_price"].replace(0, np.nan)
        _comp1  = ex2["next_comp_price"].replace(0, np.nan)
        _pmid_B = ex2["next_price"] / _comp0 * 100
        _pmid_D = ex2["price"] / _comp1 * 100
        ex2["ep"]  = ((_pmid_B - ex2["pi"]) + (ex2["next_pi"] - _pmid_D)) / 2
        ex2["ec"]  = ((ex2["next_pi"] - _pmid_B) + (_pmid_D - ex2["pi"])) / 2
        ex2["dpi"] = ex2["next_pi"] - ex2["pi"]

        for ri2,(lbl3,mask) in enumerate([
            ("Naik 2+ bucket", ex2["mv"]>=2),
            ("Naik 1 bucket",  ex2["mv"]==1),
            ("Tetap",          ex2["mv"]==0),
            ("Turun 1 bucket", ex2["mv"]==-1),
            ("Turun 2+ bucket",ex2["mv"]<=-2),
        ]):
            sub3=ex2[mask]; nm=len(sub3)
            bg="FFFFFF" if ri2%2==0 else "FAFBFF"
            vals=[lbl3,nm,
                  sub3["ep"].mean() if nm>0 else 0,
                  sub3["ec"].mean() if nm>0 else 0,
                  sub3["dpi"].mean() if nm>0 else 0]
            for ci3,v in enumerate(vals,1):
                fmt="#,##0" if ci3==2 else ("+#,##0.000;-#,##0.000;-" if ci3>2 else "General")
                cell=ws.cell(row=r,column=ci3,value=v)
                cell.number_format=fmt
                cell.font=Font(name="Calibri",size=9,
                               color=(RED_TXT if ci3>2 and isinstance(v,float) and v<-0.0001 else
                                      (GREEN_TXT if ci3>2 and isinstance(v,float) and v>0.0001 else "000000")))
                cell.fill=PatternFill("solid",fgColor=bg)
                cell.alignment=Alignment(horizontal="right" if ci3>1 else "left",
                                         vertical="center",indent=1 if ci3==1 else 0)
                cell.border=bb()
            r+=1
        return r+2

    r = _section(ex, "Overall", 1)
    for seg in ["Dry","Fresh","Frozen"]:
        r = _section(ex[ex["pricing_bl_25"]==seg], seg, r)

    ws.column_dimensions["A"].width = 18
    for i in range(2, 18):
        ws.column_dimensions[get_column_letter(i)].width = 10
    print(f"✅  Sheet 6 selesai ({time.time()-t0:.1f}s)")

# ─────────────────────────────────────────────
# SHEET 10 — STRUCTURAL LOSS
# ─────────────────────────────────────────────
def build_s10(wb, d):
    print("⏳  Sheet 10 — Structural Loss ...")
    t0 = time.time()
    ws = wb.create_sheet("10. Structural Loss")
    ex = d[d["ci_cur"].notna() & d["next_pi"].notna() & d["margin_pct_cur"].notna()].copy()
    cat_total = ex.groupby("l1_category_name").size().to_dict()
    r = 1

    for grp_label, grp_filter in [("COGS Index D (95-105)","D.95-105"),
                                   ("COGS Index E (>105)","E.>105")]:
        sec(ws, r, grp_label, 7)
        r += 1
        hdr(ws, r, ["L1 Category","n SKU","% of Category",
                    "Avg COGS Index","Avg Margin %","Avg PI (current)"], h=24)
        r += 1

        sub = ex[ex["ci_group_cur"]==grp_filter].copy()
        by_cat = sub.groupby("l1_category_name").agg(
            n_sku=("product_id","count"),
            avg_ci=("ci_cur","mean"),
            avg_mg=("margin_pct_cur","mean"),
            avg_pi=("next_pi","mean")
        ).sort_values("n_sku",ascending=False).reset_index()

        for _, row in by_cat.iterrows():
            pct = row["n_sku"] / cat_total.get(row["l1_category_name"], 1)
            bg  = "FFF0F0" if row["avg_ci"]>105 else "FFFFFF"
            for ci,v in enumerate([row["l1_category_name"],int(row["n_sku"]),pct,
                                    row["avg_ci"],row["avg_mg"],row["avg_pi"]],1):
                fmt=("#,##0" if ci==2 else "0.0%" if ci in {3,5} else "#,##0.00")
                sc(ws,r,ci,v,bg=bg,align="right" if ci>1 else "left",
                   fmt=fmt,bdr=bb(),ind=1 if ci==1 else 0,fs=9)
            r+=1

        tot_n=int(len(sub))
        for ci,v in enumerate(["TOTAL",tot_n,tot_n/len(ex) if len(ex)>0 else 0,
                                sub["ci_cur"].mean(),sub["margin_pct_cur"].mean(),
                                sub["next_pi"].mean()],1):
            fmt=("#,##0" if ci==2 else "0.0%" if ci in {3,5} else "#,##0.00")
            sc(ws,r,ci,v,bold=True,bg=NAVY,fc=WHITE,
               align="right" if ci>1 else "left",fmt=fmt,fs=9)
        r+=3

    cw(ws,{"A":28,"B":10,"C":16,"D":16,"E":14,"F":16})
    print(f"✅  Sheet 10 selesai ({time.time()-t0:.1f}s)")

# ─────────────────────────────────────────────
# SHEET 11 — COGS NEED IMPROVE
# ─────────────────────────────────────────────
def build_s11(wb, d):
    print("⏳  Sheet 11 — COGS Need Improve ...")
    t0 = time.time()
    ws = wb.create_sheet("11. COGS Need Improve")
    ex = d[d["ci_group_cur"].isin(["D.95-105","E.>105"])].copy()
    ex = ex.sort_values("ci_cur", ascending=False)

    id_cols = ["product_id","product_name","l1_category_name",
               "pricing_bl_25","pareto_classification","sku_type"]
    met_cols = ["ci_group_cur","ci_cur","next_pi","margin_pct_cur",
                "next_price","next_cogs","next_comp_price"]
    all_cols = [c for c in id_cols+met_cols if c in ex.columns]

    labels = {
        "product_id":"Product ID","product_name":"Product Name",
        "l1_category_name":"L1 Category","pricing_bl_25":"Pricing BL 25",
        "pareto_classification":"Pareto","sku_type":"SKU Type",
        "ci_group_cur":"COGS Idx Group","ci_cur":"COGS Index (cur)",
        "next_pi":"PI (current)","margin_pct_cur":"Margin % (cur)",
        "next_price":"Price (cur)","next_cogs":"COGS (cur)",
        "next_comp_price":"Comp Price (cur)"
    }

    hdr(ws, 1, [labels.get(c,c) for c in all_cols], h=28)
    ws.freeze_panes = "A2"

    pct_cols = {"margin_pct_cur"}
    num_cols = {"ci_cur","next_pi","next_price","next_cogs","next_comp_price"}

    for r_idx, row in enumerate(ex[all_cols].itertuples(index=False), 2):
        is_e = str(getattr(row,"ci_group_cur",""))=="E.>105"
        bg   = "FFF0F0" if is_e else "FFF8F0"
        for ci, c_name in enumerate(all_cols, 1):
            val = getattr(row, c_name, None)
            if isinstance(val,float) and np.isnan(val): val=None
            cell = ws.cell(row=r_idx, column=ci, value=val)
            cell.font = Font(name="Calibri", size=9)
            cell.fill = PatternFill("solid", fgColor=bg)
            cell.alignment = Alignment(
                horizontal="right" if c_name in pct_cols or c_name in num_cols else "left",
                vertical="center", indent=1 if ci==1 else 0)
            if c_name in pct_cols:    cell.number_format="0.0%"
            elif c_name in num_cols:  cell.number_format="#,##0.00"
        ws.row_dimensions[r_idx].height = 16

    widths={"product_name":28,"l1_category_name":22,"pricing_bl_25":12,
            "pareto_classification":14,"sku_type":14,"ci_group_cur":16,
            "ci_cur":16,"next_pi":12,"margin_pct_cur":14,
            "next_price":14,"next_cogs":12,"next_comp_price":16}
    for ci,c in enumerate(all_cols,1):
        ws.column_dimensions[get_column_letter(ci)].width=widths.get(c,12)

    print(f"✅  Sheet 11 selesai ({time.time()-t0:.1f}s) — {len(ex):,} SKU")

# ─────────────────────────────────────────────
# SHEET 12 — FORMULA DOCUMENTATION
# ─────────────────────────────────────────────
def build_s12(wb, ov, sr):
    print("⏳  Sheet 12 — Formula Documentation ...")
    t0   = time.time()
    ws   = wb.create_sheet("12. Formula")
    segs = ["Dry","Fresh","Frozen"]

    def _t(r, txt):
        ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=6)
        sc(ws,r,1,txt,bold=True,bg=NAVY,fc=WHITE,fs=11,ind=1)
        ws.row_dimensions[r].height=22

    def _s2(r, txt):
        ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=6)
        sc(ws,r,1,txt,bold=True,bg=LT_BLUE,fc=NAVY,fs=10,ind=1)
        ws.row_dimensions[r].height=18

    def _f(r, name, formula, note=""):
        sc(ws,r,1,name,bold=True,fs=9,ind=1)
        c2=ws.cell(row=r,column=2); c2.number_format="@"
        c2.value=formula
        c2.font=Font(name="Calibri",size=9,color=NAVY)
        c2.alignment=Alignment(horizontal="left",vertical="center",wrap_text=True)
        ws.merge_cells(start_row=r,start_column=2,end_row=r,end_column=5)
        if note:
            c6=ws.cell(row=r,column=6,value=note)
            c6.font=Font(name="Calibri",italic=True,size=8,color=MUTED)
            c6.alignment=Alignment(horizontal="left",vertical="center",wrap_text=True)
        ws.row_dimensions[r].height=18

    r=1
    _t(r,"Dokumentasi Formula — PI Analysis"); r+=2

    _s2(r,"A. Definisi Dasar"); r+=1
    _f(r,"PI",           "price x 100 / comp_price"); r+=1
    _f(r,"Margin",       "price - cogs"); r+=1
    _f(r,"Margin %",     "(price - cogs) / price"); r+=1
    _f(r,"COGS Index",   "cogs x 100 / comp_price",
       "Proxy: kalau jual di harga comp, margin kita berapa?"); r+=1
    _f(r,"Diff",         "current_value - prev_value",
       "berlaku untuk price, cogs, comp, PI, margin"); r+=1
    _f(r,"Diff %",       "diff / prev_value"); r+=2

    _s2(r,"B. Konvensi Penamaan Periode"); r+=1
    _f(r,"prev",    "Data periode lama (kolom asli: price, cogs, comp_price, pi, dst)"); r+=1
    _f(r,"current", "Data periode terbaru (kolom asli: next_price, next_cogs, dst)"); r+=2

    _s2(r,"C. Klasifikasi SKU"); r+=1
    _f(r,"Existing SKU",  "pi NOT NULL  AND  next_pi NOT NULL", f"n={ov['n_ex']:,}"); r+=1
    _f(r,"Departing SKU", "pi NOT NULL  AND  next_pi IS NULL",  f"n={ov['n_dep']:,}"); r+=1
    _f(r,"New SKU",       "pi IS NULL   AND  next_pi NOT NULL", f"n={ov['n_new']:,}"); r+=2

    _s2(r,"D. Tagging Up / Stay / Down"); r+=1
    _f(r,"Up",   "delta_abs >= 5000  OR  delta_pct >= 5%"); r+=1
    _f(r,"Down", "delta_abs <= -5000  OR  delta_pct <= -5%"); r+=1
    _f(r,"Stay", "selain Up dan Down",
       "berlaku untuk: price_tag, cogs_tag, comp_tag"); r+=2

    _s2(r,"E. Status Labels"); r+=1
    _f(r,"price_status", "Up=Price Increase | Stay=Price Stable | Down=Price Reduction"); r+=1
    _f(r,"cogs_status",  "Up=Cost Pressure | Stay=Cost Stable | Down=Cost Improvement"); r+=1
    _f(r,"comp_status",  "Up=Competitor Retreat | Stay=Market Stable | Down=Competitor Aggressive"); r+=2

    _s2(r,"F. Framework Check (current period)"); r+=1
    _f(r,"Cond 1","Fresh  AND  PI_current > 110  AND  margin_pct_current <= 15%"); r+=1
    _f(r,"Cond 2","Frozen AND  PI_current > 100  AND  margin_pct_current <= 15%"); r+=1
    _f(r,"Cond 3","Fresh  AND  PI_current > 120  AND  margin_pct_current >= 70%"); r+=1
    _f(r,"Cond 4","Dry    AND  PI_current < 105  AND  margin_pct_current <= 0%"); r+=1
    _f(r,"Cond 5","Dry    AND  PI_current > 120  AND  margin_pct_current > 40%"); r+=1
    _f(r,"Result","TRUE if any condition met, blank otherwise"); r+=2

    _s2(r,"G. PI Grouping"); r+=1
    for lbl,rng in [("A.<95","PI < 95"),("B.95-<100","95 <= PI < 100"),
                    ("C.100-105","100 <= PI < 105"),("D.105-110","105 <= PI < 110"),
                    ("E.110-120","110 <= PI < 120"),("F.>120","PI >= 120")]:
        _f(r,lbl,rng); r+=1
    _f(r,"Applies to","pi_group_prev, pi_group_cur"); r+=2

    _s2(r,"H. COGS Index Grouping"); r+=1
    for lbl,rng in [("A.<70","CI < 70"),("B.70-85","70 <= CI < 85"),
                    ("C.85-95","85 <= CI < 95"),("D.95-105","95 <= CI < 105"),
                    ("E.>105","CI >= 105  →  Structural Loss if Competitive")]:
        _f(r,lbl,rng); r+=1
    r+=1

    _s2(r,"I. Margin Grouping"); r+=1
    for lbl,rng in [("A.<-20%","< -20%"),("B.-20to-10%","-20% to -10%"),
                    ("C.-10to0%","-10% to 0%"),("D.0to10%","0% to 10%"),
                    ("E.10to20%","10% to 20%"),("F.20to30%","20% to 30%"),
                    ("G.30to50%","30% to 50%"),("H.>50%",">= 50%")]:
        _f(r,lbl,rng); r+=1
    r+=1

    _s2(r,"J. Effect per Existing SKU (Shapley Value)"); r+=1
    _f(r,"PI Mid B",       "next_price / comp_prev x 100",
       "Titik tengah: harga kita T1, comp masih T0"); r+=1
    _f(r,"PI Mid D",       "price_prev / comp_cur x 100",
       "Titik tengah: comp T1, harga kita masih T0"); r+=1
    _f(r,"Effect Price",   "((PI_mid_B - PI_prev) + (PI_cur - PI_mid_D)) / 2",
       "Shapley: rata-rata dua urutan midpoint — zero residual, tidak ada order bias"); r+=1
    _f(r,"Effect Comp (Total)", "((PI_cur - PI_mid_B) + (PI_mid_D - PI_prev)) / 2",
       "Shapley: rata-rata dua urutan midpoint — Ep + Ec = ΔPI selalu exact"); r+=1
    _f(r,"Δcomp",          "next_comp_price - comp_price",
       "Total perubahan comp price (effective/discounted)"); r+=1
    _f(r,"Δnormal_comp",   "next_normal_comp_price - normal_comp_price"); r+=1
    _f(r,"Δdiscount_comp", "(next_normal_comp - next_comp) - (normal_comp - comp)",
       "Perubahan besar diskon kompetitor antar periode"); r+=1
    _f(r,"Effect Normal Comp",   "Effect Comp × (Δnormal_comp / Δcomp)",
       "Jika Δcomp=0 maka =0. Ec_normal + Ec_discount = Ec_total exact"); r+=1
    _f(r,"Effect Discount Comp", "Effect Comp × (Δdiscount_comp / Δcomp)",
       "Bagian comp effect yang disebabkan perubahan agresivitas diskon kompetitor"); r+=2

    _s2(r,"K. Waterfall Avg PI"); r+=1
    _f(r,"A","Avg(PI_prev)     [existing + departing]"); r+=1
    _f(r,"B","Avg(PI_prev)     [existing only]"); r+=1
    _f(r,"C","Avg(PI_current)  [existing only]"); r+=1
    _f(r,"D","Avg(PI_current)  [new SKU only]"); r+=1
    _f(r,"E","(n_ex x C + n_new x D) / (n_ex + n_new)"); r+=1
    _f(r,"eff_dep","B - A"); r+=1
    _f(r,"eff_price","Avg(effect_price) [existing only] — Shapley"); r+=1
    _f(r,"eff_comp (Total)", "Avg(effect_comp)  [existing only] — Shapley"); r+=1
    _f(r,"eff_normal_comp",  "Avg(eff_comp × Δnormal_comp/Δcomp) [existing only]",
       "Sub-effect: perubahan harga normal kompetitor"); r+=1
    _f(r,"eff_discount_comp","Avg(eff_comp × Δdiscount_comp/Δcomp) [existing only]",
       "Sub-effect: perubahan agresivitas diskon kompetitor"); r+=1
    _f(r,"eff_new",  "E - C"); r+=2

    _s2(r,"L. Kontribusi Exact ke Overall"); r+=1
    _f(r,"price/comp","sum_effect_seg / n_ex_total",
       "Exact karena sum of averages = average of all"); r+=1
    _f(r,"departing",
       "sum_pi_ex_seg/n_ex_total - sum_pi_cur_seg/n_cur_total",
       "Fix dari bug lama: dua pool berbeda harus dihitung terpisah"); r+=1
    _f(r,"new SKU",
       "(sum_npi_ex_seg + sum_npi_new_seg)/n_next_total - sum_npi_ex_seg/n_ex_total",
       "Fix dari bug lama: E dan C punya denominator berbeda"); r+=2

    _s2(r,"M. SKU Count & Weight per Pricing BL 25"); r+=1
    hdr(ws,r,["Segment","n Existing","n Departing","n New","n Current","n Next",
              "w_existing","w_current","w_next"],h=20); r+=1

    for seg in segs:
        res=sr[seg]
        vals=[seg,res["n_ex"],res["n_dep"],res["n_new"],res["n_cur"],res["n_next"],
              res["n_ex"]/ov["n_ex"] if ov["n_ex"]>0 else 0,
              res["n_cur"]/ov["n_cur"] if ov["n_cur"]>0 else 0,
              res["n_ex"]/ov["n_next"] if ov["n_next"]>0 else 0]
        for ci,v in enumerate(vals,1):
            fmt="0.0%" if ci>6 else ("#,##0" if ci>1 else "General")
            sc(ws,r,ci,v,align="right" if ci>1 else "left",fmt=fmt,ind=1 if ci==1 else 0,fs=9)
        ws.row_dimensions[r].height=16; r+=1

    vals=["OVERALL",ov["n_ex"],ov["n_dep"],ov["n_new"],ov["n_cur"],ov["n_next"],1.0,1.0,1.0]
    for ci,v in enumerate(vals,1):
        fmt="0.0%" if ci>6 else ("#,##0" if ci>1 else "General")
        sc(ws,r,ci,v,bold=True,bg=NAVY,fc=WHITE,
           align="right" if ci>1 else "left",fmt=fmt,fs=9)
    ws.row_dimensions[r].height=16

    cw(ws,{"A":26,"B":44,"C":14,"D":14,"E":14,"F":14})
    print(f"✅  Sheet 12 selesai ({time.time()-t0:.1f}s)")

# ─────────────────────────────────────────────
# CALLABLE ENTRY POINT (refactored from main)
# ─────────────────────────────────────────────
def analyze(df_raw):
    """
    Main entry point for Streamlit integration.

    Args:
        df_raw: pandas DataFrame with required raw columns

    Returns:
        dict with keys:
            - df_enriched: enriched DataFrame (~+30 derived cols)
            - period_type: 'week' or 'month'
            - period_p1, period_p2: period labels (strings)
            - overall: decompose() result for entire dataset
            - segments: dict of decompose() per BL (Dry/Fresh/Frozen)
            - contribs: exact_contributions() dict
            - workbook: openpyxl Workbook (13 sheets) for download
    """
    period_type, period_col, next_col = detect_period(df_raw)

    d = enrich(df_raw)

    p1 = str(d[period_col].dropna().iloc[0])
    p2 = str(d[next_col].dropna().iloc[0])

    # Pre-compute aggregates once
    ov, sr, contribs = precompute(d)

    # Build workbook
    wb = Workbook()
    # Note: build_s1 uses wb.active to rename default sheet to "1. Raw Data"

    build_s1(wb, d, period_col, next_col)
    s1b_addr = build_s1b(wb, ov, sr, contribs)
    build_s2(wb, s1b_addr, ov, sr, contribs)
    build_s3(wb, d)
    build_s4(wb, d)
    build_s5(wb, d)
    build_s6(wb, d)
    build_s7(wb, d)
    build_s8(wb, d)
    build_s9(wb, d)
    build_s10(wb, d)
    build_s11(wb, d)
    build_s12(wb, ov, sr)

    return {
        'df_enriched': d,
        'period_type': period_type,
        'period_col': period_col,
        'next_col': next_col,
        'period_p1': p1,
        'period_p2': p2,
        'overall': ov,
        'segments': sr,
        'contribs': contribs,
        'workbook': wb,
    }


# ─────────────────────────────────────────────
# CLI ENTRY (only when run directly)
# ─────────────────────────────────────────────
def main():
    t_start = time.time()
    filepath = select_file()
    df_raw   = load_data(filepath)
    result = analyze(df_raw)
    d = result['df_enriched']
    out_name = f"PI_Analysis_{result['period_type']}_{result['period_p1']}_vs_{result['period_p2']}.xlsx"
    out_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), out_name)
    print(f"\n💾  Menyimpan {out_name} ...")
    t_save = time.time()
    result['workbook'].save(out_path)
    print(f"✅  Tersimpan ({time.time()-t_save:.1f}s)")
    print(f"\n🎉  Selesai! Total waktu: {time.time()-t_start:.1f}s")
    print(f"📄  Output: {out_name}")

if __name__ == "__main__":
    main()
